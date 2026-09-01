"""İç piyasa sevk verilerinden müşteri master datası ve rota/bölge önerisi üretir.

Kullanım:
    python -m scripts.ic_piyasa_analiz <sevk_dosyasi.xlsx> [ek_dosya.xlsx ...]

Üretilen dosya (veri/ornek/ic_piyasa_masterdata.xlsx):
  * **Müşteriler**  — bayi, il, ilçe, adres, incoterms, araç tipi geçmişi, tır girişi
  * **Rota Önerisi** — il bazında birlikte rotalanan iller ve önerilen bölge
  * **Bölgeler**    — önerilen bölge listesi ve içindeki iller
  * **Özet**        — belge kodu bazında plan istatistikleri
"""
from __future__ import annotations

import re
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel import sayfa_yaz, yeni_kitap

INCOTERMS = {"CIF", "EXW", "FOB", "DAP", "FCA", "DDP", "CPT"}
TIR_ARAC_TIPLERI = {"TIR", "TIR "}
BELGE_DESENI = re.compile(r"^(\d{2})(\d{2})([A-ZÇŞİĞÜÖ]+)(\d+)$")

# Kolon yerleşimi (sevk dosyalarında sabit)
ARAC_TIPI, SEHIR, SIPARIS_NO, BELGE_NO, DEPO = 3, 4, 5, 6, 7
STOK_KODU, ADET, BAYI, ALICI, ADRES, NOT_ALANI = 8, 10, 11, 12, 13, 14
TARIH, TELEFON, TESLIMAT, PLAN_TARIHI, DAGITIM, DESI = 15, 16, 17, 18, 19, 20
KAMYON_ANAHTAR, TIR_ANAHTAR = 1, 2


_TR_ASCII = str.maketrans("ıİşŞğĞüÜöÖçÇâÂîÎûÛ", "iIsSgGuUoOcCaAiIuU")


def metin(deger) -> str:
    return str(deger).strip() if deger is not None else ""


IL_ESANLAMLILARI = {
    "AFYON": "AFYONKARAHISAR",
    "K.MARAS": "KAHRAMANMARAS",
    "KMARAS": "KAHRAMANMARAS",
    "URFA": "SANLIURFA",
    "ICEL": "MERSIN",
}


def yer_adi(deger) -> str:
    """İl/ilçe adını tek biçime indirger.

    Kaynak veride aynı il hem 'ISTANBUL' hem 'İSTANBUL' olarak geçiyor; Türkçe
    karakterler ASCII karşılığına çevrilip büyük harfe alınmazsa aynı il iki ayrı
    il gibi sayılıyor ve bölgeler bölünüyor.
    """
    ad = metin(deger).translate(_TR_ASCII).upper()
    return IL_ESANLAMLILARI.get(ad, ad)


def not_alanini_coz(deger) -> tuple[str, str]:
    """'CIF', ' - MERKEZ' ya da 'CIF - MERKEZ' biçimini (incoterms, ilçe) olarak ayırır."""
    ham = metin(deger)
    if not ham:
        return "", ""
    if " - " in ham or ham.startswith("-") or ham.endswith("-"):
        sol, _, sag = ham.partition("-")
        sol, sag = sol.strip().upper(), sag.strip().upper()
        return (sol if sol in INCOTERMS else ""), sag.translate(_TR_ASCII)
    buyuk = ham.upper()
    if buyuk in INCOTERMS:
        return buyuk, ""
    return "", buyuk.translate(_TR_ASCII)


def belge_kodu(belge_no: str) -> str:
    eslesme = BELGE_DESENI.match(belge_no.upper())
    return eslesme.group(3) if eslesme else "?"


def satirlari_oku(dosyalar: list[Path]) -> list[tuple]:
    satirlar: list[tuple] = []
    for dosya in dosyalar:
        kitap = load_workbook(dosya, data_only=True, read_only=True)
        for sayfa in kitap.worksheets:
            satir_iter = sayfa.iter_rows(values_only=True)
            basliklar = next(satir_iter, None)
            if not basliklar or len(basliklar) <= DESI:
                continue
            if metin(basliklar[BELGE_NO]).lower() != "belge no":
                continue
            satirlar.extend(r for r in satir_iter if r[BELGE_NO])
        kitap.close()
    return satirlar


IC_PIYASA_KODLARI = {"S", "R", "K", "A"}
TIR_KARAR_ESIGI = 5
"""Bu kadar plana rağmen hiç tır yapılmamışsa müşteriye tır giremiyor kabul edilir."""


def musterileri_cikar(satirlar: list[tuple]) -> list[list]:
    """İç piyasa satırlarından müşteri master datası çıkarır.

    Ring (D) satırları dışarıda bırakılır: o dosyalarda AliciFirma sütununda adres,
    SevkAdresi sütununda ilçe geliyor; karıştırılırsa müşteri bilgisi bozulur.
    """
    musteriler: dict[str, dict] = {}
    for r in satirlar:
        if belge_kodu(metin(r[BELGE_NO])) not in IC_PIYASA_KODLARI:
            continue
        bayi = metin(r[BAYI]).upper()
        if not bayi:
            continue
        incoterms, ilce = not_alanini_coz(r[NOT_ALANI])
        kayit = musteriler.setdefault(
            bayi,
            {
                "alici": Counter(), "il": Counter(), "ilce": Counter(), "adres": Counter(),
                "telefon": Counter(), "incoterms": Counter(), "arac": Counter(),
                "belge": Counter(), "planlar": set(), "desi": 0.0, "adet": 0.0,
                "son": None,
            },
        )
        kayit["alici"][metin(r[ALICI])] += 1
        kayit["il"][yer_adi(r[SEHIR])] += 1
        if ilce:
            kayit["ilce"][ilce] += 1
        kayit["adres"][metin(r[ADRES])] += 1
        if metin(r[TELEFON]):
            kayit["telefon"][metin(r[TELEFON])] += 1
        if incoterms:
            kayit["incoterms"][incoterms] += 1
        kayit["arac"][metin(r[ARAC_TIPI]).upper()] += 1
        kayit["belge"][belge_kodu(metin(r[BELGE_NO]))] += 1
        kayit["planlar"].add(metin(r[BELGE_NO]))
        if isinstance(r[DESI], (int, float)):
            kayit["desi"] += r[DESI]
        if isinstance(r[ADET], (int, float)):
            kayit["adet"] += r[ADET]
        if r[PLAN_TARIHI] and (kayit["son"] is None or r[PLAN_TARIHI] > kayit["son"]):
            kayit["son"] = r[PLAN_TARIHI]

    def ilk(sayac: Counter) -> str:
        return sayac.most_common(1)[0][0] if sayac else ""

    satirlar_cikti = []
    for bayi, k in sorted(musteriler.items()):
        tir_sayisi = sum(adet for arac, adet in k["arac"].items() if arac in TIR_ARAC_TIPLERI)
        plan_sayisi = len(k["planlar"])
        if tir_sayisi:
            tir_girisi = "E"
        elif plan_sayisi >= TIR_KARAR_ESIGI:
            tir_girisi = "H"
        else:
            # Geçmişi az olan müşteri için "tır giremez" demek doğru olmaz.
            tir_girisi = "?"
        satirlar_cikti.append([
            bayi,
            ilk(k["alici"]),
            ilk(k["il"]),
            ilk(k["ilce"]),
            ilk(k["adres"]),
            ilk(k["telefon"]),
            ilk(k["incoterms"]) or "CIF",
            tir_girisi,
            len(k["planlar"]),
            k["belge"].get("S", 0),
            k["belge"].get("R", 0),
            k["belge"].get("K", 0),
            k["belge"].get("D", 0),
            round(k["adet"]),
            round(k["desi"]),
            k["son"].strftime("%d.%m.%Y") if k["son"] else "",
            ", ".join(f"{a}:{n}" for a, n in k["arac"].most_common(4)),
        ])
    return satirlar_cikti


def rotalari_cikar(satirlar: list[tuple]) -> tuple[list[list], list[list]]:
    planlar: dict[str, list[tuple]] = defaultdict(list)
    for r in satirlar:
        if belge_kodu(metin(r[BELGE_NO])) == "S":
            planlar[metin(r[BELGE_NO])].append(r)

    birlikte: Counter = Counter()
    il_plan: Counter = Counter()
    for satirlar_plan in planlar.values():
        iller = sorted({yer_adi(r[SEHIR]) for r in satirlar_plan if r[SEHIR]})
        for il in iller:
            il_plan[il] += 1
        for i, a in enumerate(iller):
            for b in iller[i + 1:]:
                birlikte[(a, b)] += 1

    # Güçlü bağ: en az 10 planda birlikte ve küçük ilin planlarının en az %25'i.
    baglar: dict[str, set[str]] = defaultdict(set)
    for (a, b), adet in birlikte.items():
        if adet >= 8 and adet >= 0.20 * min(il_plan[a], il_plan[b]):
            baglar[a].add(b)
            baglar[b].add(a)

    bolge_no: dict[str, int] = {}
    bolgeler: list[list[str]] = []
    for il in sorted(il_plan, key=lambda x: -il_plan[x]):
        if il in bolge_no:
            continue
        yigin, kume = [il], []
        while yigin:
            mevcut = yigin.pop()
            if mevcut in bolge_no:
                continue
            bolge_no[mevcut] = len(bolgeler) + 1
            kume.append(mevcut)
            yigin.extend(k for k in baglar[mevcut] if k not in bolge_no)
        bolgeler.append(sorted(kume))

    rota_satirlari = []
    for il in sorted(il_plan, key=lambda x: -il_plan[x]):
        komsular = sorted(
            ((b if a == il else a, n) for (a, b), n in birlikte.items() if il in (a, b)),
            key=lambda t: -t[1],
        )[:6]
        rota_satirlari.append([
            il,
            il_plan[il],
            bolge_no.get(il, 0),
            ", ".join(f"{ad} ({n})" for ad, n in komsular),
        ])

    bolge_satirlari = [
        [no, len(iller), ", ".join(iller)] for no, iller in enumerate(bolgeler, start=1)
    ]
    return rota_satirlari, bolge_satirlari


def ozeti_cikar(satirlar: list[tuple]) -> list[list]:
    planlar: dict[str, list[tuple]] = defaultdict(list)
    for r in satirlar:
        planlar[metin(r[BELGE_NO])].append(r)

    gruplar: dict[str, list] = defaultdict(list)
    for belge_no, satirlar_plan in planlar.items():
        gruplar[belge_kodu(belge_no)].append(satirlar_plan)

    cikti = []
    for kod, plan_listesi in sorted(gruplar.items(), key=lambda t: -len(t[1])):
        if len(plan_listesi) < 20:
            continue
        tir = [
            sum(r[TIR_ANAHTAR] for r in p if isinstance(r[TIR_ANAHTAR], (int, float)))
            for p in plan_listesi
        ]
        tir = [x for x in tir if x > 0]
        duraklar = [
            len({(yer_adi(r[SEHIR]), not_alanini_coz(r[NOT_ALANI])[1], metin(r[BAYI])) for r in p})
            for p in plan_listesi
        ]
        cikti.append([
            kod,
            len(plan_listesi),
            round(statistics.mean(duraklar), 2),
            round(statistics.median(tir), 3) if tir else 0,
            round(statistics.mean([len({yer_adi(r[SEHIR]) for r in p}) for p in plan_listesi]), 2),
            round(statistics.mean([len({metin(r[BAYI]) for r in p}) for p in plan_listesi]), 2),
        ])
    return cikti


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    dosyalar = [Path(a) for a in sys.argv[1:]]
    satirlar = satirlari_oku(dosyalar)
    print(f"{len(satirlar)} satır okundu")

    musteriler = musterileri_cikar(satirlar)
    rotalar, bolgeler = rotalari_cikar(satirlar)
    ozet = ozeti_cikar(satirlar)

    kitap = yeni_kitap()
    sayfa_yaz(
        kitap.create_sheet("Müşteriler"),
        ["Bayi Adı", "Alıcı Firma", "İl", "İlçe", "Sevk Adresi", "Telefon", "Incoterms",
         "Tır Girişi (E/H/?)", "Plan Sayısı", "FTL (S)", "Rutin (R)", "Kargo (K)",
         "Ring (D)", "Toplam Adet", "Toplam Desi", "Son Sevk", "Araç Tipi Geçmişi"],
        musteriler,
        [40, 40, 16, 18, 44, 16, 12, 11, 12, 10, 10, 10, 10, 13, 13, 13, 34],
    )
    sayfa_yaz(
        kitap.create_sheet("Rota Önerisi"),
        ["İl", "FTL Plan Sayısı", "Önerilen Bölge", "Birlikte Rotalandığı İller (plan sayısı)"],
        rotalar,
        [18, 16, 16, 90],
    )
    sayfa_yaz(
        kitap.create_sheet("Bölgeler"),
        ["Bölge", "İl Sayısı", "İller"],
        bolgeler,
        [10, 12, 120],
    )
    sayfa_yaz(
        kitap.create_sheet("Özet"),
        ["Belge Kodu", "Plan Sayısı", "Ort. Durak", "Medyan Tır Doluluğu",
         "Ort. İl", "Ort. Bayi"],
        ozet,
        [14, 14, 14, 22, 12, 12],
    )

    hedef = Path("veri/ornek/ic_piyasa_masterdata.xlsx")
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    print(
        f"{hedef} yazıldı · {len(musteriler)} müşteri · {len(rotalar)} il · "
        f"{len(bolgeler)} bölge önerisi"
    )


if __name__ == "__main__":
    main()
