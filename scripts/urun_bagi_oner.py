"""Birlikte sevk edilmesi gereken ürün çiftlerini önerir.

Kullanım:
    # Yalnızca ürün adlarından (klima iç/dış ünite çiftleri)
    python -m scripts.urun_bagi_oner

    # Sevk/sipariş geçmişinden de çıkarım yaparak
    python -m scripts.urun_bagi_oner <sevk_dosyasi.xlsx> [ek_dosya.xlsx ...]

Üretilen dosya (veri/ornek/urun_bagi_onerileri.xlsx) doğrudan
**Master Data > Ürün Bağları > Yükle** ekranından geri yüklenebilir; başlıklar içe
aktarımın tanıdığı başlıklardır. Öneriler **onaylanmadan** sisteme yazılmaz: dosya
gözden geçirilip yanlış satırlar silindikten sonra yüklenir.

İki kaynak var:

1. **Ürün adı** — klimada iç ve dış ünite aynı gövdeyi paylaşıp yalnızca
   "İç"/"Dış" ile ayrılıyor (ör. "A5 Inverter 18 İç" ↔ "A5 Inverter 18 Dış").
   Bu çiftler SET bağıdır: hiçbiri tek başına gitmemeli.

2. **Sevk geçmişi** — bir aksesuarın (BACA, AKSESUAR, DİRSEK grubu) hangi ana
   ürünle birlikte sipariş edildiği. Ölçüt: aksesuar o ana ürünle **birlikte
   görüldüğü siparişlerin ezici çoğunluğunda** yalnız değil. Rastlantısal
   birlikteliği elemek için hem asgari birliktelik sayısı hem asgari oran aranır.
"""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel import sayfa_yaz, yeni_kitap
from app.services.veri_formatlari import URUN_BAGI_ALANLARI

URUN_MASTERDATA = Path("veri/ornek/urun_masterdata.xlsx")
HEDEF = Path("veri/ornek/urun_bagi_onerileri.xlsx")

AKSESUAR_GRUPLARI = {"AKSESUAR", "BACA", "DİRSEK", "DIRSEK"}

ASGARI_BIRLIKTELIK = 5
"""Bir çiftin önerilmesi için en az kaç siparişte birlikte görülmesi gerekir.

Daha düşük bir eşik tek seferlik rastlantıları öneri hâline getiriyor.
"""
ASGARI_ORAN = 0.60
"""Aksesuarın göründüğü siparişlerin en az bu kadarında o ana ürün de olmalı."""


def _metin(deger) -> str:
    return str(deger).strip() if deger is not None else ""


# --------------------------------------------------------------- ürün master datası
def urunleri_oku() -> dict[str, dict]:
    if not URUN_MASTERDATA.exists():
        raise SystemExit(f"Ürün master datası bulunamadı: {URUN_MASTERDATA}")
    sayfa = load_workbook(URUN_MASTERDATA, read_only=True)["Ürünler"]
    urunler: dict[str, dict] = {}
    for satir in sayfa.iter_rows(min_row=2, values_only=True):
        kod = _metin(satir[0])
        if kod:
            urunler[kod] = {"ad": _metin(satir[1]), "grup": _metin(satir[2]).upper()}
    return urunler


# ------------------------------------------------------------------ 1. ürün adından
_IC_DIS = re.compile(r"\s*\b(İç|Iç|Ic|Dış|Dis)\b\s*", re.IGNORECASE)


def _yon(ad: str) -> str | None:
    bulunan = _IC_DIS.search(ad)
    if not bulunan:
        return None
    return "IC" if bulunan.group(1).upper() in {"İÇ", "IÇ", "IC"} else "DIS"


def addan_set_onerileri(urunler: dict[str, dict]) -> list[dict]:
    """Aynı gövdeyi paylaşan iç/dış ünite çiftleri."""
    govdeler: dict[str, dict[str, list[str]]] = defaultdict(lambda: {"IC": [], "DIS": []})
    for kod, bilgi in urunler.items():
        if bilgi["grup"] not in {"KLİMA", "KLIMA", "VRF", "ISI POMPASI"}:
            continue
        yon = _yon(bilgi["ad"])
        if yon:
            govde = _IC_DIS.sub(" ", bilgi["ad"]).strip().upper()
            govdeler[govde][yon].append(kod)

    oneriler = []
    for govde, taraflar in sorted(govdeler.items()):
        # Yalnızca birebir eşleşmeler önerilir. Bir gövdede birden çok iç ya da dış
        # ünite varsa hangisinin hangisiyle gittiği addan anlaşılmaz; elle karar
        # verilmeli, yanlış öneri doğru olanı da şüpheli hâle getirir.
        if len(taraflar["IC"]) != 1 or len(taraflar["DIS"]) != 1:
            continue
        oneriler.append({
            "ana_urun_kodu": taraflar["DIS"][0],
            "bagli_urun_kodu": taraflar["IC"][0],
            "tip": "SET",
            "aciklama": f"Ürün adından: {govde} iç/dış ünite çifti",
        })
    return oneriler


# -------------------------------------------------------------- 2. sevk geçmişinden
# Sevk dosyalarında kolon yerleşimi sabit (bkz. scripts/ic_piyasa_analiz.py).
SIPARIS_NO, STOK_KODU, BAYI = 5, 8, 11


def siparisleri_oku(dosyalar: list[Path]) -> dict[str, set[str]]:
    """Sipariş numarası -> o siparişte geçen ürün kodları."""
    siparisler: dict[str, set[str]] = defaultdict(set)
    for dosya in dosyalar:
        sayfa = load_workbook(dosya, read_only=True, data_only=True).worksheets[0]
        for satir in sayfa.iter_rows(min_row=2, values_only=True):
            if len(satir) <= STOK_KODU:
                continue
            siparis = _metin(satir[SIPARIS_NO]) or _metin(satir[BAYI])
            kod = _metin(satir[STOK_KODU])
            if siparis and kod:
                siparisler[siparis].add(kod)
    return siparisler


def gecmisten_aksesuar_onerileri(
    siparisler: dict[str, set[str]], urunler: dict[str, dict]
) -> list[dict]:
    """Aksesuarın hangi ana ürünle birlikte sipariş edildiğini bulur."""
    aksesuar_toplam: Counter[str] = Counter()
    birliktelik: Counter[tuple[str, str]] = Counter()

    for kodlar in siparisler.values():
        aksesuarlar = [
            k for k in kodlar if urunler.get(k, {}).get("grup") in AKSESUAR_GRUPLARI
        ]
        anacillar = [
            k for k in kodlar
            if k in urunler and urunler[k]["grup"] not in AKSESUAR_GRUPLARI
        ]
        for aksesuar in aksesuarlar:
            aksesuar_toplam[aksesuar] += 1
            for ana in anacillar:
                birliktelik[(ana, aksesuar)] += 1

    oneriler = []
    for (ana, aksesuar), adet in birliktelik.most_common():
        toplam = aksesuar_toplam[aksesuar]
        oran = adet / toplam if toplam else 0
        if adet < ASGARI_BIRLIKTELIK or oran < ASGARI_ORAN:
            continue
        oneriler.append({
            "ana_urun_kodu": ana,
            "bagli_urun_kodu": aksesuar,
            "tip": "AKSESUAR",
            "aciklama": (
                f"Geçmişten: {adet} siparişte birlikte "
                f"(aksesuarın siparişlerinin %{oran * 100:.0f}'i)"
            ),
        })
    return oneriler


def main() -> None:
    urunler = urunleri_oku()
    oneriler = addan_set_onerileri(urunler)
    print(f"Ürün adından set önerisi: {len(oneriler)}")

    dosyalar = [Path(a) for a in sys.argv[1:]]
    eksik = [d for d in dosyalar if not d.exists()]
    if eksik:
        raise SystemExit("Bulunamayan dosya: " + ", ".join(str(d) for d in eksik))
    if dosyalar:
        siparisler = siparisleri_oku(dosyalar)
        gecmis = gecmisten_aksesuar_onerileri(siparisler, urunler)
        print(f"Sipariş sayısı: {len(siparisler)}")
        print(f"Geçmişten aksesuar önerisi: {len(gecmis)}")
        oneriler.extend(gecmis)
    else:
        print("Sevk dosyası verilmedi; yalnızca ürün adı kuralı çalıştı.")

    basliklar = [alan.baslik for alan in URUN_BAGI_ALANLARI]
    satirlar = [[oneri[alan.ad] for alan in URUN_BAGI_ALANLARI] for oneri in oneriler]
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("Ürün Bağları")
    sayfa_yaz(sayfa, basliklar, satirlar)
    sayfa.auto_filter.ref = sayfa.dimensions
    HEDEF.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(HEDEF)
    print(f"\n{len(oneriler)} öneri yazıldı: {HEDEF}")
    print("Gözden geçirip Master Data > Ürün Bağları > Yükle ekranından yükleyin.")


if __name__ == "__main__":
    main()
