"""İhracat sevk verilerinden müşteri master datası ve kapasite ölçüleri üretir.

Kullanım:
    python -m scripts.ihracat_analiz <ihracat.xlsx> [ek_dosya.xlsx ...] \
        [--formuller <yukleme_formu.xlsx>]

`--formuller` ile verilen yükleme formu kitabının **FORMÜLLER** sayfasından müşteri
bazlı yükleme tipi, maksimum tonaj ve araç tipi okunur; verilmezse bu alanlar geçmiş
sevklerden çıkarılır.

Üretilen dosya (veri/ornek/ihracat_masterdata.xlsx):
  * **Müşteriler** — müşteri, ülke, araç tipi, taşıma modu, sefer kodu, yükleme tipi,
    maksimum tonaj, geçmiş plan sayısı ve ortalama doluluk
  * **Ülkeler**   — ülke kodu, ülke, incoterms, taşıma modu
  * **Özet**      — araç tipi bazında plan sayısı ve kapasite yüzdelikleri
"""
from __future__ import annotations

import collections
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from app.services.excel import sayfa_yaz, yeni_kitap

# Kolon yerleşimi yıla göre değişiyor (2023 dosyasında KG yok, ARAÇ TİPİ iki sütun
# solda), bu yüzden sabit indis kullanılmıyor: her sayfanın başlığı okunup eşleniyor.
ALAN_BASLIKLARI = {
    "depo": ("DEPO",),
    "axata": ("AXATA",),
    "ulke_kodu": ("ÜLKE KODU", "ULKE KODU"),
    "siparis_no": ("SİPARİŞ NO", "SIPARIS NO"),
    "urun_kodu": ("ÜRÜN KODU", "URUN KODU"),
    "urun_adi": ("ÜRÜN TANIMI", "URUN TANIMI"),
    "adet": ("ADET",),
    "musteri": ("MÜŞTERİ ADI", "MUSTERI ADI"),
    "sevk_adresi": ("SEVK ADRESİ", "SEVK ADRESI"),
    "teslimat_no": ("TESLİMAT NO", "TESLIMAT NO"),
    "desi": ("DESİ", "DESI"),
    "kg": ("KG", "AĞIRLIK", "AGIRLIK"),
    "belge_no": ("BELGE NUMARASI", "BELGE NO"),
    "plaka": ("ARAÇ PLAKASI", "ARAC PLAKASI"),
    "konteyner_no": ("KONTEYNER NUMARASI",),
    "arac_tipi": ("ARAÇ TİPİ", "ARAC TIPI"),
    "nsc": ("NSC&EXPORT", "NSC&CORE"),
    "urun_grubu": ("ÜRÜN GRUBU", "URUN GRUBU"),
    "ulke": ("ÜLKE",),
    "incoterms": ("INCOTERMS", "İNCOTERMS"),
}


def _baslik_haritasi(basliklar: tuple) -> dict[str, int] | None:
    """Sayfa başlığını alan adlarına eşler; ihracat sayfası değilse None döner."""
    yerler = {
        str(b).strip().upper(): i for i, b in enumerate(basliklar) if b is not None
    }
    harita = {
        alan: next((yerler[a] for a in adaylar if a in yerler), None)
        for alan, adaylar in ALAN_BASLIKLARI.items()
    }
    if harita["musteri"] is None or harita["belge_no"] is None:
        return None
    return harita

BELGE_DESENI = re.compile(r"^(\d{2})(\d{2})([A-Z])(\d+)$")

DENIZ_ARAC_TIPLERI = ("KONTEYNER", "40", "20", "HC", "DC")
"""Araç tipi metninde bunlardan biri geçiyorsa sevkiyat deniz yoluyladır."""


def metin(deger) -> str:
    return str(deger).strip() if deger is not None else ""


def arac_tipi(ham: str) -> str:
    """Ham araç tipi metnini KONTEYNER / TIR / PARSİYEL / KARGO'ya indirger."""
    buyuk = metin(ham).upper()
    if any(isaret in buyuk for isaret in DENIZ_ARAC_TIPLERI):
        return "KONTEYNER"
    if "TIR" in buyuk:
        return "TIR"
    if "PARS" in buyuk:
        return "PARSİYEL"
    if buyuk in {"DHL", "KARGO", "UPS"}:
        return "KARGO"
    return buyuk or "TIR"


def tasima_modu(tip: str) -> str:
    return "DENİZ" if tip == "KONTEYNER" else "KARA"


def sefer_kodu(nsc: str) -> str:
    """NSC&Export alanı sefer numarasının belge kodunu belirler.

    Geçmiş veride birebir örtüşüyor: `NSC` -> `N` (2412N4018), `Export` -> `E`
    (2608E4001). Yeni bir müşteri için alan boşsa E kabul edilir.
    """
    return "N" if metin(nsc).upper().startswith("NSC") else "E"


def sayi(deger) -> float:
    return float(deger) if isinstance(deger, (int, float)) else 0.0


def satirlari_oku(dosyalar: list[Path]) -> list[dict]:
    """Bütün dosyaları tek bir alan adı sözlüğü listesine indirger."""
    satirlar: list[dict] = []
    for dosya in dosyalar:
        kitap = load_workbook(dosya, data_only=True, read_only=True)
        for sayfa in kitap.worksheets:
            satir_iter = sayfa.iter_rows(values_only=True)
            basliklar = next(satir_iter, None)
            harita = _baslik_haritasi(basliklar) if basliklar else None
            if harita is None:
                continue
            for r in satir_iter:
                if not r or harita["belge_no"] >= len(r) or not r[harita["belge_no"]]:
                    continue
                satirlar.append(
                    {
                        alan: (r[yer] if yer is not None and yer < len(r) else None)
                        for alan, yer in harita.items()
                    }
                )
    return satirlar


def formulleri_oku(dosya: Path) -> dict[str, dict]:
    """Yükleme formu kitabının FORMÜLLER sayfasındaki müşteri kurallarını okur."""
    kitap = load_workbook(dosya, data_only=True, read_only=True)
    if "FORMÜLLER" not in kitap.sheetnames:
        return {}
    sayfa = kitap["FORMÜLLER"]
    satir_iter = sayfa.iter_rows(values_only=True)
    next(satir_iter, None)
    kurallar: dict[str, dict] = {}
    for r in satir_iter:
        ad = metin(r[0]) if r else ""
        if not ad:
            continue
        kurallar[ad.upper()] = {
            "ulke": metin(r[1]),
            "tedarikci": metin(r[2]),
            "satis_destek": metin(r[3]),
            "yukleme_tipi": metin(r[4]),
            "azami_tonaj": metin(r[5]),
            "arac_tipi": arac_tipi(metin(r[6])),
            "nsc": metin(r[7]),
        }
    return kurallar


def musterileri_cikar(satirlar: list[dict], kurallar: dict[str, dict]) -> list[list]:
    musteriler: dict[str, dict] = {}
    plan_musterileri: dict[str, set[str]] = collections.defaultdict(set)
    plan_olculeri: dict[str, dict] = {}

    for r in satirlar:
        ad = metin(r["musteri"])
        if not ad:
            continue
        belge = metin(r["belge_no"])
        plan_musterileri[belge].add(ad)
        olcu = plan_olculeri.setdefault(
            belge, {"desi": 0.0, "kg": 0.0, "arac": collections.Counter()}
        )
        olcu["desi"] += sayi(r["desi"])
        olcu["kg"] += sayi(r["kg"])
        olcu["arac"][arac_tipi(metin(r["arac_tipi"]))] += 1

        kayit = musteriler.setdefault(
            ad,
            {
                "ulke": collections.Counter(), "ulke_kodu": collections.Counter(),
                "adres": collections.Counter(), "arac": collections.Counter(),
                "nsc": collections.Counter(), "planlar": set(),
                "desi": 0.0, "kg": 0.0, "adet": 0.0,
            },
        )
        kayit["ulke"][metin(r["ulke"])] += 1
        kayit["ulke_kodu"][metin(r["ulke_kodu"])] += 1
        kayit["adres"][metin(r["sevk_adresi"])] += 1
        kayit["arac"][arac_tipi(metin(r["arac_tipi"]))] += 1
        kayit["nsc"][metin(r["nsc"])] += 1
        kayit["planlar"].add(belge)
        kayit["desi"] += sayi(r["desi"])
        kayit["kg"] += sayi(r["kg"])
        kayit["adet"] += sayi(r["adet"])

    def ilk(sayac: collections.Counter) -> str:
        return sayac.most_common(1)[0][0] if sayac else ""

    cikti = []
    for ad, k in sorted(musteriler.items()):
        kural = kurallar.get(ad.upper(), {})
        tip = kural.get("arac_tipi") or ilk(k["arac"]) or "TIR"
        # Tek müşterili planların ortalama doluluğu: araç kapasitesine oranla.
        tek_planlar = [
            plan_olculeri[b] for b in k["planlar"] if len(plan_musterileri[b]) == 1
        ]
        ort_desi = (
            sum(p["desi"] for p in tek_planlar) / len(tek_planlar) if tek_planlar else 0
        )
        ort_kg = (
            sum(p["kg"] for p in tek_planlar) / len(tek_planlar) if tek_planlar else 0
        )
        cikti.append([
            ad,
            kural.get("ulke") or ilk(k["ulke"]),
            ilk(k["ulke_kodu"]),
            ilk(k["adres"]),
            tip,
            tasima_modu(tip),
            sefer_kodu(kural.get("nsc") or ilk(k["nsc"])),
            kural.get("yukleme_tipi", ""),
            kural.get("azami_tonaj", ""),
            kural.get("tedarikci", ""),
            kural.get("satis_destek", ""),
            len(k["planlar"]),
            round(k["adet"]),
            round(ort_desi),
            round(ort_kg),
            "E",
        ])
    return cikti


def ulkeleri_cikar(satirlar: list[dict], ulke_sayfasi: dict[str, str]) -> list[list]:
    ulkeler: dict[str, dict] = {}
    for r in satirlar:
        kod = metin(r["ulke_kodu"])
        if not kod:
            continue
        kayit = ulkeler.setdefault(
            kod, {"ad": collections.Counter(), "arac": collections.Counter()}
        )
        kayit["ad"][metin(r["ulke"])] += 1
        kayit["arac"][arac_tipi(metin(r["arac_tipi"]))] += 1
    cikti = []
    for kod, k in sorted(ulkeler.items()):
        tip = k["arac"].most_common(1)[0][0]
        cikti.append([
            kod,
            k["ad"].most_common(1)[0][0],
            ulke_sayfasi.get(kod, ""),
            tasima_modu(tip),
            tip,
        ])
    return cikti


def ozeti_cikar(satirlar: list[dict]) -> list[list]:
    planlar: dict[str, dict] = {}
    for r in satirlar:
        belge = metin(r["belge_no"])
        p = planlar.setdefault(
            belge, {"desi": 0.0, "kg": 0.0, "arac": collections.Counter(), "musteri": set()}
        )
        p["desi"] += sayi(r["desi"])
        p["kg"] += sayi(r["kg"])
        p["arac"][arac_tipi(metin(r["arac_tipi"]))] += 1
        p["musteri"].add(metin(r["musteri"]))

    gruplar: dict[str, list[dict]] = collections.defaultdict(list)
    for p in planlar.values():
        gruplar[p["arac"].most_common(1)[0][0]].append(p)

    def yuzdelik(degerler: list[float], oran: float) -> float:
        if not degerler:
            return 0.0
        sirali = sorted(degerler)
        return sirali[min(len(sirali) - 1, int(len(sirali) * oran))]

    cikti = []
    for tip, ps in sorted(gruplar.items(), key=lambda x: -len(x[1])):
        desiler = [p["desi"] for p in ps if p["desi"] > 0]
        kglar = [p["kg"] for p in ps if p["kg"] > 0]
        tek = sum(1 for p in ps if len(p["musteri"]) == 1)
        cikti.append([
            tip, len(ps), round(tek * 100 / len(ps), 1),
            round(yuzdelik(desiler, 0.5)), round(yuzdelik(desiler, 0.9)),
            round(max(desiler) if desiler else 0),
            round(yuzdelik(kglar, 0.5)), round(yuzdelik(kglar, 0.9)),
            round(max(kglar) if kglar else 0),
        ])
    return cikti


def main() -> int:
    argumanlar = sys.argv[1:]
    formuller_dosyasi = None
    if "--formuller" in argumanlar:
        indis = argumanlar.index("--formuller")
        formuller_dosyasi = Path(argumanlar[indis + 1])
        argumanlar = argumanlar[:indis] + argumanlar[indis + 2 :]
    if not argumanlar:
        print(__doc__)
        return 1

    dosyalar = [Path(a) for a in argumanlar]
    satirlar = satirlari_oku(dosyalar)
    print(f"{len(satirlar)} satır okundu")

    kurallar = formulleri_oku(formuller_dosyasi) if formuller_dosyasi else {}
    if formuller_dosyasi:
        print(f"{len(kurallar)} müşteri kuralı okundu ({formuller_dosyasi.name})")

    # "Ülke" sayfası varsa incoterms oradan alınır.
    ulke_sayfasi: dict[str, str] = {}
    for dosya in dosyalar:
        kitap = load_workbook(dosya, data_only=True, read_only=True)
        if "Ülke" not in kitap.sheetnames:
            continue
        it = kitap["Ülke"].iter_rows(values_only=True)
        next(it, None)
        for r in it:
            if r and r[0]:
                ulke_sayfasi[metin(r[0])] = metin(r[2])

    musteriler = musterileri_cikar(satirlar, kurallar)
    ulkeler = ulkeleri_cikar(satirlar, ulke_sayfasi)
    ozet = ozeti_cikar(satirlar)

    kitap = yeni_kitap()
    sayfa_yaz(
        kitap.create_sheet("Müşteriler"),
        ["Müşteri Adı", "Ülke", "Ülke Kodu", "Sevk Adresi", "Araç Tipi", "Taşıma Modu",
         "Sefer Kodu", "Yükleme Tipi", "Azami Tonaj", "Tedarikçi", "Satış Destek",
         "Plan Sayısı", "Toplam Adet", "Ort. Desi", "Ort. KG", "Aktif"],
        musteriler,
        [38, 16, 10, 24, 13, 13, 11, 26, 14, 18, 16, 12, 13, 12, 12, 8],
    )
    sayfa_yaz(
        kitap.create_sheet("Ülkeler"),
        ["Ülke Kodu", "Ülke", "Incoterms", "Taşıma Modu", "Araç Tipi"],
        ulkeler,
        [12, 20, 12, 14, 13],
    )
    sayfa_yaz(
        kitap.create_sheet("Özet"),
        ["Araç Tipi", "Plan Sayısı", "Tek Müşterili %", "Desi p50", "Desi p90",
         "Desi max", "KG p50", "KG p90", "KG max"],
        ozet,
        [14, 13, 17, 12, 12, 12, 12, 12, 12],
    )

    hedef = Path("veri/ornek/ihracat_masterdata.xlsx")
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    print(
        f"{hedef} yazıldı · {len(musteriler)} müşteri · {len(ulkeler)} ülke · "
        f"{len(ozet)} araç tipi"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
