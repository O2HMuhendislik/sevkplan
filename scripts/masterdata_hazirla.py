"""Kaynak `masterdata` sayfasından sisteme yüklenmeye hazır master data üretir.

Kullanım:
    python -m scripts.masterdata_hazirla <kaynak.xlsx> [hedef.xlsx]

Yapılanlar:
  * '#N/A' ve benzeri formül hataları boşa çevrilir
  * ürün grubu büyük harfe normalize edilir
  * planlanabilirlik kontrol edilir (palet içi adet / kamyon / tır yükleme adeti)
  * çıktı iki sayfalı olur: "Ürünler" ve eksik veri gerekçeleriyle "Eksik Veri"
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl.styles import Font

from app.services import excel
from app.services.excel import sayfa_yaz, yeni_kitap
from app.services.veri_formatlari import URUN_ALANLARI, URUN_ALIAS, zorunlu_alanlar

CIKTI_KOLONLARI = [alan for alan in URUN_ALANLARI if alan.ad != "aktif"]


def hazirla(kaynak: Path, hedef: Path) -> dict[str, int]:
    kayitlar = excel.satirlari_oku(kaynak, URUN_ALIAS, zorunlu_alanlar(URUN_ALANLARI))

    temiz: list[list] = []
    eksik: list[list] = []
    gruplar: Counter[str] = Counter()
    gorulen: set[str] = set()

    for kayit in kayitlar:
        urun_kodu = excel.metin(kayit.get("urun_kodu"))
        urun_adi = excel.metin(kayit.get("urun_adi"))
        if not urun_kodu or not urun_adi:
            continue
        if urun_kodu in gorulen:
            continue
        gorulen.add(urun_kodu)

        grup = (excel.metin(kayit.get("urun_grubu")) or "").upper() or None
        palet_ici = excel.tam_sayi_ya_da(kayit.get("palet_ici_adet"))
        kamyon = excel.tam_sayi_ya_da(kayit.get("kamyon_yukleme_adeti"))
        tir = excel.tam_sayi_ya_da(kayit.get("tir_yukleme_adeti"))

        satir = [
            urun_kodu, urun_adi, grup, palet_ici, kamyon,
            excel.tam_sayi_ya_da(kayit.get("kamyon_palet")), tir,
            excel.tam_sayi_ya_da(kayit.get("tir_palet")),
            excel.sayi_ya_da(kayit.get("agirlik")),
            excel.sayi_ya_da(kayit.get("desi")),
            excel.sayi_ya_da(kayit.get("m3")),
            excel.tam_sayi_ya_da(kayit.get("palet_en")),
            excel.tam_sayi_ya_da(kayit.get("palet_boy")),
            excel.tam_sayi_ya_da(kayit.get("palet_yukseklik")),
            excel.metin(kayit.get("header_kod")),
        ]
        satir = [float(h) if hasattr(h, "quantize") else h for h in satir]
        temiz.append(satir)
        gruplar[grup or "(grup yok)"] += 1

        gerekceler = []
        if not grup:
            gerekceler.append("ürün grubu boş")
        if not palet_ici:
            gerekceler.append("palet içi adet yok → depo 64 (palet) planlamasına giremez")
        if not tir:
            gerekceler.append("tır yükleme adeti yok → depo 74 (anahtar) planlamasına giremez")
        if not (palet_ici or kamyon or tir):
            gerekceler.append("HİÇBİR kapasite verisi yok → hiç planlanamaz")
        if gerekceler:
            eksik.append([urun_kodu, urun_adi, grup, palet_ici, kamyon, tir,
                          "; ".join(gerekceler)])

    kitap = yeni_kitap()
    sayfa_yaz(
        kitap.create_sheet("Ürünler"),
        [alan.baslik for alan in CIKTI_KOLONLARI],
        temiz,
        [max(16, len(alan.baslik) + 4) for alan in CIKTI_KOLONLARI],
    )
    eksik_sayfa = kitap.create_sheet("Eksik Veri")
    sayfa_yaz(
        eksik_sayfa,
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet",
         "Kamyon yükleme adeti", "Tır yükleme adeti", "Eksik olan"],
        eksik,
        [16, 46, 16, 14, 20, 18, 84],
    )
    ozet = kitap.create_sheet("Özet")
    ozet["A1"] = "Ürün Master Data Özeti"
    ozet["A1"].font = Font(bold=True, size=13)
    ozet.append([])
    ozet.append(["Toplam ürün", len(temiz)])
    ozet.append(["Eksik verisi olan", len(eksik)])
    ozet.append([])
    ozet.append(["Ürün Grubu", "Adet"])
    for grup, adet in gruplar.most_common():
        ozet.append([grup, adet])
    ozet.column_dimensions["A"].width = 26
    ozet.column_dimensions["B"].width = 12

    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return {"urun": len(temiz), "eksik": len(eksik), "grup": len(gruplar)}


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    kaynak = Path(sys.argv[1])
    hedef = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("veri/ornek/urun_masterdata.xlsx")
    sonuc = hazirla(kaynak, hedef)
    print(
        f"{hedef} yazıldı · {sonuc['urun']} ürün · {sonuc['grup']} grup · "
        f"{sonuc['eksik']} üründe eksik veri"
    )


if __name__ == "__main__":
    main()
