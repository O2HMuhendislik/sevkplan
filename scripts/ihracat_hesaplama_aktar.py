"""Şirketin `Hesaplama.xlsx` dosyasını sistemin ihracat master datasına çevirir.

Kullanım:
    python -m scripts.ihracat_hesaplama_aktar [Hesaplama.xlsx]

İki çıktı üretir:

* ``veri/ornek/ihracat_urun_masterdata.xlsx`` — `Ürün` sayfasının birebir karşılığı.
  Doluluk hesabının girdisi budur: palet içi adet, tır ve konteyner yükleme adetleri
  (yeni ve eski hesaplama olmak üzere iki set), desi, ağırlık, ölçüler.
* ``veri/ornek/ihracat_masterdata.xlsx`` — mevcut müşteri master datası, `Müşteriler`
  sayfasındaki yükleme tipi, azami tonaj, tedarikçi, satış destek ve **notlar** ile
  zenginleştirilir. Notlar hangi hesap sürümünün geçerli olduğunu söyler
  ("ESKİ HESAPLAMA", "PALET YÜKSELTMELİ", "TONAJ ÖNEMLİ" ...); sistem bu metni
  `app/domain/ihracat_hesap.yukleme_kurali_coz` ile kurala çevirir.

Geçmiş sevk verisinden gelen ülke kodu, sevk adresi ve plan istatistikleri korunur;
hesaplama dosyasında olup geçmişte olmayan müşteriler yeni satır olarak eklenir.
"""
from __future__ import annotations

import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from app.domain.ihracat import arac_tipi_coz
from app.domain.ihracat_hesap import yukleme_kurali_coz
from app.domain.iller import yer_adi

URUN_HEDEFI = Path("veri/ornek/ihracat_urun_masterdata.xlsx")
MUSTERI_HEDEFI = Path("veri/ornek/ihracat_masterdata.xlsx")

URUN_BASLIKLARI = (
    "ÜRÜN KODU", "ÜRÜN", "PALET İÇİ ADET", "TIR", "KONTEYNER", "DESİ", "AĞIRLIK",
    "EN", "BOY", "YÜKSEKLİK", "Ürün Grubu", "TIR-2", "KONTEYNER-2",
    "PALET İÇİ ADET-2", "Dökme",
)
"""Hedef dosyanın kolon düzeni — `Hesaplama.xlsx` ile aynı, `veri_formatlari` bunu tanır."""

MUSTERI_BASLIKLARI = (
    "Müşteri Adı", "Ülke", "Ülke Kodu", "Sevk Adresi", "Araç Tipi", "Taşıma Modu",
    "Sefer Kodu", "Yükleme Tipi", "Azami Tonaj", "Hesaplama", "Notlar", "Tedarikçi",
    "Satış Destek", "Plan Sayısı", "Toplam Adet", "Ort. Desi", "Ort. KG", "Aktif",
)


def _sayi(deger) -> Decimal | None:
    """'#N/A' ve boş hücreler None döner; sayı olmayan hiçbir değer yazılmaz."""
    if deger is None or deger == "":
        return None
    try:
        return Decimal(str(deger).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _metin(deger) -> str:
    if deger is None:
        return ""
    metin = str(deger).strip()
    return "" if metin in {"-", "#N/A"} else metin


def _basliklari_yaz(ws, basliklar: tuple[str, ...]) -> None:
    ws.append(list(basliklar))
    for hucre in ws[1]:
        hucre.font = Font(bold=True)
    ws.freeze_panes = "A2"


def urunleri_yaz(kaynak: openpyxl.Workbook) -> int:
    ws = kaynak["Ürün"]
    hedef = openpyxl.Workbook()
    sayfa = hedef.active
    sayfa.title = "Ürünler"
    _basliklari_yaz(sayfa, URUN_BASLIKLARI)

    sayi_sutunlari = {2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14}
    adet = 0
    for satir in ws.iter_rows(min_row=2, max_col=15, values_only=True):
        kod = _metin(satir[0])
        if not kod:
            continue
        cikti = []
        for sutun, deger in enumerate(satir):
            if sutun in sayi_sutunlari:
                cikti.append(_sayi(deger))
            else:
                cikti.append(_metin(deger) or None)
        sayfa.append(cikti)
        adet += 1

    for sutun, genislik in zip("ABCDEFGHIJKLMNO", (14, 40, 12, 10, 12, 10, 10, 8, 8, 10, 14, 10, 12, 14, 10)):
        sayfa.column_dimensions[sutun].width = genislik
    URUN_HEDEFI.parent.mkdir(parents=True, exist_ok=True)
    hedef.save(URUN_HEDEFI)
    return adet


def _mevcut_musteriler() -> dict[str, dict]:
    """Geçmiş sevk verisinden üretilmiş müşteri satırlarını okur (varsa)."""
    if not MUSTERI_HEDEFI.exists():
        return {}
    wb = openpyxl.load_workbook(MUSTERI_HEDEFI, data_only=True)
    if "Müşteriler" not in wb.sheetnames:
        return {}
    ws = wb["Müşteriler"]
    basliklar = [_metin(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    kayitlar: dict[str, dict] = {}
    for satir in ws.iter_rows(min_row=2, values_only=True):
        kayit = dict(zip(basliklar, satir))
        anahtar = yer_adi(kayit.get("Müşteri Adı"))
        if anahtar:
            kayitlar[anahtar] = kayit
    return kayitlar


def musterileri_yaz(kaynak: openpyxl.Workbook) -> tuple[int, int]:
    ws = kaynak["Müşteriler"]
    mevcutlar = _mevcut_musteriler()
    # Ülke -> kod eşlemesi geçmiş veriden; hesaplama dosyasında ülke kodu yok.
    ulke_kodlari = {
        yer_adi(k.get("Ülke")): _metin(k.get("Ülke Kodu"))
        for k in mevcutlar.values()
        if _metin(k.get("Ülke Kodu"))
    }

    satirlar: dict[str, list] = {}
    yeni = 0
    for ham in ws.iter_rows(min_row=2, max_col=10, values_only=True):
        ad = _metin(ham[2])
        if not ad:
            continue
        anahtar = yer_adi(ad)
        onceki = mevcutlar.pop(anahtar, {})
        if not onceki:
            yeni += 1

        ulke = _metin(ham[1]) or _metin(onceki.get("Ülke"))
        yukleme_tipi = _metin(ham[5])
        notlar = _metin(ham[9])
        arac_tipi = arac_tipi_coz(_metin(ham[7]) or _metin(onceki.get("Araç Tipi")))
        kural = yukleme_kurali_coz(yukleme_tipi, notlar)
        satirlar[anahtar] = [
            ad,
            ulke,
            _metin(onceki.get("Ülke Kodu")) or ulke_kodlari.get(yer_adi(ulke), ""),
            _metin(onceki.get("Sevk Adresi")),
            arac_tipi.value,
            arac_tipi.tasima_modu,
            "N" if _metin(ham[0]).upper().startswith("N") else "E",
            yukleme_tipi,
            _metin(ham[6]),
            kural.ad,
            notlar,
            _metin(ham[3]),
            _metin(ham[4]),
            onceki.get("Plan Sayısı") or 0,
            onceki.get("Toplam Adet") or 0,
            onceki.get("Ort. Desi") or 0,
            onceki.get("Ort. KG") or 0,
            # "AKTİF MÜŞTERİ DEĞİL" notu olan müşteri planlamaya girmesin.
            "H" if "AKTIF MUSTERI DEGIL" in yer_adi(notlar) else "E",
        ]

    # Hesaplama dosyasında bulunmayan ama geçmişte sevk yapılmış müşteriler korunur.
    for anahtar, onceki in mevcutlar.items():
        arac_tipi = arac_tipi_coz(_metin(onceki.get("Araç Tipi")))
        satirlar[anahtar] = [
            _metin(onceki.get("Müşteri Adı")),
            _metin(onceki.get("Ülke")),
            _metin(onceki.get("Ülke Kodu")),
            _metin(onceki.get("Sevk Adresi")),
            arac_tipi.value,
            arac_tipi.tasima_modu,
            _metin(onceki.get("Sefer Kodu")) or "E",
            _metin(onceki.get("Yükleme Tipi")),
            _metin(onceki.get("Azami Tonaj")),
            yukleme_kurali_coz(_metin(onceki.get("Yükleme Tipi")), "").ad,
            "",
            _metin(onceki.get("Tedarikçi")),
            _metin(onceki.get("Satış Destek")),
            onceki.get("Plan Sayısı") or 0,
            onceki.get("Toplam Adet") or 0,
            onceki.get("Ort. Desi") or 0,
            onceki.get("Ort. KG") or 0,
            _metin(onceki.get("Aktif")) or "E",
        ]

    eski_kitap = (
        openpyxl.load_workbook(MUSTERI_HEDEFI) if MUSTERI_HEDEFI.exists() else None
    )
    hedef = openpyxl.Workbook()
    sayfa = hedef.active
    sayfa.title = "Müşteriler"
    _basliklari_yaz(sayfa, MUSTERI_BASLIKLARI)
    for anahtar in sorted(satirlar, key=lambda a: (satirlar[a][1], satirlar[a][0])):
        sayfa.append(satirlar[anahtar])
    for sutun, genislik in zip(
        "ABCDEFGHIJKLMNOPQR",
        (38, 16, 10, 18, 12, 12, 10, 26, 14, 34, 60, 18, 18, 11, 12, 10, 10, 7),
    ):
        sayfa.column_dimensions[sutun].width = genislik

    # Ülke ve özet sayfaları geçmiş analizden geliyor; olduğu gibi taşınır.
    if eski_kitap is not None:
        for ad in ("Ülkeler", "Özet"):
            if ad in eski_kitap.sheetnames:
                kaynak_sayfa = eski_kitap[ad]
                yeni_sayfa = hedef.create_sheet(ad)
                for satir in kaynak_sayfa.iter_rows(values_only=True):
                    yeni_sayfa.append(list(satir))
                for hucre in yeni_sayfa[1]:
                    hucre.font = Font(bold=True)

    hedef.save(MUSTERI_HEDEFI)
    return len(satirlar), yeni


def main() -> int:
    kaynak_yolu = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Hesaplama.xlsx")
    if not kaynak_yolu.exists():
        print(f"Hesaplama dosyası bulunamadı: {kaynak_yolu}")
        return 1

    kaynak = openpyxl.load_workbook(kaynak_yolu, data_only=True)
    urun_adedi = urunleri_yaz(kaynak)
    print(f"{urun_adedi} ürün → {URUN_HEDEFI}")
    musteri_adedi, yeni = musterileri_yaz(kaynak)
    print(f"{musteri_adedi} müşteri ({yeni} yeni) → {MUSTERI_HEDEFI}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
