"""Yükleme formunun depo operasyonun kullandığı düzene uyduğunu doğrular."""
from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.services import plan_servisi, yukleme_formu
from tests.conftest import satir_ekle, urun_ekle


def _plan_hazirla(db, depo_kodu="64", miktar=25):
    urun_ekle(db, "KMB-24", palet_ici_adet=10, tir_yukleme_adeti=100)
    for i in range(4):
        satir_ekle(db, f"TSL-{i}", "KMB-24", miktar, depo_kodu=depo_kodu, siparis_no=f"S{i}")
    return plan_servisi.plan_uret(
        db, plan_tarihi=date(2026, 8, 31), depo_kodu=depo_kodu
    ).planlar[0]


def test_form_duzeni_ve_alanlari(db, tmp_path):
    plan = _plan_hazirla(db)
    plan_servisi.axata_no_gir(db, plan, "5202")

    hedef = yukleme_formu.form_uret(plan, tmp_path / "form.xlsx")
    sayfa = load_workbook(hedef)["D-RİNG"]

    assert sayfa["J1"].value == "FORM NO : "
    assert sayfa["K1"].value == yukleme_formu.FORM_NO
    assert sayfa["B2"].value == "YÜKLEME FORMLARI"
    assert sayfa["E2"].value == "Depo "
    assert sayfa["F2"].value == "AXATA"
    assert sayfa["B3"].value == "SEFER NO"
    assert sayfa["C3"].value == "2608D1001"
    assert sayfa["B6"].value.date() == date(2026, 8, 31)

    # Depo/AXATA kutusu: numara plan deposunun satırına yazılır.
    depo_kutusu = {sayfa.cell(row=3 + i, column=5).value: sayfa.cell(row=3 + i, column=6).value
                   for i in range(5)}
    assert list(depo_kutusu) == list(yukleme_formu.DEPO_SATIRLARI)
    assert depo_kutusu["64-D DEPO"] == "5202"
    assert depo_kutusu["74-DEPO"] is None

    basliklar = [sayfa.cell(row=8, column=k).value for k in range(1, 13)]
    assert basliklar[:9] == [
        "No", "İl Adi", "Sipariş No", "Belge No", "Depo ", "Ürün Kodu", "Ürün Adi",
        "Adet", "Bayii Adı",
    ]
    assert basliklar[11] == "Teslimat"

    assert sayfa["A9"].value == "RİNG"
    assert sayfa["D9"].value == "2608D1001"
    assert sayfa["F9"].value == "KMB-24"
    assert sayfa["H9"].value == 25

    # Toplam adet ve imza alanları
    metinler = {h.value for satir in sayfa.iter_rows() for h in satir}
    assert "TOPLAM ADET" in metinler
    assert 100 in metinler
    assert "PLANLAYAN" in metinler
    assert "Sevk Kontrol" in metinler


def test_74_deposu_axata_numarasini_kendi_satirina_yazar(db, tmp_path):
    plan = _plan_hazirla(db, depo_kodu="74")
    plan_servisi.axata_no_gir(db, plan, "7788")
    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    depo_kutusu = {sayfa.cell(row=3 + i, column=5).value: sayfa.cell(row=3 + i, column=6).value
                   for i in range(5)}
    assert depo_kutusu["74-DEPO"] == "7788"
    assert depo_kutusu["64-D DEPO"] is None


def test_birden_cok_plan_tek_kitapta_alt_alta_yazilir(db, tmp_path):
    urun_ekle(db, "KMB-24", palet_ici_adet=10)
    for i in range(8):
        satir_ekle(db, f"TSL-{i}", "KMB-24", 25, siparis_no=f"S{i}")
    planlar = plan_servisi.plan_uret(
        db, plan_tarihi=date(2026, 8, 31), depo_kodu="64"
    ).planlar
    assert len(planlar) == 2

    hedef = yukleme_formu.formlari_uret(planlar, tmp_path / "hepsi.xlsx")
    sayfa = load_workbook(hedef)["D-RİNG"]
    form_satirlari = [h.row for h in sayfa["J"] if h.value == "FORM NO : "]
    assert len(form_satirlari) == 2
    sefer_nolar = {sayfa.cell(row=satir + 2, column=3).value for satir in form_satirlari}
    assert sefer_nolar == {"2608D1001", "2608D1002"}
    assert len(sayfa.row_breaks.brk) == 1  # her form ayrı sayfaya basılır


def test_depo_etiketi_eslesmesi():
    assert yukleme_formu.depo_etiketi("64") == "64-D DEPO"
    assert yukleme_formu.depo_etiketi("64-V") == "64-V DEPO"
    assert yukleme_formu.depo_etiketi("74") == "74-DEPO"
    assert yukleme_formu.depo_etiketi("34") == "34-DEPO"


def test_kutuda_satiri_olmayan_depo_icin_ek_satir_acilir(db, tmp_path):
    """Depo 03/36 gibi formun standart kutusunda bulunmayan depolarda Axata kaybolmaz."""
    urun_ekle(db, "KMB-24", palet_ici_adet=10)
    for i in range(4):
        satir_ekle(db, f"TSL-{i}", "KMB-24", 25, depo_kodu="03", siparis_no=f"S{i}")
    plan = plan_servisi.plan_uret(
        db, plan_tarihi=date(2026, 8, 31), depo_kodu="03"
    ).planlar[0]
    plan_servisi.axata_no_gir(db, plan, "5202")

    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    kutu = {
        sayfa.cell(row=3 + i, column=5).value: sayfa.cell(row=3 + i, column=6).value
        for i in range(6)
    }
    assert kutu["03-DEPO"] == "5202"
    assert kutu["64-D DEPO"] is None


def test_axata_numarasi_kutunun_disinda_da_yazilir(db, tmp_path):
    """Kutu gözden kaçsa bile numara formun üst bölümünde görünür."""
    plan = _plan_hazirla(db)
    plan_servisi.axata_no_gir(db, plan, "5202")
    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    assert sayfa["G3"].value == "AXATA NO"
    assert sayfa["H3"].value == "5202"


def test_axata_girilmemis_planda_alanlar_bos_kalir(db, tmp_path):
    plan = _plan_hazirla(db)
    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    assert not sayfa["H3"].value
    kutu = {
        sayfa.cell(row=3 + i, column=5).value: sayfa.cell(row=3 + i, column=6).value
        for i in range(5)
    }
    assert not kutu["64-D DEPO"]


def test_birden_fazla_axata_numarasi_girilebilir(db, tmp_path):
    """Depo toplama işini kolaylaştırmak için plana birden çok numara verilebilir."""
    plan = _plan_hazirla(db)
    plan_servisi.axata_no_gir(db, plan, "5322, 5323", aciklama="Panel grubu")
    plan_servisi.axata_no_gir(db, plan, "5324", aciklama="Kombi grubu")

    assert [a.numara for a in plan.axata_numaralari] == ["5322", "5323", "5324"]
    assert plan.axata_ozeti == "5322, 5323, 5324"
    assert plan.axata_no == "5322, 5323, 5324"  # arama için birleşik hâli

    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    metinler = {str(h.value) for satir in sayfa.iter_rows() for h in satir if h.value}
    assert "AXATA NUMARALARI" in metinler
    assert any("5322 — Panel grubu" == m for m in metinler)
    assert "5322, 5323, 5324" in metinler


def test_mukerrer_axata_numarasi_eklenmez(db):
    plan = _plan_hazirla(db)
    plan_servisi.axata_no_gir(db, plan, "5322")
    with pytest.raises(plan_servisi.PlanHatasi, match="zaten kayıtlı"):
        plan_servisi.axata_no_gir(db, plan, "5322")


def test_axata_numarasi_silinebilir(db):
    plan = _plan_hazirla(db)
    plan_servisi.axata_no_gir(db, plan, "5322, 5323")
    silinecek = plan.axata_numaralari[0]
    plan_servisi.axata_no_sil(db, plan, silinecek.id)
    assert [a.numara for a in plan.axata_numaralari] == ["5323"]
    assert plan.axata_no == "5323"


def test_marka_payi_yukleme_formuna_yazilir(db, tmp_path):
    """Navlun faturası dağıtımı için marka yüzdeleri forma işlenir.

    Ring planları tek depodan yüklendiği için pay %100 tek markadır. Karışık pay,
    çok depolu ortak yükleme yapılan iç piyasa planlarında oluşacak.
    """
    plan = _plan_hazirla(db, depo_kodu="64-V")
    plan_servisi.axata_no_gir(db, plan, "5322")

    assert plan.marka_paylari == {"VAİLLANT": Decimal("1.0000")}
    assert plan.marka_ozeti == "VAİLLANT %100"

    sayfa = load_workbook(yukleme_formu.form_uret(plan, tmp_path / "f.xlsx"))["D-RİNG"]
    metinler = {str(h.value) for satir in sayfa.iter_rows() for h in satir if h.value}
    assert "FATURA YÜZDESİ" in metinler
    assert "VAİLLANT" in metinler


def test_marka_payi_depo_koduna_gore_bolunur():
    """Çok depolu araçta pay, anahtar değere göre markalar arasında paylaşılır."""
    from app.domain.marka import marka, paylari_hesapla

    assert marka("64") == "DEMİRDÖKÜM"
    assert marka("-1") == "DEMİRDÖKÜM"
    assert marka("64-V") == "VAİLLANT"
    assert marka("64-P") == "PROTHERM"

    paylar = paylari_hesapla({"64": Decimal("0.25"), "64-V": Decimal("0.75")})
    assert paylar == {"DEMİRDÖKÜM": Decimal("0.2500"), "VAİLLANT": Decimal("0.7500")}

    uclu = paylari_hesapla(
        {"64": Decimal("0.50"), "64-V": Decimal("0.25"), "64-P": Decimal("0.25")}
    )
    assert uclu == {
        "DEMİRDÖKÜM": Decimal("0.5000"),
        "PROTHERM": Decimal("0.2500"),
        "VAİLLANT": Decimal("0.2500"),
    }
