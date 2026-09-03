"""İhracat planlama motorunun kuralları.

İhracatta araç **tek noktaya** gider ve kapasite iki boyutludur: doluluk ve ağırlık.

Doluluk şirketin `Hesaplama.xlsx` dosyasındaki formülle ölçülür —
Σ(miktar / yükleme adeti), 1,00 = araç %100 dolu. Yükleme adeti araç tipine (tır /
konteyner) ve müşterinin hesap sürümüne (yeni / eski) göre değişir; palet
yükseltmeli yüklemede sonuç 1,2'ye bölünür. Ağırlık ikinci sınırdır: tır 22.000 kg,
konteyner 19.500 kg (müşteriye özel tonaj bunun önüne geçer).
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest

from app.domain.ihracat import (
    AGIRLIK_KAPASITELERI,
    AracTipi,
    MusteriYuku,
    arac_tipi_coz,
    planla,
)
from app.domain.iller import yer_adi
from app.domain.kapasite import IHRACAT_KONTEYNER, IHRACAT_TIR
from app.domain.planlama import Teslimat


def teslimat(no, doluluk, agirlik=0, depo="34", satir_id=None):
    """`doluluk` teslimatın araç payıdır: 0,5 = yarım araç."""
    pay = Decimal(str(doluluk))
    return Teslimat(
        teslimat_no=no,
        depo_kodu=depo,
        planlama_anahtari="U1",
        urun_kodu="U1",
        urun_adi="U1",
        miktar=Decimal(10),
        birim=pay,
        anahtar=pay,
        agirlik=Decimal(str(agirlik)),
        desi=pay * 20000,
        oncelik_tarihi=date(2026, 9, 1),
        satir_idleri=(satir_id or int(no[-1]) if no[-1].isdigit() else 1,),
        depo_katkilari={depo: pay},
    )


def yuk(
    ad, doluluk, agirlik=0, tip=AracTipi.TIR, teslimatlar=None, azami=None,
    ulke="ALMANYA",
):
    teslimatlar = teslimatlar or (teslimat(f"{ad}-1", doluluk, agirlik),)
    return MusteriYuku(
        anahtar=ad,
        musteri_adi=ad,
        ulke=ulke,
        ulke_kodu="DE",
        sevk_adresi="BERLIN",
        teslimatlar=tuple(teslimatlar),
        doluluk=Decimal(str(doluluk)),
        desi=sum((t.desi for t in teslimatlar), Decimal(0)),
        agirlik=Decimal(str(agirlik)),
        adet=sum((t.miktar for t in teslimatlar), Decimal(0)),
        arac_tipi=tip,
        azami_agirlik=Decimal(str(azami)) if azami else None,
    )


# ------------------------------------------------------------------- araç tipi


def test_konteyner_deniz_tir_kara_yoludur():
    """Taşıma modu araç tipinden çıkar: Şili konteyner (deniz), Romanya tır (kara)."""
    assert AracTipi.KONTEYNER.tasima_modu == "DENİZ"
    assert AracTipi.TIR.tasima_modu == "KARA"
    assert AracTipi.KONTEYNER.deniz_mi
    assert not AracTipi.TIR.deniz_mi


def test_arac_tipi_serbest_metinden_cozulur():
    """Sahada '1X40 DC', '1X40 HC', 'Konteyner' hepsi konteyner demek."""
    assert arac_tipi_coz("1X40 DC") is AracTipi.KONTEYNER
    assert arac_tipi_coz("1X40 HC") is AracTipi.KONTEYNER
    assert arac_tipi_coz("Konteyner") is AracTipi.KONTEYNER
    assert arac_tipi_coz("TIR") is AracTipi.TIR
    assert arac_tipi_coz("PARSİYEL") is AracTipi.PARSIYEL
    assert arac_tipi_coz("DHL") is AracTipi.KARGO
    assert arac_tipi_coz("") is AracTipi.TIR


def test_doluluk_bir_olunca_arac_dolar():
    """Kapasite farkı yükleme adetlerinde: her iki araç da 1,00 doluluğa gider."""
    assert IHRACAT_TIR.ust_limit == Decimal(1)
    assert IHRACAT_KONTEYNER.ust_limit == Decimal(1)
    assert AGIRLIK_KAPASITELERI[AracTipi.TIR] == Decimal(22000)
    assert AGIRLIK_KAPASITELERI[AracTipi.KONTEYNER] == Decimal(19500)


# ------------------------------------------------------------------- planlama


def test_her_musteri_kendi_aracina_yuklenir():
    """İhracat aracı tek noktaya gider: iki müşteri aynı araca binmez."""
    sonuc = planla([yuk("A", "0.95", 15000), yuk("B", "0.90", 14000)])
    assert len(sonuc.planlar) == 2
    assert {p.musteri.musteri_adi for p in sonuc.planlar} == {"A", "B"}


def test_araci_asan_musteri_birden_fazla_araca_bolunur():
    """Teslimat bölünmez; araç sayısını bölünmezlik belirler.

    Yarım araçtan büyük üç teslimat ikişer birleşemez (%110 > %100), o yüzden üç araç
    çıkar ve hepsi yarım kalır. Bu kuralın ihlali değil, bölünmezliğin sonucudur —
    alt limit araç başına değil müşteri toplamına uygulanır.
    """
    teslimatlar = tuple(teslimat(f"T{i}", "0.55", 8000, satir_id=i) for i in range(3))
    sonuc = planla([yuk("DEV", "1.65", 24000, teslimatlar=teslimatlar)])
    assert len(sonuc.planlar) == 3
    assert all(p.hacim <= IHRACAT_TIR.ust_limit for p in sonuc.planlar)
    assert sum(len(p.teslimatlar) for p in sonuc.planlar) == 3
    assert not any(p.alt_limit_esnetildi for p in sonuc.planlar)


def test_agirlik_sinira_takilirsa_hacim_dolmadan_arac_kapanir():
    """Ağır ama küçük hacimli yük: aracı ağırlık doldurur, doluluk değil."""
    teslimatlar = (
        teslimat("T1", "0.30", 12000, satir_id=1),
        teslimat("T2", "0.30", 12000, satir_id=2),
    )
    sonuc = planla([yuk("AĞIR", "0.60", 24000, teslimatlar=teslimatlar)])
    assert len(sonuc.planlar) == 2
    for plan in sonuc.planlar:
        assert plan.agirlik <= Decimal(22000)
        assert plan.kisitlayan == "AĞIRLIK"


def test_musteriye_ozel_azami_tonaj_varsayilanin_onune_gecer():
    """Marka kılavuzundaki azami tonaj araç tipinin varsayılanını ezer."""
    teslimatlar = (
        teslimat("T1", "0.25", 10000, satir_id=1),
        teslimat("T2", "0.25", 10000, satir_id=2),
    )
    dusuk = yuk("SINIRLI", "0.50", 20000, teslimatlar=teslimatlar, azami=19500)
    assert dusuk.agirlik_kapasitesi == Decimal(19500)
    sonuc = planla([dusuk])
    assert len(sonuc.planlar) == 2


def test_alt_limiti_dolduramayan_musteri_beklemede_kalir():
    """Alt limit müşteri toplamına bakar: bu müşteriye bugün araç kaldırmaya değer mi?"""
    sonuc = planla([yuk("KÜÇÜK", "0.25", 3000)])
    assert sonuc.planlar == []
    assert "Yeterli hacim yok" in sonuc.bekleyenler[0].sebep
    assert "%25.0 araç" in sonuc.bekleyenler[0].sebep
    assert "alt limit %75" in sonuc.bekleyenler[0].sebep


def test_kalanlari_zorla_alt_limiti_atlar():
    sonuc = planla([yuk("KÜÇÜK", "0.25", 3000)], kalanlari_zorla=True)
    assert len(sonuc.planlar) == 1
    assert sonuc.planlar[0].alt_limit_esnetildi


def test_tek_basina_araci_asan_teslimat_istisna_olur():
    buyuk = yuk("DEV", "1.40", 10000, teslimatlar=(teslimat("T1", "1.40", 10000),))
    sonuc = planla([buyuk])
    assert len(sonuc.planlar) == 1
    assert sonuc.planlar[0].istisna_asim
    assert sonuc.planlar[0].doluluk_yuzdesi > 100


# --------------------------------------------------------------- servis katmanı


@pytest.fixture()
def ihracat_veri(db):
    from app.models import IhracatMusterisi

    db.add(
        IhracatMusterisi(
            anahtar=yer_adi("VAILLANT D.O.O."),
            musteri_adi="VAILLANT D.O.O.",
            ulke="HIRVATİSTAN",
            ulke_kodu="HR",
            arac_tipi="TIR",
            sefer_kodu="N",
            yukleme_tipi="PALET YÜKSELTME",
            aciklama="hava yastığı kullanılacak",
        )
    )
    db.add(
        IhracatMusterisi(
            anahtar=yer_adi("ANSAL REFRIGERACION SA"),
            musteri_adi="ANSAL REFRIGERACION SA",
            ulke="ARJANTİN",
            ulke_kodu="AR",
            arac_tipi="KONTEYNER",
            sefer_kodu="E",
            aciklama="silika jel",
        )
    )
    db.flush()
    return db


def _ihracat_satiri(db, siparis, teslimat_no, musteri, desi, kg, urun="U1", depo="34"):
    from app.models import SiparisSatiri

    satir = SiparisSatiri(
        siparis_no=siparis,
        siparis_satir_no=urun,
        teslimat_no=teslimat_no,
        urun_kodu=urun,
        urun_adi=f"{urun} ürünü",
        miktar=Decimal(10),
        depo_kodu=depo,
        bayi_adi=musteri,
        sehir="HIRVATİSTAN",
        ulke_kodu="HR",
        sevk_adresi="OSIJEK",
        desi=Decimal(str(desi)),
        agirlik=Decimal(str(kg)),
        modul="IHRACAT",
    )
    db.add(satir)
    db.flush()
    return satir


def test_sefer_numarasinin_belge_kodu_musteriden_gelir(ihracat_veri):
    """NSC müşterisi N, Export müşterisi E kodlu sefer alır."""
    from app.services import ihracat_servisi

    db = ihracat_veri
    _ihracat_satiri(db, "S1", "T1", "VAILLANT D.O.O.", 20000, 15000)
    _ihracat_satiri(db, "S2", "T2", "ANSAL REFRIGERACION SA", 14000, 12000, urun="U2")

    sonuc = ihracat_servisi.plan_uret(
        db, plan_tarihi=date(2026, 9, 1), kullanici="test"
    )
    kodlar = {p.musteri_adi: p.sefer_no[4] for p in sonuc.planlar}
    assert kodlar["VAILLANT D.O.O."] == "N"
    assert kodlar["ANSAL REFRIGERACION SA"] == "E"


def test_musteri_master_datasi_arac_ve_notu_belirler(ihracat_veri):
    from app.services import ihracat_servisi

    db = ihracat_veri
    _ihracat_satiri(db, "S1", "T1", "ANSAL REFRIGERACION SA", 14000, 12000)

    sonuc = ihracat_servisi.plan_uret(
        db, plan_tarihi=date(2026, 9, 1), kullanici="test"
    )
    plan = sonuc.planlar[0]
    assert plan.arac_tipi == "KONTEYNER"
    assert plan.tasima_modu == "DENİZ"
    assert plan.musteri_aciklamasi == "silika jel"
    assert plan.modul == "IHRACAT"
    assert plan.plan_tipi == "IHRACAT_KONTEYNER"


def test_master_datada_olmayan_musteri_tir_varsayilir(ihracat_veri):
    from app.services import ihracat_servisi

    db = ihracat_veri
    _ihracat_satiri(db, "S9", "T9", "YENİ MÜŞTERİ LTD", 20000, 15000)

    sonuc = ihracat_servisi.plan_uret(
        db, plan_tarihi=date(2026, 9, 1), kullanici="test"
    )
    assert "YENİ MÜŞTERİ LTD" in sonuc.tanimsiz_musteriler
    assert sonuc.planlar[0].arac_tipi == "TIR"
    assert "master datada yok" in sonuc.ozet()


def test_ihracat_formu_musteri_notunu_ve_arac_bloklarini_yazar(ihracat_veri, tmp_path):
    from openpyxl import load_workbook

    from app.services import ihracat_servisi, ihracat_yukleme_formu, plan_servisi

    db = ihracat_veri
    _ihracat_satiri(db, "S1", "T1", "VAILLANT D.O.O.", 20000, 15000)
    sonuc = ihracat_servisi.plan_uret(db, plan_tarihi=date(2026, 9, 1))
    plan = sonuc.planlar[0]
    plan_servisi.axata_no_gir(db, plan, "2735", "test")
    ihracat_servisi.arac_bilgisi_kaydet(
        db, plan, "OMSAN", "34 ABC 12", "MSCU1234567", "MÜHÜR-9", "Ali", "test"
    )

    hedef = ihracat_yukleme_formu.form_uret(plan, tmp_path / "form.xlsx")
    sayfa = load_workbook(hedef)["EXPORT"]
    metinler = [h.value for satir in sayfa.iter_rows() for h in satir if h.value]

    assert plan.sefer_no in metinler
    assert "hava yastığı kullanılacak" in metinler
    assert "DORSE/KONTEYNER NO" in metinler
    assert "MSCU1234567" in metinler
    assert "MÜHÜR NO" in metinler
    assert "PALET YÜKSELTME" in metinler
    assert "34-DEPO AXATA" in metinler


# ------------------------------------------------- Hesaplama.xlsx doluluk modeli


def _olcu(**degerler):
    from app.domain.ihracat_hesap import UrunOlcusu

    varsayilan = {
        "urun_kodu": "313041213",
        "palet_ici_adet": Decimal(120),
        "tir_yukleme_adeti": Decimal(3000),
        "konteyner_yukleme_adeti": Decimal(2640),
        "palet_ici_adet_eski": Decimal(112),
        "tir_yukleme_adeti_eski": Decimal(2800),
        "konteyner_yukleme_adeti_eski": Decimal(2464),
        "desi": Decimal("2.5"),
        "agirlik": Decimal(12),
    }
    varsayilan.update(degerler)
    return UrunOlcusu(**varsayilan)


def test_doluluk_yukleme_adetinden_hesaplanir():
    """Şirketin formülü: DOLULUK = Σ(miktar / yükleme adeti), yuvarlama yok."""
    from app.domain.ihracat_hesap import Kalem, hesapla

    sonuc = hesapla([Kalem("313041213", Decimal(1500), _olcu())], "TIR")
    assert sonuc.doluluk == Decimal("0.5")  # 1500 / 3000
    assert sonuc.palet == Decimal("12.5")  # 1500 / 120
    assert sonuc.desi == Decimal("3750.0")  # 2,5 × 1500
    assert sonuc.agirlik == Decimal(18000)  # 12 × 1500


def test_konteyner_ve_tir_ayri_yukleme_adeti_kullanir():
    from app.domain.ihracat_hesap import Kalem, hesapla

    kalem = Kalem("313041213", Decimal(2640), _olcu())
    assert hesapla([kalem], "KONTEYNER").doluluk == Decimal(1)
    assert hesapla([kalem], "TIR").doluluk == Decimal("0.88")


def test_eski_hesaplama_iki_numarali_sutunlari_kullanir():
    """Notu "ESKİ HESAPLAMA" olan müşteri `-2` sütunlarıyla hesaplanır."""
    from app.domain.ihracat_hesap import Kalem, hesapla, yukleme_kurali_coz

    kural = yukleme_kurali_coz("STANDART", "ESKİ HESAPLAMA")
    kalem = Kalem("313041213", Decimal(2800), _olcu())
    assert hesapla([kalem], "TIR", kural).doluluk == Decimal(1)
    # Yeni hesaplamada aynı miktar tırı doldurmaz: 2800 / 3000.
    assert hesapla([kalem], "TIR").doluluk < Decimal(1)


def test_palet_yukseltmeli_yuklemede_doluluk_bire_iki_bolunur():
    """Paletler üst üste istiflenince araca %20 daha fazla yük girer."""
    from app.domain.ihracat_hesap import Kalem, hesapla, yukleme_kurali_coz

    kalem = Kalem("313041213", Decimal(3000), _olcu())
    kural = yukleme_kurali_coz("PALET YÜKSETLME", "YENİ HESAPLAMA")
    assert kural.palet_yukseltme
    assert hesapla([kalem], "TIR", kural).doluluk == Decimal(1) / Decimal("1.2")


def test_yukleme_kurali_serbest_metinden_cozulur():
    """Kaynak dosyada yazım değişiyor: YÜKSELTME / YÜKSETLME / YÜKSELTMELİ."""
    from app.domain.ihracat_hesap import HesaplamaTipi, yukleme_kurali_coz

    kural = yukleme_kurali_coz(
        "PALET YÜKSELTME-KÖŞEBENT",
        "ESKİ HESAPLAMA PALET YÜKSELTMELİ VE DÖKME TONAJ ÖNEMLİ",
    )
    assert kural.hesaplama is HesaplamaTipi.ESKI
    assert kural.palet_yukseltme and kural.dokme and kural.kosebent
    assert kural.tonaj_onemli
    # "YENİ HESAPLAMA" geçen not eski sürüme düşmez.
    assert yukleme_kurali_coz("STANDART", "YENİ HESAPLAMA").hesaplama is HesaplamaTipi.YENI


def test_olcusu_olmayan_urun_desiden_yaklasik_hesaplanir():
    """Master datada tır/konteyner adedi boş olan ürünler planlamayı durdurmaz."""
    from app.domain.ihracat_hesap import DESI_KAPASITELERI, Kalem, hesapla

    kalem = Kalem("YOK-1", Decimal(100), None, desi=Decimal(10750))
    sonuc = hesapla([kalem], "TIR")
    assert sonuc.doluluk == Decimal(10750) / DESI_KAPASITELERI["TIR"]
    assert sonuc.olcusuz_kodlar == ("YOK-1",)


@pytest.fixture()
def ihracat_urun(db):
    from app.models import IhracatUrunu

    db.add(
        IhracatUrunu(
            urun_kodu="U1",
            urun_adi="25 DD S 22 300 0400 V0 A1 G1",
            urun_grubu="Radiator",
            palet_ici_adet=Decimal(120),
            tir_yukleme_adeti=Decimal(3000),
            konteyner_yukleme_adeti=Decimal(2640),
            tir_yukleme_adeti_eski=Decimal(2800),
            konteyner_yukleme_adeti_eski=Decimal(2464),
            palet_ici_adet_eski=Decimal(112),
            desi=Decimal("2.5"),
            agirlik=Decimal(6),
        )
    )
    db.flush()
    return db


def test_plan_dolulugu_urun_master_datasindan_gelir(ihracat_veri, ihracat_urun):
    """Sipariş satırındaki desi değil, ürünün yükleme adeti belirleyicidir."""
    from app.models import SiparisSatiri
    from app.services import ihracat_servisi

    db = ihracat_veri
    satir = _ihracat_satiri(db, "S1", "T1", "VAILLANT D.O.O.", 0, 0)
    satir.miktar = Decimal(3000)
    db.flush()

    sonuc = ihracat_servisi.plan_uret(db, plan_tarihi=date(2026, 9, 1))
    plan = sonuc.planlar[0]
    # 3000 adet bir tırı doldurur; müşteri palet yükseltmeli olduğu için %83,3.
    assert plan.doluluk_yuzdesi == Decimal("83.33")
    assert plan.toplam_palet == Decimal(25)  # 3000 / 120
    assert plan.toplam_desi == Decimal(7500)  # 2,5 × 3000
    assert plan.toplam_agirlik == Decimal(18000)  # 6 × 3000
    assert db.get(SiparisSatiri, satir.id).plan_id == plan.id


def test_ihracat_urun_master_datasi_hesaplama_dosyasindan_yuklenir(db, tmp_path):
    """`Hesaplama.xlsx` → `Ürün` sayfası olduğu gibi yüklenebilmeli."""
    from openpyxl import Workbook

    from app.models import IhracatUrunu
    from app.services import ice_aktarim

    kitap = Workbook()
    sayfa = kitap.active
    sayfa.append([
        "ÜRÜN KODU", "ÜRÜN", "PALET İÇİ ADET", "TIR", "KONTEYNER", "DESİ", "AĞIRLIK",
        "EN", "BOY", "YÜKSEKLİK", "Ürün Grubu", "TIR-2", "KONTEYNER-2",
        "PALET İÇİ ADET-2", "Dökme",
    ])
    sayfa.append([
        313041213, "25 DD S 22 300 0400 V0 A1 G1", 120, 3000, 2640, 2.5, 12,
        60, 120, 77, "Radiator", 2800, 2464, 112, None,
    ])
    # Ölçüsü olmayan ürün: kayıt alınır, uyarı verilir.
    sayfa.append([999, "ÖLÇÜSÜZ", None, None, None, None, None, None, None, None,
                  "Acc.", None, None, None, None])
    dosya = tmp_path / "hesaplama.xlsx"
    kitap.save(dosya)

    sonuc = ice_aktarim.ihracat_urunlerini_aktar(db, dosya, "hesaplama.xlsx")
    assert sonuc.eklenen == 2
    assert sonuc.hatali == 0
    assert len(sonuc.uyarilar) == 1

    urun = db.query(IhracatUrunu).filter_by(urun_kodu="313041213").one()
    assert urun.tir_yukleme_adeti == Decimal(3000)
    assert urun.tir_yukleme_adeti_eski == Decimal(2800)
    assert urun.urun_grubu == "Radiator"
    assert urun.olculebilir_mi


def test_tekrar_eden_urun_satiri_dolu_olcuyu_silmez(db, tmp_path):
    """Aynı ürün kodu dosyada iki kez geçiyor ve ikincisi boş: ölçüler korunur.

    Şirketin hesaplama dosyasında üç ürün böyle; son satır kazansaydı o ürünlerin
    yükleme adetleri silinirdi.
    """
    from openpyxl import Workbook

    from app.models import IhracatUrunu
    from app.services import ice_aktarim

    kitap = Workbook()
    sayfa = kitap.active
    sayfa.append(["ÜRÜN KODU", "ÜRÜN", "PALET İÇİ ADET", "TIR", "KONTEYNER"])
    sayfa.append([813182240, "22-300 180CM", 32, 672, 576])
    sayfa.append([813182240, "22-300 180CM", None, None, None])
    dosya = tmp_path / "tekrar.xlsx"
    kitap.save(dosya)

    sonuc = ice_aktarim.ihracat_urunlerini_aktar(db, dosya, "tekrar.xlsx")
    assert sonuc.eklenen == 1
    assert sonuc.birlestirilen == 1
    urun = db.query(IhracatUrunu).filter_by(urun_kodu="813182240").one()
    assert urun.tir_yukleme_adeti == Decimal(672)
    assert urun.konteyner_yukleme_adeti == Decimal(576)


def test_gomulu_master_data_ilk_kurulumda_yuklenir(db):
    """Program kendi master datasıyla gelir: ilk açılışta dosya yüklemek gerekmez."""
    from app.models import IhracatMusterisi, IhracatUrunu
    from app.services import gomulu_veri

    mesajlar = gomulu_veri.eksikleri_yukle(db)
    assert any("İhracat ürün master datası" in m for m in mesajlar)
    assert db.query(IhracatUrunu).count() > 2500
    assert db.query(IhracatMusterisi).count() > 150

    # Hesaplama dosyasındaki örnek ürün ölçüleriyle birlikte gelmeli.
    urun = db.query(IhracatUrunu).filter_by(urun_kodu="313041213").one()
    assert urun.tir_yukleme_adeti == Decimal(3000)
    assert urun.konteyner_yukleme_adeti == Decimal(2640)
    assert urun.tir_yukleme_adeti_eski == Decimal(2800)

    # İkinci çağrı hiçbir şeyi ezmez: tablolar artık dolu.
    urun.tir_yukleme_adeti = Decimal(1)
    db.flush()
    assert gomulu_veri.eksikleri_yukle(db) == []
    assert db.query(IhracatUrunu).filter_by(urun_kodu="313041213").one().tir_yukleme_adeti == Decimal(1)


def test_gomulu_musteri_datasi_hesap_surumunu_tasir(db):
    """Notlar hangi hesabın geçerli olduğunu söyler; sistem bunu kurala çeviriyor."""
    from app.domain.ihracat_hesap import HesaplamaTipi
    from app.models import IhracatMusterisi
    from app.services import gomulu_veri

    gomulu_veri.eksikleri_yukle(db)
    eskiler = [
        m for m in db.query(IhracatMusterisi).all()
        if m.yukleme_kurali.hesaplama is HesaplamaTipi.ESKI
    ]
    # Master datada 60'tan fazla müşteri eski hesaplamayla yükleniyor.
    assert len(eskiler) > 60
    assert any(m.yukleme_kurali.palet_yukseltme for m in eskiler)
