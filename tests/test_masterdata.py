"""Master Data modülü: filtreleme, dışa aktarma, tekil güncelleme, depo ve ayarlar."""
from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import load_workbook

from app.models import Ayar, Depo, Urun
from app.services import ice_aktarim, masterdata_servisi as md
from app.services.veri_formatlari import URUN_ALANLARI
from tests.conftest import urun_ekle


@pytest.fixture()
def urunler(db):
    urun_ekle(db, "TAM", palet_ici_adet=10, tir_yukleme_adeti=100, grup="PANEL")
    tam = db.query(Urun).filter_by(urun_kodu="TAM").one()
    tam.agirlik = Decimal(20)
    tam.desi = Decimal(5)
    tam.palet_en, tam.palet_boy, tam.palet_yukseklik = 80, 120, 160
    tam.kamyon_yukleme_adeti, tam.tir_palet, tam.kamyon_palet = 50, 10, 5

    urun_ekle(db, "OLCUSUZ", palet_ici_adet=10, tir_yukleme_adeti=100, grup="KOMBİ")
    urun_ekle(db, "PALETSIZ", palet_ici_adet=None, tir_yukleme_adeti=100, grup="PANEL")
    db.flush()
    return db


def test_eksik_alan_filtresi_yalnizca_o_eksigi_olanlari_getirir(urunler):
    palet_eksik = md.urunleri_getir(urunler, md.UrunFiltresi(eksik="PALET"))
    assert {u.urun_kodu for u in palet_eksik} == {"PALETSIZ"}

    olcu_eksik = md.urunleri_getir(urunler, md.UrunFiltresi(eksik="OLCU"))
    assert {u.urun_kodu for u in olcu_eksik} == {"OLCUSUZ", "PALETSIZ"}

    herhangi = md.urunleri_getir(urunler, md.UrunFiltresi(eksik=md.EKSIK_HERHANGI))
    assert "TAM" not in {u.urun_kodu for u in herhangi}


def test_grup_ve_arama_filtreleri_birlikte_calisir(urunler):
    sonuc = md.urunleri_getir(urunler, md.UrunFiltresi(urun_grubu="PANEL"))
    assert {u.urun_kodu for u in sonuc} == {"TAM", "PALETSIZ"}
    sonuc = md.urunleri_getir(urunler, md.UrunFiltresi(urun_grubu="PANEL", arama="TAM"))
    assert {u.urun_kodu for u in sonuc} == {"TAM"}


def test_eksik_ozeti_alan_bazinda_sayar(urunler):
    ozet = {o["kod"]: o["sayi"] for o in md.urun_eksik_ozeti(urunler)}
    assert ozet["PALET"] == 1          # PALETSIZ
    assert ozet["OLCU"] == 2           # OLCUSUZ + PALETSIZ
    assert ozet["AGIRLIK"] == 2


def test_indirilen_dosya_geri_yuklenebilir(urunler, tmp_path):
    """Dışa aktarım içe aktarımın başlıklarını kullanır: dosya tur atıp geri gelir."""
    hedef = md.disari_aktar(
        md.urunleri_getir(urunler, md.UrunFiltresi()),
        URUN_ALANLARI,
        md.URUN_DEGERLERI,
        tmp_path / "urun.xlsx",
        "Ürünler",
    )
    sayfa = load_workbook(hedef).active
    basliklar = [h.value for h in sayfa[1]]
    assert basliklar == [alan.baslik for alan in URUN_ALANLARI]

    # Eksik ölçüyü dosyada doldur, aynı dosyayı geri yükle.
    kolon = basliklar.index("Palet içi adet") + 1
    kod_kolonu = basliklar.index("StokKodu") + 1
    for satir in range(2, sayfa.max_row + 1):
        if sayfa.cell(row=satir, column=kod_kolonu).value == "PALETSIZ":
            sayfa.cell(row=satir, column=kolon, value=24)
    sayfa.parent.save(hedef)

    sonuc = ice_aktarim.urunleri_aktar(urunler, hedef, "urun.xlsx")
    assert sonuc.eklenen == 0 and sonuc.guncellenen == 3
    assert urunler.query(Urun).filter_by(urun_kodu="PALETSIZ").one().palet_ici_adet == 24


def test_tekil_guncellemede_bos_alan_silinir(urunler):
    md.urunu_guncelle(
        urunler,
        "TAM",
        {"urun_adi": "TAM ürün", "urun_grubu": "PANEL", "palet_ici_adet": "10",
         "tir_yukleme_adeti": "100", "agirlik": "", "desi": "7,5"},
    )
    urun = urunler.query(Urun).filter_by(urun_kodu="TAM").one()
    assert urun.agirlik is None            # boş bırakıldı -> silindi
    assert urun.desi == Decimal("7.5")     # virgül kabul edilir
    assert urun.palet_ici_adet == 10


def test_toplu_yuklemede_bos_sutun_silmez(urunler, tmp_path):
    """Excel'de boş gelen alan mevcut veriyi silmemeli; tekil formun tersi kural."""
    from openpyxl import Workbook

    kitap = Workbook()
    sayfa = kitap.active
    sayfa.append(["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Ağırlık"])
    sayfa.append(["TAM", "TAM ürün", "PANEL", 10, None])
    yol = tmp_path / "kismi.xlsx"
    kitap.save(yol)

    ice_aktarim.urunleri_aktar(urunler, yol, "kismi.xlsx")
    assert urunler.query(Urun).filter_by(urun_kodu="TAM").one().agirlik == Decimal(20)


def test_olcusuz_urun_kaydedilemez(urunler):
    with pytest.raises(md.MasterDataHatasi, match="en az biri"):
        md.urunu_guncelle(
            urunler,
            "YENI",
            {"urun_adi": "Yeni", "palet_ici_adet": "", "tir_yukleme_adeti": "",
             "kamyon_yukleme_adeti": ""},
        )


def test_depo_tanimlari_varsayilanlarla_yuklenir(db):
    assert md.depolari_yukle(db) == len(md.VARSAYILAN_DEPOLAR)
    assert md.depolari_yukle(db) == 0          # ikinci çağrı mevcutlara dokunmaz

    kodlar = {d.kod for d in md.depolari_getir(db)}
    assert {"64", "64-V", "74", "34", "44", "-1"} <= kodlar
    # Bayi ortak deposunda Axata iş emri açılmaz.
    assert db.query(Depo).filter_by(kod="-1").one().axata_var is False
    # Form satırları eski sabit listeyle aynı sırada başlar.
    assert md.form_depo_satirlari(db)[:5] == [
        "34-DEPO", "44-DEPO", "64-D DEPO", "64-V DEPO", "74-DEPO"
    ]


def test_yeni_depo_yukleme_formunun_kutusuna_dusuyor(db):
    md.depolari_yukle(db)
    md.depoyu_kaydet(
        db, "80",
        {"ad": "Yeni depo", "tesis": "ESKİŞEHİR", "form_etiketi": "80-DEPO",
         "sira": "70", "axata_var": "E", "parsiyel_yapilir": "H", "aktif": "E"},
    )
    assert "80-DEPO" in md.form_depo_satirlari(db)

    # Pasife alınan depo formda görünmez.
    md.depoyu_kaydet(db, "80", {"ad": "Yeni depo", "form_etiketi": "80-DEPO",
                                "sira": "70", "aktif": "H"})
    assert "80-DEPO" not in md.form_depo_satirlari(db)


def test_ayarlar_kayit_yoksa_varsayilani_verir(db):
    kurallar = md.kurallari_kur(db)
    from app.domain.ic_piyasa import VARSAYILAN_KURALLAR

    assert kurallar == VARSAYILAN_KURALLAR


def test_ayar_degisikligi_planlama_kurallarina_yansir(db):
    degisenler = md.ayarlari_kaydet(
        db, {"kargo_desi_siniri": "25", "azami_sapma_km": "150"}, "admin"
    )
    assert len(degisenler) == 2
    kurallar = md.kurallari_kur(db)
    assert kurallar.kargo_desi_siniri == Decimal(25)
    assert kurallar.azami_sapma_km == 150
    # Dokunulmayan ayar varsayılanda kalır.
    assert kurallar.azami_durak == 5


def test_gecersiz_ayar_reddedilir(db):
    with pytest.raises(md.MasterDataHatasi, match="sayı değil"):
        md.ayarlari_kaydet(db, {"azami_durak": "beş"}, "admin")
    with pytest.raises(md.MasterDataHatasi, match="sıfırdan büyük"):
        md.ayarlari_kaydet(db, {"azami_durak": "0"}, "admin")
    assert db.query(Ayar).count() == 0


# ------------------------------------------------------------- ürün grupları


def test_turkce_buyuk_harf_grubu_ikiye_bolmez(db, tmp_path):
    """'Klima' ile 'KLİMA' aynı gruptur; Python'un upper()'ı bunları ayırıyordu.

    `'Klima'.upper()` -> 'KLIMA' (noktasız I), 'KLİMA' ile eşleşmez.
    """
    from openpyxl import Workbook

    kitap = Workbook()
    sayfa = kitap.active
    sayfa.append(["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"])
    sayfa.append(["A1", "Klima A", "Klima", 10, 100])
    sayfa.append(["A2", "Klima B", "KLİMA", 10, 100])
    sayfa.append(["A3", "Kombi", "kombi", 10, 100])
    yol = tmp_path / "u.xlsx"
    kitap.save(yol)

    ice_aktarim.urunleri_aktar(db, yol, "u.xlsx")
    gruplar = {u.urun_grubu for u in db.query(Urun).all()}
    assert gruplar == {"KLİMA", "KOMBİ"}


def test_grup_adi_degisince_butun_urunler_guncellenir(db, urunler):
    """Ad değişikliği master datanın tamamına işler."""
    sayi = md.grubu_yeniden_adlandir(db, md.IC_PIYASA, "PANEL", "Radyatör")
    assert sayi == 2
    # İç piyasada ad Türkçe kurallarına göre büyük harfe çevrilir.
    assert {u.urun_grubu for u in db.query(Urun).all()} == {"RADYATÖR", "KOMBİ"}


def test_ayni_ada_tasinan_grup_birlesir(db, urunler):
    """Yazım hatasıyla ikiye bölünmüş grubu toplamanın yolu budur."""
    db.query(Urun).filter_by(urun_kodu="OLCUSUZ").update({"urun_grubu": "KLIMA"})
    db.flush()
    md.grubu_yeniden_adlandir(db, md.IC_PIYASA, "KLIMA", "PANEL")
    assert {u.urun_grubu for u in db.query(Urun).all()} == {"PANEL"}
    assert db.query(Urun).filter_by(urun_grubu="PANEL").count() == 3


def test_gruptan_cikarilan_urun_silinmez(db, urunler):
    sayi = md.grubu_sil(db, md.IC_PIYASA, "PANEL")
    assert sayi == 2
    assert db.query(Urun).count() == 3
    assert db.query(Urun).filter(Urun.urun_grubu.is_(None)).count() == 2


def test_grup_ozeti_yalnizca_yazimla_ayrisani_isaretler(db, urunler):
    db.query(Urun).filter_by(urun_kodu="OLCUSUZ").update({"urun_grubu": "KLIMA"})
    db.query(Urun).filter_by(urun_kodu="PALETSIZ").update({"urun_grubu": "KLİMA"})
    db.flush()
    ozet = {g["ad"]: g for g in md.urun_gruplari_ozeti(db) if g["kapsam"] == md.IC_PIYASA}
    assert ozet["KLIMA"]["cakisanlar"] == ["KLİMA"]
    assert ozet["KLİMA"]["cakisanlar"] == ["KLIMA"]
    assert ozet["PANEL"]["cakisanlar"] == []


def test_ihracat_gruplari_ic_piyasadan_ayridir(db, urunler):
    """İhracat grupları İngilizce ve karışık yazımlı; büyük harfe çevrilmez."""
    from app.models import IhracatUrunu

    db.add(IhracatUrunu(urun_kodu="E1", urun_adi="Rad", urun_grubu="Towel Heater",
                        tir_yukleme_adeti=Decimal(3000)))
    db.flush()
    md.grubu_yeniden_adlandir(db, md.IHRACAT, "Towel Heater", "Towel Radiator")
    assert db.query(IhracatUrunu).one().urun_grubu == "Towel Radiator"
    # İç piyasa grupları etkilenmez.
    assert {u.urun_grubu for u in db.query(Urun).all()} == {"PANEL", "KOMBİ"}


def test_bilinmeyen_kapsam_reddedilir(db):
    with pytest.raises(md.MasterDataHatasi, match="Bilinmeyen kapsam"):
        md.grubu_yeniden_adlandir(db, "BASKA", "A", "B")


# ------------------------------------------------------------ ihracat ürünleri


@pytest.fixture()
def ihracat_urunler(db):
    from app.models import IhracatUrunu

    db.add_all(
        [
            IhracatUrunu(urun_kodu="TAM", urun_adi="Radiator", urun_grubu="Radiator",
                         palet_ici_adet=Decimal(120), tir_yukleme_adeti=Decimal(3000),
                         konteyner_yukleme_adeti=Decimal(2640), desi=Decimal("2.5"),
                         agirlik=Decimal(12), en=60, boy=120),
            IhracatUrunu(urun_kodu="OLCUSUZ", urun_adi="Valve", urun_grubu="Acc."),
            IhracatUrunu(urun_kodu="DESILI", urun_adi="Sensor", desi=Decimal("0.1")),
        ]
    )
    db.flush()
    return db


def test_ihracat_eksik_filtresi(db, ihracat_urunler):
    olcusuz = md.ihracat_urunleri_getir(
        db, md.IhracatUrunFiltresi(eksik="OLCUSUZ")
    )
    assert {u.urun_kodu for u in olcusuz} == {"OLCUSUZ"}

    grupsuz = md.ihracat_urunleri_getir(db, md.IhracatUrunFiltresi(eksik="GRUP"))
    assert {u.urun_kodu for u in grupsuz} == {"DESILI"}

    gruplu = md.ihracat_urunleri_getir(
        db, md.IhracatUrunFiltresi(urun_grubu="Radiator")
    )
    assert {u.urun_kodu for u in gruplu} == {"TAM"}


def test_ihracat_urunu_ekrandan_guncellenir(db, ihracat_urunler):
    from app.models import IhracatUrunu

    md.ihracat_urununu_guncelle(
        db,
        "OLCUSUZ",
        {"urun_adi": "Valve", "urun_grubu": "Acc.", "tir_yukleme_adeti": "2800",
         "konteyner_yukleme_adeti": "2464", "desi": "0,4", "en": "60", "boy": "120"},
    )
    urun = db.query(IhracatUrunu).filter_by(urun_kodu="OLCUSUZ").one()
    assert urun.tir_yukleme_adeti == Decimal(2800)
    assert urun.desi == Decimal("0.4")
    assert (urun.en, urun.boy) == (60, 120)


def test_olcusuz_ihracat_urunu_kaydedilemez(db):
    with pytest.raises(md.MasterDataHatasi, match="en az biri"):
        md.ihracat_urununu_guncelle(
            db, "YENI",
            {"urun_adi": "X", "tir_yukleme_adeti": "", "konteyner_yukleme_adeti": "",
             "desi": ""},
        )
