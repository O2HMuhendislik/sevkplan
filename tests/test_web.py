"""Web katmanı duman testi: ekranlar açılıyor ve uçtan uca akış çalışıyor mu?"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models
from app.db import oturum_bagimliligi
from app.main import uygulama


@pytest.fixture()
def fabrika(tmp_path):
    motor = create_engine(f"sqlite:///{tmp_path/'test.db'}", future=True)
    models.Temel.metadata.create_all(motor)
    return sessionmaker(bind=motor, autoflush=False, expire_on_commit=False)


@pytest.fixture()
def ham_istemci(fabrika, monkeypatch):
    """Giriş yapılmamış istemci. Başlangıçta yalnızca yönetici hesabı vardır."""
    from app import main as ana

    monkeypatch.setattr(ana, "OturumFabrikasi", fabrika)
    # Gömülü master data (2.800 ürün, 5.100 müşteri) her web testinde yüklenmesin;
    # yüklendiğini tests/test_ihracat.py doğrudan doğruluyor.
    monkeypatch.setattr(ana.gomulu_veri, "eksikleri_yukle", lambda db: [])

    def oturum_ver():
        db = fabrika()
        try:
            yield db
        finally:
            db.close()

    uygulama.dependency_overrides[oturum_bagimliligi] = oturum_ver
    with TestClient(uygulama) as istemci:
        yield istemci
    uygulama.dependency_overrides.clear()


@pytest.fixture()
def istemci(ham_istemci, fabrika):
    """Yönetici olarak giriş yapmış, parolası belirlenmiş istemci."""
    from app.services import kullanici_servisi

    with fabrika() as db:
        yonetici = kullanici_servisi.kullanici_getir(db, "admin")
        yonetici.parola_ozeti = __import__(
            "app.guvenlik", fromlist=["parola_ozeti"]
        ).parola_ozeti("Yonetici2026!")
        yonetici.parola_degistirmeli = False
        db.commit()
    cevap = ham_istemci.post(
        "/giris", data={"kullanici_adi": "admin", "parola": "Yonetici2026!"}
    )
    assert cevap.status_code == 200
    return ham_istemci


def sorgu(cevap) -> str:
    """Yönlendirme sonrası URL'deki mesaj/hata parametresini okunur metne çevirir."""
    from urllib.parse import unquote_plus

    ham = cevap.url.query
    return unquote_plus(ham.decode() if isinstance(ham, bytes) else ham)


def kitap(basliklar, satirlar) -> BytesIO:
    calisma_kitabi = Workbook()
    sayfa = calisma_kitabi.active
    sayfa.append(basliklar)
    for satir in satirlar:
        sayfa.append(satir)
    tampon = BytesIO()
    calisma_kitabi.save(tampon)
    tampon.seek(0)
    return tampon


@pytest.mark.parametrize(
    "yol",
    [
        "/", "/masterdata/urunler", "/ring", "/ring/siparisler", "/ring/planlar",
        "/ring/raporlar", "/raporlama/izleme", "/veri-yonetimi", "/yonetim/kullanicilar",
        "/rota", "/rota/siparisler", "/rota/planlar", "/masterdata/musteriler",
        "/rota/raporlar",
        "/ring/manuel-plan", "/rota/manuel-plan", "/ihracat/manuel-plan",
        "/raporlama", "/raporlama/siparisler", "/raporlama/planlar",
        "/ihracat/raporlar",
        "/ihracat", "/ihracat/siparisler", "/ihracat/planlar", "/masterdata/ihracat-musteriler",
    ],
)
def test_ekranlar_acilir(istemci, yol):
    cevap = istemci.get(yol)
    assert cevap.status_code == 200
    assert "Nakliye Yönetim Sistemi" in cevap.text


@pytest.mark.parametrize(
    "yol", ["/", "/masterdata/urunler", "/ring/planlar", "/yonetim/kullanicilar"]
)
def test_giris_yapmadan_erisilemez(ham_istemci, yol):
    cevap = ham_istemci.get(yol, follow_redirects=False)
    assert cevap.status_code == 303
    assert "/giris" in cevap.headers["location"]


def test_hatali_giris_reddedilir(ham_istemci):
    cevap = ham_istemci.post(
        "/giris", data={"kullanici_adi": "admin", "parola": "yanlis"}
    )
    assert "hatalı" in sorgu(cevap)


def test_ilk_giriste_parola_degistirme_zorunlu(ham_istemci, fabrika):
    from app.guvenlik import parola_ozeti
    from app.services import kullanici_servisi

    with fabrika() as db:
        yonetici = kullanici_servisi.kullanici_getir(db, "admin")
        yonetici.parola_ozeti = parola_ozeti("Gecici2026!")
        db.commit()
    ham_istemci.post("/giris", data={"kullanici_adi": "admin", "parola": "Gecici2026!"})
    cevap = ham_istemci.get("/ring/planlar", follow_redirects=False)
    assert cevap.status_code == 303
    assert cevap.headers["location"] == "/sifre-degistir"


def test_yetkisiz_kullanici_modulu_goremez(ham_istemci, fabrika):
    from app.guvenlik import parola_ozeti
    from app.models import Rol
    from app.services import kullanici_servisi

    with fabrika() as db:
        kullanici, _ = kullanici_servisi.kullanici_olustur(
            db, "depo", "Depo Görevlisi", Rol.DEPO, parola="DepoParola1!"
        )
        kullanici.parola_degistirmeli = False
        db.commit()
    ham_istemci.post("/giris", data={"kullanici_adi": "depo", "parola": "DepoParola1!"})
    assert ham_istemci.get("/ring/planlar").status_code == 403
    assert ham_istemci.get("/yonetim/kullanicilar").status_code == 403
    # Modül seçim ekranı herkese açık; yetkisi olmayan modül "yetkiniz yok" gösterir.
    assert "YETKİNİZ YOK" in ham_istemci.get("/").text


def test_goruntuleme_yetkisi_duzenlemeye_izin_vermez(ham_istemci, fabrika):
    from app.guvenlik import parola_ozeti
    from app.models import Rol
    from app.services import kullanici_servisi

    with fabrika() as db:
        kullanici, _ = kullanici_servisi.kullanici_olustur(
            db, "izleyici", "İzleyici", Rol.IZLEYICI, parola="IzleyiciP1!"
        )
        kullanici.parola_degistirmeli = False
        kullanici_servisi.yetkileri_ayarla(db, kullanici, {"RING": "GORUNTULE"})
        db.commit()
    ham_istemci.post(
        "/giris", data={"kullanici_adi": "izleyici", "parola": "IzleyiciP1!"}
    )
    assert ham_istemci.get("/ring/planlar").status_code == 200
    assert ham_istemci.post("/ring/planlar/uret", data={"depo_kodu": "64"}).status_code == 403


def test_uctan_uca_akis(istemci):
    urun_dosyasi = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["KMB-24", "Kombi 24 kW", "KOMBİ", 10, 100]],
    )
    cevap = istemci.post(
        "/masterdata/urunler/yukle", files={"dosya": ("urunler.xlsx", urun_dosyasi)}
    )
    assert cevap.status_code == 200 and "1 yeni" in sorgu(cevap)

    siparis_dosyasi = kitap(
        ["Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu", "Termin Tarihi"],
        [[f"SIP-{i}", f"TSL-{i}", "KMB-24", 25, "64", "05.09.2026"] for i in range(4)],
    )
    istemci.post("/ring/siparisler/yukle", files={"dosya": ("siparis.xlsx", siparis_dosyasi)})

    cevap = istemci.post(
        "/ring/planlar/uret", data={"plan_tarihi": "2026-08-31", "depo_kodu": "64"}
    )
    assert "1 plan üretildi" in sorgu(cevap)

    planlar = istemci.get("/ring/planlar")
    assert "2608D1001" in planlar.text

    detay = istemci.get("/ring/planlar/1")
    assert detay.status_code == 200
    assert "2608D1001" in detay.text

    # Axata numarası girilmeden mail gönderilemez.
    cevap = istemci.post("/ring/planlar/1/mail")
    assert "Axata" in sorgu(cevap)

    istemci.post("/ring/planlar/1/axata", data={"axata_no": "AX-5501"})
    cevap = istemci.post("/ring/planlar/1/mail")
    assert "hata=" not in sorgu(cevap)

    form = istemci.get("/ring/planlar/1/form")
    assert form.status_code == 200
    assert form.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )

    gunluk = istemci.get("/ring/gunluk-form", params={"tarih": "2026-08-31"})
    assert gunluk.status_code == 200

    izleme = istemci.get("/raporlama/izleme", params={"anahtar": "TSL-1"})
    assert "2608D1001" in izleme.text

    # Ekran Raporlama'ya taşındı; eski adres yer imlerinde kalmış olabilir.
    eski = istemci.get("/ring/izleme", params={"anahtar": "TSL-1"})
    assert "2608D1001" in eski.text


def test_veri_silme_onay_ister(istemci):
    cevap = istemci.post("/veri-yonetimi/sil", data={"islem": "bekleyen", "onay": "evet"})
    assert "SIL yazmalısınız" in sorgu(cevap)


def test_veri_silme_onayla_calisir(istemci):
    urun_dosyasi = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["KMB-24", "Kombi 24 kW", "KOMBİ", 10, 100]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("urunler.xlsx", urun_dosyasi)})
    siparis_dosyasi = kitap(
        ["Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu", "Termin Tarihi"],
        [["SIP-1", "TSL-1", "KMB-24", 25, "64", "05.09.2026"]],
    )
    istemci.post("/ring/siparisler/yukle", files={"dosya": ("siparis.xlsx", siparis_dosyasi)})

    cevap = istemci.post(
        "/veri-yonetimi/sil", data={"islem": "bekleyen", "onay": "SIL"}
    )
    assert "1 sipariş satırı" in sorgu(cevap)
    assert "0" in istemci.get("/veri-yonetimi").text


# --------------------------------------------------- İç piyasa modülü uçtan uca


def ring_verisi_yukle(istemci):
    """Ring havuzuna tek depolu (64) bir plan üretir."""
    urunler = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["U1", "Kombi A", "KOMBİ", 10, 100]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("urun.xlsx", urunler)})
    siparisler = kitap(
        ["Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu", "SehirAdi",
         "BayiAdi"],
        [[f"S{i}", f"T{i}", "U1", 20, "64", "ESKİŞEHİR", "BAYİ A"] for i in range(5)],
    )
    istemci.post("/ring/siparisler/yukle", files={"dosya": ("s.xlsx", siparisler)})
    istemci.post("/ring/planlar/uret", data={"depo_kodu": "64"})


def ic_piyasa_verisi_yukle(istemci):
    """Ürün + müşteri master datası ve iki müşterili bir sipariş dosyası yükler."""
    urunler = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti",
         "Ürün Desi"],
        [
            ["U1", "Kombi A", "KOMBİ", 10, 100, 12],
            ["U2", "Panel B", "PANEL", 10, 100, 12],
        ],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("urun.xlsx", urunler)})

    musteriler = kitap(
        ["Bayi Adı", "İl", "İlçe", "Tır Girişi (E/H/?)"],
        [
            ["EGE ISITMA", "İZMİR", "BORNOVA", "E"],
            ["MANİSA TESİSAT", "MANİSA", "MERKEZ", "H"],
        ],
    )
    istemci.post("/masterdata/musteriler/yukle", files={"dosya": ("m.xlsx", musteriler)})

    siparisler = kitap(
        ["Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu", "SehirAdi",
         "BayiAdi", "AliciFirma", "SevkAdresi", "Not"],
        [
            ["S1", "T1", "U1", 60, "64", "İZMİR", "EGE ISITMA", "EGE ISITMA A.Ş.",
             "1234 SOK. NO:5", "CIF - BORNOVA"],
            ["S2", "T2", "U2", 40, "74", "MANİSA", "MANİSA TESİSAT",
             "MANİSA TESİSAT LTD.", "SANAYİ CAD. NO:8", "CIF - MERKEZ"],
        ],
    )
    # İki modül aynı havuzu kullanır; iç piyasa ekranından yüklemek de aynı sonucu verir.
    return istemci.post(
        "/rota/siparisler/yukle", files={"dosya": ("siparis.xlsx", siparisler)}
    )


def test_ic_piyasa_plani_uretilir_ve_formu_indirilir(istemci, fabrika):
    from app.models import SevkiyatPlani

    ic_piyasa_verisi_yukle(istemci)

    onizleme = istemci.get("/rota/siparisler")
    assert "EGE ISITMA" in onizleme.text

    cevap = istemci.post(
        "/rota/planlar/uret", data={"tipler": ["FTL"], "plan_tarihi": "2026-09-01"}
    )
    assert "plan üretildi" in sorgu(cevap)

    with fabrika() as db:
        plan = db.query(SevkiyatPlani).filter_by(modul="ROTA").one()
        # İzmir ve Manisa aynı bölgede (Ege), tek araca binerler. Duraklar
        # Eskişehir'e uzaklığa göre sıralanır: Manisa 350 km, İzmir 400 km.
        assert plan.sevkiyat_tipi == "FTL"
        assert plan.sefer_no[4] == "S"
        assert plan.durak_sayisi == 2
        assert plan.son_ugrak == "IZMIR"
        assert plan.iller_metni == "MANISA, IZMIR"
        assert plan.ilce_metni == "MERKEZ+BORNOVA"
        # Araçtaki her il yazılır, sonuna toplam durak: "MANISA+IZMIR2YER".
        assert plan.il_yeri_metni == "MANISA+IZMIR2YER"
        # 64 hacmin çoğunu taşıyor; 74'teki mal oraya getirilecek.
        assert plan.yukleme_deposu == "64"
        plan_id = plan.id

    detay = istemci.get(f"/rota/planlar/{plan_id}")
    assert "MANİSA TESİSAT" in detay.text
    assert "64 depoya gönderilmelidir" in detay.text
    assert "Tır girişi olmayan müşteri var" in detay.text

    # Planda 64 ve 74 depoları birlikte; numara depoya bağlanmalı.
    istemci.post(
        f"/rota/planlar/{plan_id}/axata",
        data={"axata_no": "3299, 3300", "depo_kodu": "64"},
    )
    istemci.post(
        f"/rota/planlar/{plan_id}/arac",
        data={"nakliyeci": "OMSAN", "plaka": "34 ABC 12", "surucu": "Ali Veli",
              "surucu_telefon": "555"},
    )

    form = istemci.get(f"/rota/planlar/{plan_id}/form")
    assert form.status_code == 200
    assert form.headers["content-type"].startswith("application/")

    gunluk = istemci.get("/rota/gunluk-form?tarih=2026-09-01")
    assert gunluk.status_code == 200


def test_ic_piyasa_formu_sevkiyat_tipine_gore_sayfalanir(istemci, fabrika, tmp_path):
    from openpyxl import load_workbook

    from app.models import SevkiyatPlani
    from app.services import ic_yukleme_formu

    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"plan_tarihi": "2026-09-01"})

    with fabrika() as db:
        planlar = db.query(SevkiyatPlani).filter_by(modul="ROTA").all()
        hedef = ic_yukleme_formu.formlari_uret(planlar, tmp_path / "form.xlsx")

    sayfa = load_workbook(hedef)["S-FTL Sevk"]
    metinler = [h.value for satir in sayfa.iter_rows() for h in satir if h.value]
    assert "Yer Miktarı" in metinler
    assert "MANISA+IZMIR2YER" in metinler
    assert "MERKEZ+BORNOVA" in metinler
    assert "Nak.Firma" in metinler
    assert "Yükleme yapacak depolar" in metinler
    assert any("depoya gönderilmelidir" in str(m) for m in metinler)
    # Kılavuz çizgileri kapalı ve form kalın çerçeve içinde.
    assert sayfa.sheet_view.showGridLines is False
    assert sayfa["A1"].border.left.style == "medium"


def test_musteri_ekranindan_tir_girisi_guncellenir(istemci, fabrika):
    from app.models import Musteri

    ic_piyasa_verisi_yukle(istemci)
    with fabrika() as db:
        musteri = db.query(Musteri).filter_by(bayi_adi="EGE ISITMA").one()
        musteri_id = musteri.id
        assert musteri.tir_girisi == "E"

    istemci.post(
        f"/masterdata/musteriler/{musteri_id}",
        data={"tir_girisi": "H", "bolge_kodu": "", "il": "İZMİR", "ilce": "BORNOVA",
              "notlar": "AVM içi", "aktif": "true"},
    )
    with fabrika() as db:
        musteri = db.get(Musteri, musteri_id)
        assert musteri.tir_girisi == "H"
        assert musteri.notlar == "AVM içi"
        assert musteri.il == "IZMIR"


def test_ic_piyasa_ekranindan_siparis_yuklenir(istemci, fabrika):
    """Yalnızca ROTA yetkisi olan kullanıcı Ring ekranını açamaz; yükleme burada da olmalı."""
    from app.models import SiparisSatiri

    cevap = ic_piyasa_verisi_yukle(istemci)
    assert "Sipariş aktarımı" in sorgu(cevap)

    with fabrika() as db:
        assert db.query(SiparisSatiri).count() == 2

    ekran = istemci.get("/rota/siparisler")
    assert "Sipariş dosyası yükle" in ekran.text
    assert istemci.get("/rota/siparisler/sablon").status_code == 200


def test_alinamayan_satirlar_gerekcesiyle_gosterilir(istemci, fabrika):
    """Master datada olmayan ürün planlamaya girmez; sebebi ekranda görünmeli."""
    ic_piyasa_verisi_yukle(istemci)
    tanimsiz = kitap(
        ["Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu", "SehirAdi",
         "BayiAdi"],
        [["S9", "T9", "YOK-99", 10, "64", "İZMİR", "EGE ISITMA"]],
    )
    istemci.post("/rota/siparisler/yukle", files={"dosya": ("s.xlsx", tanimsiz)})
    istemci.post("/rota/planlar/uret", data={"plan_tarihi": "2026-09-01"})

    ekran = istemci.get("/rota/siparisler")
    assert "Alınamayan satırlar" in ekran.text
    assert "YOK-99" in ekran.text
    assert "master datada tanımlı değil" in ekran.text


def test_baslikta_marka_ve_logo_var(istemci):
    """Başlık her ekranda kurumsal kimliği taşır; logo tek dosyadan gelir."""
    cevap = istemci.get("/")
    assert "Vaillant Group" in cevap.text
    assert "Nakliye Yönetim Sistemi" in cevap.text
    assert "/marka/logo" in cevap.text
    # Logo yüklenmemişken depodaki yer tutucu döner.
    logo = istemci.get("/marka/logo")
    assert logo.status_code == 200
    assert logo.headers["content-type"].startswith("image/svg")


def test_giris_ekraninda_da_marka_gorunur(ham_istemci):
    cevap = ham_istemci.get("/giris")
    assert "Nakliye Yönetim Sistemi" in cevap.text
    assert "/marka/logo" in cevap.text
    # Giriş ekranında oturum yok; logo yine de açılmalı.
    assert ham_istemci.get("/marka/logo").status_code == 200


def test_resmi_logo_ekrandan_yuklenir(istemci, tmp_path, monkeypatch):
    """Resmî logo markadır; depoda yer tutucu durur, gerçeği ekrandan yüklenir."""
    from app.services import marka

    monkeypatch.setattr(marka, "LOGO_DIZINI", tmp_path / "marka")

    tek_piksel = bytes.fromhex(
        "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
        "890000000a49444154789c6360000002000100ffff03000006000557bfabd400"
        "00000049454e44ae426082"
    )
    cevap = istemci.post(
        "/yonetim/logo", files={"dosya": ("vaillant.png", tek_piksel, "image/png")}
    )
    assert "Logo güncellendi" in sorgu(cevap)

    logo = istemci.get("/marka/logo")
    assert logo.headers["content-type"] == "image/png"
    assert logo.content == tek_piksel

    # Kaldırılınca yer tutucuya dönülür.
    cevap = istemci.post("/yonetim/logo/sil")
    assert "yer tutucuya" in sorgu(cevap)
    assert istemci.get("/marka/logo").headers["content-type"].startswith("image/svg")


def test_gecersiz_logo_dosyasi_reddedilir(istemci, tmp_path, monkeypatch):
    from app.services import marka

    monkeypatch.setattr(marka, "LOGO_DIZINI", tmp_path / "marka")

    cevap = istemci.post(
        "/yonetim/logo", files={"dosya": ("logo.txt", b"merhaba", "text/plain")}
    )
    assert "Desteklenmeyen dosya türü" in sorgu(cevap)
    assert marka.yuklenen_logo() is None


def test_moduller_ayri_siparis_havuzu_kullanir(istemci, fabrika):
    """İç piyasadan yüklenen sipariş Ring ekranında görünmemeli."""
    from app.models import SiparisSatiri

    ic_piyasa_verisi_yukle(istemci)

    ring = istemci.get("/ring/siparisler")
    assert "EGE ISITMA" not in ring.text

    rota = istemci.get("/rota/siparisler")
    assert "EGE ISITMA" in rota.text

    # Raporlama ekranı hepsini bir arada gösterir.
    hepsi = istemci.get("/raporlama/siparisler")
    assert "EGE ISITMA" in hepsi.text
    assert istemci.get("/raporlama/siparisler?modul=RING").text.count("EGE ISITMA") == 0

    with fabrika() as db:
        assert {s.modul for s in db.query(SiparisSatiri).all()} == {"ROTA"}


def test_ring_planlamasi_ic_piyasa_siparisini_almaz(istemci, fabrika):
    from app.models import SevkiyatPlani

    ic_piyasa_verisi_yukle(istemci)
    cevap = istemci.post("/ring/planlar/uret", data={"depo_kodu": "64"})
    assert "Plan üretilemedi" in sorgu(cevap)

    with fabrika() as db:
        assert db.query(SevkiyatPlani).count() == 0


def test_plana_alinma_kpisi_hesaplanir(istemci, fabrika):
    """Sipariş sisteme girdikten kaç gün sonra plana alındı?"""
    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"plan_tarihi": "2026-09-01"})

    ekran = istemci.get("/raporlama")
    assert "Plana alınma süresi" in ekran.text
    assert "İç Piyasa" in ekran.text

    with fabrika() as db:
        from app.services import rapor_servisi

        kpiler = {k.modul: k for k in rapor_servisi.planlama_kpi(db)}
        kpi = kpiler["ROTA"]
        assert kpi.planlanan == 2
        # Aynı gün yüklenip aynı gün planlandı.
        assert kpi.ortalama_gun == 0
        assert kpi.ayni_gun == 2


def ihracat_verisi_yukle(istemci):
    """İhracat müşteri master datası ve iki müşterili bir sipariş dosyası yükler."""
    musteriler = kitap(
        ["Müşteri Adı", "Ülke", "Ülke Kodu", "Araç Tipi", "Sefer Kodu", "Yükleme Tipi",
         "Azami Tonaj", "Açıklama"],
        [
            ["VAILLANT D.O.O.", "HIRVATİSTAN", "HR", "TIR", "NSC", "STANDART",
             "22.000 KG", ""],
            ["ANSAL REFRIGERACION SA", "ARJANTİN", "AR", "KONTEYNER", "Export",
             "PALET YÜKSELTME", "19.500 KG", "silika jel konulacak"],
        ],
    )
    istemci.post("/masterdata/ihracat-musteriler/yukle", files={"dosya": ("m.xlsx", musteriler)})

    siparisler = kitap(
        ["DEPO", "ÜLKE KODU", "SİPARİŞ NO", "ÜRÜN KODU", "ÜRÜN TANIMI", "ADET",
         "MÜŞTERİ ADI", "SEVK ADRESİ", "TESLİMAT NO", "Desi", "KG", "ÜLKE"],
        [
            ["34", "HR", "9002842146", "916041211", "Panel 600", 100,
             "VAILLANT D.O.O.", "OSIJEK", "9106800933", 20000, 15000, "HIRVATİSTAN"],
            ["34", "AR", "9002842200", "916101211", "Panel 1000", 80,
             "ANSAL REFRIGERACION SA", "BUENOS AIRES", "9106800940", 14000, 12000,
             "ARJANTİN"],
        ],
    )
    return istemci.post(
        "/ihracat/siparisler/yukle", files={"dosya": ("ihracat.xlsx", siparisler)}
    )


def test_ihracat_plani_uretilir_ve_formu_indirilir(istemci, fabrika):
    from app.models import SevkiyatPlani

    cevap = ihracat_verisi_yukle(istemci)
    assert "Sipariş aktarımı" in sorgu(cevap)

    onizleme = istemci.get("/ihracat/siparisler")
    assert "VAILLANT D.O.O." in onizleme.text
    assert "DENİZ" in onizleme.text  # Arjantin konteyner ile gider

    cevap = istemci.post("/ihracat/planlar/uret", data={"plan_tarihi": "2026-09-01"})
    assert "araç planlandı" in sorgu(cevap)

    with fabrika() as db:
        planlar = {
            p.musteri_adi: p
            for p in db.query(SevkiyatPlani).filter_by(modul="IHRACAT").all()
        }
        hirvat = planlar["VAILLANT D.O.O."]
        arjantin = planlar["ANSAL REFRIGERACION SA"]
        # Sefer belge kodu müşteriden gelir: NSC -> N, Export -> E.
        assert hirvat.sefer_no[4] == "N"
        assert arjantin.sefer_no[4] == "E"
        # Araç tipi taşıma modunu belirler.
        assert hirvat.tasima_modu == "KARA"
        assert arjantin.tasima_modu == "DENİZ"
        assert arjantin.musteri_aciklamasi == "silika jel konulacak"
        plan_id = arjantin.id

    detay = istemci.get(f"/ihracat/planlar/{plan_id}")
    assert "silika jel konulacak" in detay.text
    assert "DENİZ" in detay.text

    istemci.post(f"/ihracat/planlar/{plan_id}/axata", data={"axata_no": "2735"})
    istemci.post(
        f"/ihracat/planlar/{plan_id}/arac",
        data={"nakliyeci": "OMSAN", "plaka": "34 ABC 12",
              "konteyner_no": "MSCU1234567", "muhur_no": "M-9", "surucu": "Ali"},
    )
    assert istemci.get(f"/ihracat/planlar/{plan_id}/form").status_code == 200
    assert istemci.get("/ihracat/gunluk-form?tarih=2026-09-01").status_code == 200


def test_uc_modul_ayri_havuz_kullanir(istemci, fabrika):
    """Ring, iç piyasa ve ihracat siparişleri birbirinin ekranında görünmez."""
    from app.models import SiparisSatiri

    ic_piyasa_verisi_yukle(istemci)
    ihracat_verisi_yukle(istemci)

    assert "VAILLANT D.O.O." not in istemci.get("/rota/siparisler").text
    assert "EGE ISITMA" not in istemci.get("/ihracat/siparisler").text
    assert "EGE ISITMA" not in istemci.get("/ring/siparisler").text

    hepsi = istemci.get("/raporlama/siparisler").text
    assert "EGE ISITMA" in hepsi and "VAILLANT D.O.O." in hepsi

    with fabrika() as db:
        assert {s.modul for s in db.query(SiparisSatiri).all()} == {"ROTA", "IHRACAT"}


def test_axatali_plan_silinebilir(istemci, fabrika):
    """Plana Axata numarası girilmişse silme yabancı anahtar hatası vermemeli.

    Toplu `delete()` ORM ilişkilerini çalıştırmadığı için plana bağlı Axata
    satırları elle silinmezse veritabanı planı silmeyi reddeder ve ekran
    "Internal Server Error" gösterir.
    """
    from app.models import AxataNumarasi, SevkiyatPlani

    ihracat_verisi_yukle(istemci)
    istemci.post("/ihracat/planlar/uret", data={"plan_tarihi": "2026-09-01"})
    with fabrika() as db:
        plan_id = db.query(SevkiyatPlani).filter_by(modul="IHRACAT").first().id
    istemci.post(f"/ihracat/planlar/{plan_id}/axata", data={"axata_no": "2735"})
    with fabrika() as db:
        assert db.query(AxataNumarasi).count() == 1

    cevap = istemci.post(
        "/veri-yonetimi/sil", data={"islem": "siparis_ve_planlar", "onay": "SIL"}
    )
    assert cevap.status_code == 200
    assert "silindi" in sorgu(cevap)
    with fabrika() as db:
        assert db.query(SevkiyatPlani).count() == 0
        assert db.query(AxataNumarasi).count() == 0


def test_her_seyi_sil_master_datayi_da_temizler(istemci, fabrika):
    """"Her şeyi sil" ihracat ürün ve müşteri master datasını da kapsamalı."""
    from app.models import IhracatMusterisi, IhracatUrunu, SiparisSatiri

    ihracat_verisi_yukle(istemci)
    with fabrika() as db:
        assert db.query(IhracatMusterisi).count() > 0

    cevap = istemci.post("/veri-yonetimi/sil", data={"islem": "hepsi", "onay": "SIL"})
    assert cevap.status_code == 200
    with fabrika() as db:
        assert db.query(SiparisSatiri).count() == 0
        assert db.query(IhracatMusterisi).count() == 0
        assert db.query(IhracatUrunu).count() == 0


def test_ring_eskisehir_disi_siparisi_almaz(istemci, fabrika):
    """Ring Eskişehir içi dağıtımdır; başka il yüklenirse alınmaz ve kullanıcı uyarılır."""
    from app.models import SiparisSatiri

    urunler = kitap(
        ["StokKodu", "StokAdi", "Palet içi adet", "Tır yükleme adeti"],
        [["U1", "Kombi A", 10, 100]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("urun.xlsx", urunler)})

    dosya = kitap(
        ["SehirAdi", "Sipariş No", "Teslimat No", "StokKodu", "Adet", "Depo  Kodu"],
        [
            ["ESKİŞEHİR", "S1", "T1", "U1", 100, "64"],
            ["İZMİR", "S2", "T2", "U1", 100, "64"],
            ["", "S3", "T3", "U1", 100, "64"],
        ],
    )
    cevap = istemci.post("/ring/siparisler/yukle", files={"dosya": ("s.xlsx", dosya)})
    assert "modül kapsamı dışında" in sorgu(cevap)

    with fabrika() as db:
        siparisler = {s.siparis_no for s in db.query(SiparisSatiri).all()}
    # Eskişehir ve şehri boş satır alınır; İzmir alınmaz.
    assert siparisler == {"S1", "S3"}


def test_ring_raporu_baska_modulun_bekleyenini_gostermez(istemci, fabrika):
    """Bekleyen listesi modüle göre daralır; yoksa 'esneme çalışmıyor' sanılıyor."""
    ihracat_verisi_yukle(istemci)

    cevap = istemci.get("/ring/raporlar")
    assert cevap.status_code == 200
    # İhracat havuzundaki bekleyen ürün Ring raporunda görünmemeli.
    assert "916041211" not in cevap.text


def test_plan_raporu_ozet_urun_grubu_ve_sevk_durumu_uretir(istemci, fabrika, tmp_path):
    from openpyxl import load_workbook

    ihracat_verisi_yukle(istemci)
    istemci.post("/ihracat/planlar/uret", data={"plan_tarihi": "2026-09-01"})

    cevap = istemci.get("/rapor/plan-raporu?modul=IHRACAT")
    assert cevap.status_code == 200
    hedef = tmp_path / "rapor.xlsx"
    hedef.write_bytes(cevap.content)

    kitap_ = load_workbook(hedef)
    assert kitap_.sheetnames == ["Özet", "Ürün Grubu", "Planlar", "Sevk Durumu"]
    ozet = [h.value for satir in kitap_["Özet"].iter_rows() for h in satir if h.value]
    assert "İhracat" in ozet and "TOPLAM" in ozet
    assert kitap_["Ürün Grubu"].max_row > 1
    planlar = kitap_["Planlar"]
    assert planlar["A1"].value == "#" and planlar["A2"].value == 1
    assert "Konteyner No" in [h.value for h in kitap_["Sevk Durumu"][1]]


def _bekleyen_sayaci(sayfa: str) -> int:
    """Gösterge panelindeki "Bekleyen sipariş satırı" metriğinin değeri."""
    import re

    eslesme = re.search(
        r"Bekleyen sipariş satırı</div>\s*<div class=\"deger\">(\d+)<", sayfa
    )
    assert eslesme, "Bekleyen sipariş satırı metriği bulunamadı"
    return int(eslesme.group(1))


def test_baska_modulun_siparisi_ring_sayacinda_gorunmez(istemci, fabrika):
    """Havuzlar ayrı: iç piyasaya yüklenen sipariş Ring gösterge panelinde sayılmaz.

    Sayılırsa kullanıcı "planlanmayı bekleyen iş var ama listede yok" sanıyor.
    """
    from app.models import SiparisSatiri
    from app.services import rapor_servisi

    ic_piyasa_verisi_yukle(istemci)
    with fabrika() as db:
        assert db.query(SiparisSatiri).filter_by(modul="ROTA").count() > 0
        ring = rapor_servisi.gosterge_paneli(db, modul="RING")
        rota = rapor_servisi.gosterge_paneli(db, modul="ROTA")

    assert ring["siparis"] == {}
    assert rota["siparis"].get("BEKLEMEDE", 0) > 0

    # Ekranda da görünmemeli: Ring panelindeki sayaç sıfır, iç piyasada dolu.
    assert _bekleyen_sayaci(istemci.get("/ring").text) == 0
    assert _bekleyen_sayaci(istemci.get("/rota").text) > 0


def test_manuel_planlama_secilen_teslimati_planlar(istemci, fabrika):
    """Kullanıcı listeden teslimat seçip planlayabilmeli; seçilmeyen beklemede kalır."""
    from app.models import SevkiyatPlani, SiparisSatiri

    ic_piyasa_verisi_yukle(istemci)

    ekran = istemci.get("/rota/manuel-plan")
    assert ekran.status_code == 200
    # Liste teslimat bazında; iki teslimat da seçilebilir durumda.
    assert 'value="T1"' in ekran.text
    assert 'value="T2"' in ekran.text

    cevap = istemci.post(
        "/rota/manuel-plan/uret",
        data={
            "teslimat_nolar": ["T1"],
            "plan_tarihi": "2026-09-01",
            "kalanlari_zorla": "1",
        },
    )
    assert "Manuel planlama" in sorgu(cevap)

    with fabrika() as db:
        planlar = db.query(SevkiyatPlani).filter_by(modul="ROTA").all()
        assert len(planlar) == 1
        planli = {s.teslimat_no for s in planlar[0].satirlar}
        assert planli == {"T1"}
        # Seçilmeyen teslimat plana girmez, beklemede kalır.
        t2 = db.query(SiparisSatiri).filter_by(teslimat_no="T2").one()
        assert t2.plan_id is None


def test_manuel_planlamada_secim_yoksa_uyari_verilir(istemci):
    ic_piyasa_verisi_yukle(istemci)
    cevap = istemci.post("/rota/manuel-plan/uret", data={"plan_tarihi": "2026-09-01"})
    assert "teslimat seçilmedi" in sorgu(cevap)


def test_manuel_planlama_planlanmis_teslimati_listelemez(istemci, fabrika):
    """Plana giren teslimat listeden düşer; iki kez planlanamaz."""
    ic_piyasa_verisi_yukle(istemci)
    istemci.post(
        "/rota/manuel-plan/uret",
        data={"teslimat_nolar": ["T1"], "kalanlari_zorla": "1"},
    )
    ekran = istemci.get("/rota/manuel-plan")
    assert 'value="T1"' not in ekran.text
    assert 'value="T2"' in ekran.text


def test_manuel_planlama_moduller_arasi_sizmaz(istemci):
    """İç piyasa siparişi Ring'in manuel planlama listesinde görünmemeli."""
    ic_piyasa_verisi_yukle(istemci)
    ring = istemci.get("/ring/manuel-plan")
    assert 'value="T1"' not in ring.text
    assert 'value="T2"' not in ring.text


def test_manuel_planlama_aramayla_daraltilir(istemci):
    ic_piyasa_verisi_yukle(istemci)
    cevap = istemci.get("/rota/manuel-plan", params={"arama": "MANİSA"})
    assert 'value="T2"' in cevap.text
    assert 'value="T1"' not in cevap.text


def test_cok_depolu_planda_axata_deposu_secilmeden_girilemez(istemci, fabrika):
    """64 + 74 planında numara hangi depoya ait belli olmalı.

    Aksi hâlde yükleme formunda numara hangi depo satırına yazılacak bilinmiyor ve
    depo yanlış iş emriyle toplama yapıyor.
    """
    from app.models import SevkiyatPlani

    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"tipler": ["FTL"]})
    with fabrika() as db:
        plan = db.query(SevkiyatPlani).filter_by(modul="ROTA").one()
        assert plan.axata_depolari == ["64", "74"]
        plan_id = plan.id

    cevap = istemci.post(f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3299"})
    assert "hangi depoya ait olduğu" in sorgu(cevap)
    with fabrika() as db:
        assert not db.get(SevkiyatPlani, plan_id).axata_numaralari

    # Planda olmayan depo da reddedilir; numara formda hiçbir satıra düşmezdi.
    hatali = istemci.post(
        f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3299", "depo_kodu": "34"}
    )
    assert "34 deposu bu planda yok" in sorgu(hatali)

    istemci.post(
        f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3299", "depo_kodu": "64"}
    )
    istemci.post(
        f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3400", "depo_kodu": "74"}
    )
    with fabrika() as db:
        plan = db.get(SevkiyatPlani, plan_id)
        assert plan.depo_axata_ozeti("64") == "3299"
        assert plan.depo_axata_ozeti("74") == "3400"
        assert plan.axata_ozeti == "64: 3299, 74: 3400"
        assert plan.axatasiz_depolar == []


def test_yukleme_formunda_her_axata_kendi_depo_satirina_yazilir(istemci, fabrika, tmp_path):
    from openpyxl import load_workbook

    from app.models import SevkiyatPlani

    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"tipler": ["FTL"]})
    with fabrika() as db:
        plan_id = db.query(SevkiyatPlani).filter_by(modul="ROTA").one().id
    istemci.post(
        f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3299", "depo_kodu": "64"}
    )
    istemci.post(
        f"/rota/planlar/{plan_id}/axata", data={"axata_no": "3400", "depo_kodu": "74"}
    )

    dosya = tmp_path / "form.xlsx"
    dosya.write_bytes(istemci.get(f"/rota/planlar/{plan_id}/form").content)
    sayfa = load_workbook(dosya).active

    kutu = {}
    for satir in sayfa.iter_rows(min_col=5, max_col=6, values_only=True):
        if satir[0] and str(satir[0]).endswith("DEPO"):
            kutu[str(satir[0])] = satir[1]
    assert kutu["64-D DEPO"] == "3299"
    assert kutu["74-DEPO"] == "3400"
    assert not kutu["34-DEPO"]


def test_tek_depolu_planda_axata_deposu_zorunlu_degil(istemci, fabrika):
    """Ring planlarında tek depo var; kullanıcı her seferinde depo seçmek zorunda kalmaz."""
    from app.models import SevkiyatPlani

    ring_verisi_yukle(istemci)
    with fabrika() as db:
        plan = db.query(SevkiyatPlani).filter_by(modul="RING").first()
        assert plan.axata_depolari == ["64"]
        assert plan.cok_depolu_mu is False
        plan_id = plan.id

    cevap = istemci.post(f"/ring/planlar/{plan_id}/axata", data={"axata_no": "AX-1"})
    assert "Axata numaraları" in sorgu(cevap)
    with fabrika() as db:
        plan = db.get(SevkiyatPlani, plan_id)
        # Deposu boş numara bütün depolar için geçerlidir; formda plan deposuna yazılır.
        assert plan.axata_numaralari[0].depo_kodu is None
        assert plan.depo_axata_ozeti("64") == "AX-1"



@pytest.mark.parametrize(
    "yol",
    ["/masterdata", "/masterdata/urunler", "/masterdata/musteriler",
     "/masterdata/ihracat-musteriler", "/masterdata/ihracat-urunler",
     "/masterdata/gruplar", "/masterdata/depolar", "/masterdata/sistem"],
)
def test_masterdata_ekranlari_acilir(istemci, yol):
    cevap = istemci.get(yol)
    assert cevap.status_code == 200
    assert "Nakliye Yönetim Sistemi" in cevap.text


def test_eski_masterdata_adresleri_yonlendirilir(istemci):
    """Kayıtlı bağlantılar ve yer imleri kırılmasın."""
    for eski, yeni in (
        ("/urunler", "/masterdata/urunler"),
        ("/rota/musteriler", "/masterdata/musteriler"),
        ("/ihracat/musteriler", "/masterdata/ihracat-musteriler"),
        ("/ihracat/urunler", "/masterdata/ihracat-urunler"),
    ):
        cevap = istemci.get(eski, follow_redirects=False)
        assert cevap.status_code == 308
        assert cevap.headers["location"] == yeni


def test_filtrelenen_liste_indirilir(istemci, fabrika, tmp_path):
    """Ekranda görülen filtre ile inen dosya aynı kayıtları içerir."""
    from openpyxl import load_workbook

    urunler = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["P1", "Panel 1", "PANEL", 10, 100],
         ["P2", "Panel 2", "PANEL", None, 100],
         ["K1", "Kombi 1", "KOMBİ", 15, 360]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("u.xlsx", urunler)})

    ekran = istemci.get("/masterdata/urunler", params={"urun_grubu": "PANEL"})
    assert "P1" in ekran.text and "K1" not in ekran.text

    dosya = tmp_path / "filtre.xlsx"
    dosya.write_bytes(
        istemci.get("/masterdata/urunler/excel", params={"urun_grubu": "PANEL"}).content
    )
    sayfa = load_workbook(dosya).active
    kodlar = {sayfa.cell(row=r, column=1).value for r in range(2, sayfa.max_row + 1)}
    assert kodlar == {"P1", "P2"}

    # Eksik alan filtresi de dosyaya uygulanır.
    dosya2 = tmp_path / "eksik.xlsx"
    dosya2.write_bytes(
        istemci.get("/masterdata/urunler/excel", params={"eksik": "PALET"}).content
    )
    sayfa2 = load_workbook(dosya2).active
    kodlar2 = {sayfa2.cell(row=r, column=1).value for r in range(2, sayfa2.max_row + 1)}
    assert kodlar2 == {"P2"}


def test_urun_ekrandan_guncellenir(istemci, fabrika):
    from app.models import Urun

    urunler = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["P1", "Panel 1", "PANEL", 10, 100]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("u.xlsx", urunler)})

    assert istemci.get("/masterdata/urunler/P1").status_code == 200
    cevap = istemci.post(
        "/masterdata/urunler/kaydet",
        data={"geri": "/masterdata/urunler", "urun_kodu": "P1", "urun_adi": "Panel 1",
              "urun_grubu": "PANEL", "palet_ici_adet": "32", "tir_yukleme_adeti": "704",
              "tir_palet": "22", "agirlik": "34,5", "palet_en": "105",
              "palet_boy": "123", "aktif": "E"},
    )
    assert "güncellendi" in sorgu(cevap)
    with fabrika() as db:
        urun = db.query(Urun).filter_by(urun_kodu="P1").one()
        assert (urun.palet_ici_adet, urun.tir_palet, urun.palet_en) == (32, 22, 105)
        assert str(urun.agirlik) == "34.500"


def test_sistem_ayari_planlama_kurallarina_gecer(istemci, fabrika):
    from app.services import masterdata_servisi

    cevap = istemci.post("/masterdata/sistem", data={"kargo_desi_siniri": "25"})
    assert "Kargo desi sınırı: 10 → 25" in sorgu(cevap)
    with fabrika() as db:
        assert masterdata_servisi.kurallari_kur(db).kargo_desi_siniri == Decimal(25)

    istemci.post("/masterdata/sistem", data={"varsayilana_don": "1"})
    with fabrika() as db:
        assert masterdata_servisi.kurallari_kur(db).kargo_desi_siniri == Decimal(10)


def test_yerlesim_plani_ekrani_acilir(istemci, fabrika):
    """Plan detayından yerleşim planına geçilir; ekran duraklarla birlikte gelir."""
    from app.models import SevkiyatPlani

    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"tipler": ["FTL"]})
    with fabrika() as db:
        plan_id = db.query(SevkiyatPlani).filter_by(modul="ROTA").one().id

    detay = istemci.get(f"/rota/planlar/{plan_id}")
    assert f"/rota/planlar/{plan_id}/yerlesim" in detay.text

    cevap = istemci.get(f"/rota/planlar/{plan_id}/yerlesim")
    assert cevap.status_code == 200
    assert "Araç İçi Yerleşim" in cevap.text
    assert "KABİN — önce yüklenir" in cevap.text
    # İki müşteri de durak tablosunda görünür.
    assert "EGE ISITMA" in cevap.text and "MANİSA TESİSAT" in cevap.text
    # Çizim üretildi.
    assert "<svg" in cevap.text and "<rect" in cevap.text


def test_yerlesimde_son_durak_once_yuklenir(istemci, fabrika):
    """İç piyasa planında son uğrak dibe konur; yükleme sırası 1 ona ait olur."""
    from app.models import SevkiyatPlani
    from app.services import istif_servisi

    ic_piyasa_verisi_yukle(istemci)
    istemci.post("/rota/planlar/uret", data={"tipler": ["FTL"]})
    with fabrika() as db:
        plan = db.query(SevkiyatPlani).filter_by(modul="ROTA").one()
        # Rota: MANISA (350 km) sonra IZMIR (400 km) — son uğrak İzmir.
        assert plan.son_ugrak == "IZMIR"
        istif = istif_servisi.istif_plani(db, plan)
        dipteki = min(istif.yerlesimler, key=lambda y: (y.x, y.y))
        assert dipteki.yuk.durak.il == "IZMIR"
        assert dipteki.yukleme_sirasi == 1


def test_urun_gruplari_ekranindan_ad_degistirilir(istemci, fabrika):
    """Ad değişikliği master datanın tamamına işlemeli."""
    from app.models import Urun

    urunler = kitap(
        ["StokKodu", "StokAdi", "Ürün Grubu", "Palet içi adet", "Tır yükleme adeti"],
        [["P1", "Panel 1", "PANEL", 10, 100],
         ["P2", "Panel 2", "PANEL", 10, 100],
         ["K1", "Klima 1", "Klima", 10, 100]],
    )
    istemci.post("/masterdata/urunler/yukle", files={"dosya": ("u.xlsx", urunler)})

    ekran = istemci.get("/masterdata/gruplar")
    assert ekran.status_code == 200
    # 'Klima' Türkçe kurallarıyla KLİMA olur; KLIMA diye ikinci grup açılmaz.
    assert "KLİMA" in ekran.text and "KLIMA<" not in ekran.text

    cevap = istemci.post(
        "/masterdata/gruplar/ad",
        data={"kapsam": "IC_PIYASA", "eski_ad": "PANEL", "yeni_ad": "Radyatör"},
    )
    assert "2 ürün güncellendi" in sorgu(cevap)
    with fabrika() as db:
        assert {u.urun_grubu for u in db.query(Urun).all()} == {"RADYATÖR", "KLİMA"}


def test_ihracat_urunu_ekrandan_duzenlenir(istemci, fabrika):
    from app.models import IhracatUrunu

    urunler = kitap(
        ["ÜRÜN KODU", "ÜRÜN", "Ürün Grubu", "TIR", "DESİ"],
        [["E1", "Radiator", "Radiator", 3000, 2.5],
         ["E2", "Valve", "Acc.", None, None]],
    )
    istemci.post("/masterdata/ihracat-urunler/yukle", files={"dosya": ("i.xlsx", urunler)})

    liste = istemci.get("/masterdata/ihracat-urunler", params={"eksik": "OLCUSUZ"})
    assert "E2" in liste.text and "E1" not in liste.text

    assert istemci.get("/masterdata/ihracat-urunler/E2").status_code == 200
    cevap = istemci.post(
        "/masterdata/ihracat-urunler/kaydet",
        data={"geri": "/masterdata/ihracat-urunler", "urun_kodu": "E2",
              "urun_adi": "Valve", "urun_grubu": "Acc.", "tir_yukleme_adeti": "2800",
              "desi": "0,4", "en": "60", "boy": "120", "aktif": "E"},
    )
    assert "güncellendi" in sorgu(cevap)
    with fabrika() as db:
        urun = db.query(IhracatUrunu).filter_by(urun_kodu="E2").one()
        assert urun.tir_yukleme_adeti == Decimal(2800)
        assert urun.desi == Decimal("0.4")


def test_ihracat_urun_indirmesi_filtreyi_uygular(istemci, tmp_path):
    """Ekranda görülen filtre inen dosyaya da uygulanmalı."""
    from openpyxl import load_workbook

    urunler = kitap(
        ["ÜRÜN KODU", "ÜRÜN", "Ürün Grubu", "TIR", "DESİ"],
        [["E1", "Radiator", "Radiator", 3000, 2.5],
         ["E2", "Valve", "Acc.", None, None]],
    )
    istemci.post("/masterdata/ihracat-urunler/yukle", files={"dosya": ("i.xlsx", urunler)})

    dosya = tmp_path / "ih.xlsx"
    dosya.write_bytes(
        istemci.get(
            "/masterdata/ihracat-urunler/excel", params={"eksik": "OLCUSUZ"}
        ).content
    )
    sayfa = load_workbook(dosya).active
    kodlar = {sayfa.cell(row=r, column=1).value for r in range(2, sayfa.max_row + 1)}
    assert kodlar == {"E2"}


def test_uc_planlama_modulu_ayni_ekranlari_sunar(istemci):
    """Ring, iç piyasa ve ihracat aynı işi yapıyor; ekran takımları da aynı olmalı.

    Ekranlar tek tek eklendiği için modüller arasında boşluklar oluşmuştu: ihracatın
    rapor ekranı yoktu, ring'in günlük yükleme formu yoktu. Bu test o boşlukların
    geri gelmesini engeller.
    """
    for kok in ("/ring", "/rota", "/ihracat"):
        for ekran in ("", "/siparisler", "/planlar", "/manuel-plan",
                      "/bekleyenler", "/raporlar"):
            cevap = istemci.get(kok + ekran)
            assert cevap.status_code == 200, kok + ekran
        # Plan listesi her modülde raporların altında duruyor.
        assert istemci.get(f"{kok}/raporlar/plan-excel").status_code == 200


def test_ring_planinda_yerlesim_ekrani_yok(istemci, fabrika):
    """Yerleşim planı iç piyasa ve ihracatta var, ring'de yok.

    Ring planı tek üründen ve tek depo çıkışından oluşur; araç tek noktaya
    boşaltılır, durak sırası yoktur. Palet palet yerleşim çizmenin depoya
    kattığı bir bilgi olmuyordu.
    """
    yollar = {r.path for r in uygulama.routes if hasattr(r, "path")}
    assert "/ring/planlar/{plan_id}/yerlesim" not in yollar
    assert "/rota/planlar/{plan_id}/yerlesim" in yollar
    assert "/ihracat/planlar/{plan_id}/yerlesim" in yollar

    # Plan detayında da düğmesi kalmamalı.
    kombi_plan_id, _ = _urun_bagi_ortami(fabrika)
    detay = istemci.get(f"/ring/planlar/{kombi_plan_id}")
    assert "Yerleşim planı" not in detay.text


def test_ihracat_plan_listesi_eski_adresten_filtresiyle_yonlenir(istemci):
    """Eski /ihracat/plan-excel adresi filtreyi kaybetmeden yeni adrese gider."""
    cevap = istemci.get(
        "/ihracat/plan-excel", params={"durum": "TASLAK", "arama": "VAILLANT"},
        follow_redirects=False,
    )
    assert cevap.status_code == 307
    assert cevap.headers["location"].startswith("/ihracat/raporlar/plan-excel?")
    assert "durum=TASLAK" in cevap.headers["location"]
    assert "arama=VAILLANT" in cevap.headers["location"]


# --------------------------------------------------- birlikte sevk edilecek ürünler
def _urun_bagi_ortami(fabrika):
    """Aynı bayinin kombisi ve bacası, ayrı teslimatlarda ve ayrı planlarda."""
    from datetime import date
    from decimal import Decimal as D

    from app.models import SevkiyatPlani, SiparisDurumu, SiparisSatiri, Urun

    def plan(sefer, anahtar, urun):
        return SevkiyatPlani(
            sefer_no=sefer, donem="2609", depo_kodu="64", planlama_anahtari=anahtar,
            urun_kodlari=urun, toplam_birim=D(1), doluluk_yuzdesi=D(100), modul="RING",
            plan_tipi="RING_PALET", plan_tarihi=date(2026, 9, 1),
        )

    with fabrika() as db:
        db.add(Urun(urun_kodu="KOMBI-1", urun_adi="ecoTEC 24", urun_grubu="KOMBİ",
                    palet_ici_adet=10, tir_yukleme_adeti=100))
        db.add(Urun(urun_kodu="BACA-1", urun_adi="Baca Seti 60/100", urun_grubu="BACA",
                    palet_ici_adet=20, tir_yukleme_adeti=400))
        kombi_plan = plan("2609D9001", "KOMBI-1", "KOMBI-1")
        baca_plan = plan("2609D9002", "BACA-1", "BACA-1")
        db.add_all([kombi_plan, baca_plan])
        db.flush()
        db.add(SiparisSatiri(
            siparis_no="S1", siparis_satir_no="10", teslimat_no="T1",
            urun_kodu="KOMBI-1", urun_adi="ecoTEC 24", miktar=D(10), depo_kodu="64",
            bayi_adi="EGE ISITMA", modul="RING", plan_id=kombi_plan.id,
            durum=SiparisDurumu.PLANLANDI))
        db.add(SiparisSatiri(
            siparis_no="S1", siparis_satir_no="20", teslimat_no="T2",
            urun_kodu="BACA-1", urun_adi="Baca Seti 60/100", miktar=D(10),
            depo_kodu="64", bayi_adi="EGE ISITMA", modul="RING",
            plan_id=baca_plan.id, durum=SiparisDurumu.PLANLANDI))
        db.commit()
        return kombi_plan.id, baca_plan.id


def test_aksesuar_ana_urunsuz_plana_girince_uyarir(istemci, fabrika):
    """Sahadaki şikâyet: kombi gidiyor, bacası iki gün sonra gidiyor.

    Teslimat numaraları farklı olduğu için teslimat bazlı koruma bu durumu
    yakalamıyordu; bağ ürün kodu üzerinden kurulunca yakalanıyor.
    """
    kombi_plan_id, baca_plan_id = _urun_bagi_ortami(fabrika)

    istemci.post(
        "/masterdata/urun-baglari/kaydet",
        data={"ana_urun_kodu": "KOMBI-1", "bagli_urun_kodu": "BACA-1",
              "tip": "AKSESUAR", "aciklama": "ecoTEC baca seti"},
    )
    liste = istemci.get("/masterdata/urun-baglari")
    assert "BACA-1" in liste.text and "AKSESUAR" in liste.text

    # Baca planı: aksesuar ana ürünü olmadan gidiyor → ağır uyarı, kombinin seferi.
    baca = istemci.get(f"/ring/planlar/{baca_plan_id}")
    assert "Eksik parça" in baca.text
    assert "KOMBI-1" in baca.text and "2609D9001" in baca.text

    # Kombi planı: aksesuarı başka planda → uyarı var ama ağır değil.
    kombi = istemci.get(f"/ring/planlar/{kombi_plan_id}")
    assert "BACA-1" in kombi.text and "2609D9002" in kombi.text
    assert "Aksesuar uyarısı" in kombi.text


def test_set_bagi_iki_yone_de_uyarir(istemci, fabrika):
    """Klima iç/dış ünite: hangisi planda olursa olsun diğeri aranır."""
    kombi_plan_id, baca_plan_id = _urun_bagi_ortami(fabrika)
    istemci.post(
        "/masterdata/urun-baglari/kaydet",
        data={"ana_urun_kodu": "KOMBI-1", "bagli_urun_kodu": "BACA-1", "tip": "SET"},
    )
    for plan_id in (kombi_plan_id, baca_plan_id):
        cevap = istemci.get(f"/ring/planlar/{plan_id}")
        assert "Eksik parça" in cevap.text, plan_id
        assert "SET" in cevap.text


def test_urun_bagi_excel_gidip_geri_yuklenir(istemci, fabrika):
    """İnen dosya doğrudan geri yüklenebilmeli: başlıklar tek kaynaktan geliyor."""
    _urun_bagi_ortami(fabrika)
    istemci.post(
        "/masterdata/urun-baglari/kaydet",
        data={"ana_urun_kodu": "KOMBI-1", "bagli_urun_kodu": "BACA-1", "tip": "SET"},
    )
    inen = istemci.get("/masterdata/urun-baglari/excel")
    assert inen.status_code == 200

    geri = istemci.post(
        "/masterdata/urun-baglari/yukle",
        files={"dosya": ("urun_baglari.xlsx", BytesIO(inen.content),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert "1 güncellendi" in sorgu(geri) or "güncellendi" in sorgu(geri)
    assert istemci.get("/masterdata/urun-baglari").text.count("KOMBI-1") >= 1


def test_urun_bagi_tanimsiz_urunu_reddeder(istemci):
    cevap = istemci.post(
        "/masterdata/urun-baglari/kaydet",
        data={"ana_urun_kodu": "YOK-1", "bagli_urun_kodu": "YOK-2", "tip": "SET"},
    )
    assert "tanımlı olmayan ürün kodu" in sorgu(cevap)


def test_urun_bagi_kendine_baglanmaz(istemci, fabrika):
    from app.models import Urun

    with fabrika() as db:
        db.add(Urun(urun_kodu="TEK-1", urun_adi="Tek", urun_grubu="KOMBİ"))
        db.commit()
    cevap = istemci.post(
        "/masterdata/urun-baglari/kaydet",
        data={"ana_urun_kodu": "TEK-1", "bagli_urun_kodu": "TEK-1", "tip": "SET"},
    )
    assert "kendisine bağlanamaz" in sorgu(cevap)


def test_bagi_olmayan_planda_uyari_cikmaz(istemci, fabrika):
    """Bağ tanımlanmadıysa ekran hiçbir uyarı göstermemeli."""
    kombi_plan_id, _ = _urun_bagi_ortami(fabrika)
    cevap = istemci.get(f"/ring/planlar/{kombi_plan_id}")
    assert "Eksik parça" not in cevap.text
    assert "Aksesuar uyarısı" not in cevap.text
