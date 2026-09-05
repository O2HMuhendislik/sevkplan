"""Cari kod ve sevk tipi listelerinin müşteri master datasına işlenmesi."""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.models import Musteri
from app.services import musteri_ek_bilgi as ek
from app.services.musteri_eslestirme import MusteriEslestirici


def musteri_ekle(db, ad, **alanlar):
    from app.domain.iller import yer_adi

    kayit = Musteri(anahtar=yer_adi(ad), bayi_adi=ad, **alanlar)
    db.add(kayit)
    db.flush()
    return kayit


def dosya(tmp_path, cari=(), tipler=(), ad="veri.xlsx"):
    kitap = Workbook()
    kitap.remove(kitap.active)
    if cari:
        s = kitap.create_sheet("CARİ KODLAR")
        s.append(["Teslim Yapılacak Müşteri Adı", "Müşteri Kodu"])
        for satir in cari:
            s.append(list(satir))
    if tipler:
        s = kitap.create_sheet("TESLİMAT TİPİ")
        s.append(["BAYİ ADI", "SEVK TİPİ", "ÖZEL DURUM"])
        for satir in tipler:
            s.append(list(satir) + [None] * (3 - len(satir)))
    yol = tmp_path / ad
    kitap.save(yol)
    return yol


@pytest.mark.parametrize(
    "metin, tir, cumartesi, eirsaliye",
    [
        ("TIR", "E", True, False),
        ("KAMYON", "H", True, False),
        ("KAMYONET", "H", True, False),
        ("RUTİN", "H", True, False),
        ("SADECE KAMYON", "H", True, False),
        ("TIR-EİRSALİYE", "E", True, True),
        ("TIR-C.TESİ YOK-EİRSALİYE", "E", False, True),
        ("KAMYON-C.TESİ YOK-EİRSALİYE", "H", False, True),
        ("TIRE-C.TESİ YOK-EİRSALİYE", "E", False, True),   # yazım hatası
        ("ZORDA KALIRSAN TIR", "E", True, False),
        ("SOR", "?", True, False),
    ],
)
def test_sevk_tipi_metni_alanlara_ayrilir(metin, tir, cumartesi, eirsaliye):
    """Sahanın tek sütunda yazdığı metin üç bilgiyi birden taşıyor."""
    cozum = ek.sevk_tipini_coz(metin)
    assert cozum["tir_girisi"] == tir
    assert cozum["cumartesi_teslimat"] is cumartesi
    assert cozum["e_irsaliye"] is eirsaliye
    assert cozum["sevk_tipi"] == metin


def test_cari_kod_ve_sevk_tipi_islenir(db, tmp_path):
    musteri_ekle(db, "EGE ISITMA SOĞUTMA")
    yol = dosya(
        tmp_path,
        cari=[("EGE ISITMA SOĞUTMA", "20001362")],
        tipler=[("EGE ISITMA SOĞUTMA", "KAMYON-C.TESİ YOK-EİRSALİYE", "CUMARTESİ YOK")],
    )
    sonuc = ek.ek_bilgileri_aktar(db, yol)

    kayit = db.query(Musteri).one()
    assert kayit.bayi_kodu == "20001362"
    assert kayit.tir_girisi == "H"
    assert kayit.cumartesi_teslimat is False
    assert kayit.e_irsaliye is True
    assert kayit.ozel_durum == "CUMARTESİ YOK"
    assert (sonuc.cari_yazilan, sonuc.tip_yazilan, sonuc.tir_degisen) == (1, 1, 1)


def test_noktalama_ve_unvan_farki_eslesmeyi_bozmaz(db, tmp_path):
    """'X MÜH. LTD. ŞTİ.' ile 'X MÜHENDİSLİK' aynı bayi."""
    musteri_ekle(db, "ÇARMIK MAKİNE MÜHENDİSLİK")
    yol = dosya(tmp_path, cari=[("ÇARMIK MAKİNE MÜH. LTD.ŞTİ.", "21900574")])
    ek.ek_bilgileri_aktar(db, yol)
    assert db.query(Musteri).one().bayi_kodu == "21900574"


def test_benzer_ad_otomatik_eslestirilmez(db, tmp_path):
    """'AKKAŞ ISI DEPO' ile 'ARSE ISI DEPO' %81 benziyor ama farklı bayiler."""
    musteri_ekle(db, "ARSE ISI DEPO")
    yol = dosya(tmp_path, cari=[("AKKAŞ ISI DEPO", "20870113")])
    sonuc = ek.ek_bilgileri_aktar(db, yol)

    assert db.query(Musteri).one().bayi_kodu is None
    assert sonuc.cari_yazilan == 0
    assert sonuc.eslesmeyenler[0]["sebep"] == "eşleşme yok"


def test_ayni_musteriye_farkli_deger_yazilmaz(db, tmp_path):
    """Dosyada aynı bayi iki satırda farklı yazılıysa hangisi doğru belli değil."""
    musteri_ekle(db, "AS MÜHENDİSLİK")
    yol = dosya(
        tmp_path,
        tipler=[("AS MÜHENDİSLİK", "KAMYON"), ("AS MÜHENDİSLİK LTD.", "TIR")],
    )
    sonuc = ek.ek_bilgileri_aktar(db, yol)

    kayit = db.query(Musteri).one()
    assert kayit.tir_girisi == "?"           # dokunulmadı
    assert kayit.sevk_tipi is None
    assert sonuc.tip_yazilan == 0
    assert all(
        k["sebep"].startswith("çakışma") for k in sonuc.eslesmeyenler
    )


def test_belirsiz_ad_iki_musteriye_uyuyorsa_yazilmaz(db, tmp_path):
    """Bir kademede iki kayda birden uyan ad yazılmaz; hangisi olduğu bilinemez.

    Sistemde noktalamayla ayrılan iki kayıt var; dosyadaki ad ikisine de birebir
    uymuyor ama noktalama yok sayılınca ikisine de uyuyor.
    """
    musteri_ekle(db, "KAYA İKLİMLENDİRME", il="BURSA")
    musteri_ekle(db, "KAYA-İKLİMLENDİRME", il="IZMIR")
    yol = dosya(tmp_path, cari=[("KAYA. İKLİMLENDİRME", "20001507")])
    sonuc = ek.ek_bilgileri_aktar(db, yol)

    assert all(m.bayi_kodu is None for m in db.query(Musteri).all())
    assert sonuc.eslesmeyenler[0]["sebep"] == "birden fazla aday"
    assert "KAYA" in sonuc.eslesmeyenler[0]["adaylar"]


def test_birebir_ad_daha_gevsek_eslesmeye_tercih_edilir(db, tmp_path):
    """Ad birebir tutuyorsa, noktalamasız hâli başkasına da uysa bile o kazanır."""
    musteri_ekle(db, "DENGE MÜH DEPO", il="BURSA")
    musteri_ekle(db, "DENGE MÜH. DEPO", il="IZMIR")
    yol = dosya(tmp_path, cari=[("DENGE MÜH DEPO", "20001507")])
    ek.ek_bilgileri_aktar(db, yol)

    kodlu = [m for m in db.query(Musteri).all() if m.bayi_kodu]
    assert len(kodlu) == 1 and kodlu[0].il == "BURSA"


def test_yer_tutucu_cari_kod_yazilmaz(db, tmp_path):
    musteri_ekle(db, "EGE ISITMA")
    yol = dosya(tmp_path, cari=[("EGE ISITMA", "-")])
    sonuc = ek.ek_bilgileri_aktar(db, yol)
    assert db.query(Musteri).one().bayi_kodu is None
    assert sonuc.cari_okunan == 0


def test_alici_firma_da_aday_sayilir(db, tmp_path):
    musteri_ekle(db, "KAYA İKL.", alici_firma="KAYA İKLİMLENDİRME")
    yol = dosya(tmp_path, tipler=[("KAYA İKLİMLENDİRME", "TIR")])
    ek.ek_bilgileri_aktar(db, yol)
    assert db.query(Musteri).one().sevk_tipi == "TIR"


def test_eslesmeyen_raporu_adaylariyla_yazilir(db, tmp_path):
    from openpyxl import load_workbook

    musteri_ekle(db, "ARSE ISI DEPO")
    yol = dosya(tmp_path, cari=[("AKKAŞ ISI DEPO", "20870113")])
    sonuc = ek.ek_bilgileri_aktar(db, yol)
    rapor = ek.eslesmeyen_raporu(sonuc, tmp_path / "eslesmeyen.xlsx")

    sayfa = load_workbook(rapor).active
    assert [h.value for h in sayfa[1]] == [
        "Kaynak", "Dosyadaki ad", "Değer", "Sebep", "Sistemdeki aday kayıtlar"
    ]
    assert sayfa.cell(row=2, column=2).value == "AKKAŞ ISI DEPO"
    assert sayfa.cell(row=2, column=3).value == "20870113"


def test_eslestirici_kesin_kademeleri_kullanir(db):
    musteri_ekle(db, "EGE ISITMA SOĞUTMA")
    e = MusteriEslestirici(db)
    for ad in ("EGE ISITMA SOĞUTMA", "EGE ISITMA SOGUTMA", "E.G.E ISITMA-SOĞUTMA",
               "EGE ISITMA SOĞUTMA LTD. ŞTİ."):
        kayit, adaylar = e.esle(ad)
        assert kayit is not None, ad
    assert e.esle("BAŞKA BİR BAYİ") == (None, [])


def test_disa_aktarilan_dosya_sevk_tipiyle_geri_yuklenir(db, tmp_path):
    """İndir → doldur → geri yükle akışı yeni alanları da taşır."""
    from openpyxl import Workbook

    from app.services import ice_aktarim

    musteri_ekle(db, "EGE ISITMA")
    kitap = Workbook()
    s = kitap.active
    s.append(["Bayi Adı", "İl", "Sevk Tipi", "E-posta", "Özel Durum"])
    s.append(["EGE ISITMA", "İZMİR", "KAMYON-C.TESİ YOK-EİRSALİYE",
              "bayi@ornek.com.tr", "CUMARTESİ MAL KABUL YOK"])
    yol = tmp_path / "musteri.xlsx"
    kitap.save(yol)

    ice_aktarim.musterileri_aktar(db, yol, "musteri.xlsx")

    kayit = db.query(Musteri).one()
    assert kayit.sevk_tipi == "KAMYON-C.TESİ YOK-EİRSALİYE"
    assert kayit.tir_girisi == "H"
    assert kayit.cumartesi_teslimat is False
    assert kayit.e_irsaliye is True
    assert kayit.eposta == "bayi@ornek.com.tr"
    assert kayit.ozel_durum == "CUMARTESİ MAL KABUL YOK"


def test_tir_girisi_sutunu_sevk_tipinin_onune_gecer(db, tmp_path):
    """Kullanıcı sütunu elle doldurduysa sevk tipinden türetilen değer ezmez."""
    from openpyxl import Workbook

    from app.services import ice_aktarim

    musteri_ekle(db, "EGE ISITMA")
    kitap = Workbook()
    s = kitap.active
    s.append(["Bayi Adı", "İl", "Sevk Tipi", "Tır Girişi (E/H/?)"])
    s.append(["EGE ISITMA", "İZMİR", "KAMYON", "E"])
    yol = tmp_path / "m.xlsx"
    kitap.save(yol)

    ice_aktarim.musterileri_aktar(db, yol, "m.xlsx")
    kayit = db.query(Musteri).one()
    assert kayit.tir_girisi == "E"
    assert kayit.sevk_tipi == "KAMYON"


def test_gomulu_ek_bilgi_bir_kez_islenir(db, tmp_path, monkeypatch):
    """Kurulumda bir kez işlenir; sonraki açılışta ekran düzeltmelerini ezmez."""
    from app.services import gomulu_veri

    musteri_ekle(db, "EGE ISITMA")
    yol = dosya(tmp_path, tipler=[("EGE ISITMA", "KAMYON")], ad="ek.xlsx")
    monkeypatch.setattr(gomulu_veri, "EK_BILGI_DOSYASI", yol)

    assert "246" not in (gomulu_veri._musteri_ek_bilgisi(db) or "")
    kayit = db.query(Musteri).one()
    assert kayit.tir_girisi == "H" and kayit.sevk_tipi == "KAMYON"

    # Kullanıcı ekrandan düzeltti; ikinci açılış üzerine yazmamalı.
    kayit.tir_girisi = "E"
    db.flush()
    assert gomulu_veri._musteri_ek_bilgisi(db) is None
    assert db.query(Musteri).one().tir_girisi == "E"
