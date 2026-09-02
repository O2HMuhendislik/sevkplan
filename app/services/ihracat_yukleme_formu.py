"""İhracat yükleme formu.

Düzen sahadaki "… İHRACAT Yükleme Formları" kitabının **EXPORT** sayfasının
karşılığıdır. İç piyasa formundan farkları:

* Üst blokta ülke, araç tipi ve depo/AXATA kutuları; sefer numarası müşteriye göre
  `N` ya da `E` belge koduyla üretilir (`2608E4001`).
* Satır tablosu ihracat sütunlarıyla: ülke, sipariş, ürün, adet, müşteri, sevk adresi,
  teslimat, desi.
* Alt blokta iki sütun: solda **sevk ve araç bilgileri** (desi, araç tipi, yükleme
  tipi, azami tonaj, çekici plakası, dorse/konteyner, mühür), sağda **sipariş ve depo
  bilgileri** (planlama tarihi, satış destek, araç tedarik, toplayan, sevk kontrol,
  vardiya amiri) ve müşteriye özel yükleme notu (hava yastığı, silika jel, dökme …).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Font
from openpyxl.worksheet.pagebreak import Break

from app.config import CIKTI_DIZIN
from app.models import SevkiyatPlani
from app.services.excel import sayfa_yaz, yeni_kitap
from app.services.yukleme_formu import KALIN, KENAR, KUCUK, ORTALI, SOLA, BASLIK_DOLGU

FORM_NO = "8101058098.02"
"""İhracat formunun doküman numarası (iç piyasa formundan farklı)."""

HASAR_UYARISI = (
    "İHRACAT YÜKLEMELERİNDE HASAR RİSKİNE KARŞI PALET YÜKSELTME VE SABİTLEME "
    "TALİMATLARINA UYULMASI GEREKMEKTEDİR"
)

BASLIKLAR = [
    "NO", "DEPO", "AXATA", "ÜLKE", "SİPARİŞ NO", "ÜRÜN KODU", "ÜRÜN TANIMI",
    "ADET", "MÜŞTERİ ADI", "SEVK ADRESİ", "TESLİMAT NO", "Desi",
]
KOLON_GENISLIKLERI = {
    "A": 6.0, "B": 8.0, "C": 12.0, "D": 9.0, "E": 15.0, "F": 14.0,
    "G": 42.0, "H": 8.0, "I": 34.0, "J": 20.0, "K": 15.0, "L": 12.0, "M": 22.0,
}
SUTUN_SAYISI = len(BASLIKLAR)
ASGARI_SATIR = 12


def _depo_kutusu(plan: SevkiyatPlani) -> list[tuple[str, str]]:
    """Formdaki depo/AXATA kutusu: hangi depodan yükleniyorsa Axata onun karşısına."""
    depolar = sorted({satir.depo_kodu for satir in plan.satirlar})
    if not depolar:
        depolar = [plan.depo_kodu or "34"]
    axata = plan.axata_ozeti or ""
    return [(f"{depo}-DEPO AXATA", axata if depo == depolar[0] else "") for depo in depolar]


def _ust_blok(sayfa, plan: SevkiyatPlani, ust: int) -> int:
    sayfa.cell(row=ust, column=12, value="Doküman numarası").font = KUCUK

    sayfa.cell(row=ust, column=1, value="SEVK TARİHİ").font = KALIN
    tarih = sayfa.cell(row=ust, column=3, value=plan.plan_tarihi)
    tarih.number_format = "DD.MM.YYYY"
    tarih.font = KALIN

    sayfa.cell(row=ust + 1, column=1, value="SEFER NO").font = KALIN
    sayfa.cell(row=ust + 1, column=3, value=plan.sefer_no).font = Font(bold=True, size=12)
    sayfa.cell(row=ust + 1, column=12, value=FORM_NO).font = KUCUK

    for sira, (etiket, deger) in enumerate(_depo_kutusu(plan)):
        satir = ust + sira
        hucre = sayfa.cell(row=satir, column=5, value=etiket)
        hucre.font = KALIN
        hucre.border = KENAR
        axata = sayfa.cell(row=satir, column=6, value=deger)
        axata.font = KALIN
        axata.border = KENAR

    sayfa.cell(row=ust, column=7, value=plan.ulke or "").font = Font(bold=True, size=12)
    sayfa.cell(row=ust, column=10, value=plan.arac_tipi or "").font = KALIN
    return ust + max(2, len(_depo_kutusu(plan)))


def _satir_tablosu(sayfa, plan: SevkiyatPlani, baslik_satiri: int) -> int:
    for sutun, ad in enumerate(BASLIKLAR, start=1):
        hucre = sayfa.cell(row=baslik_satiri, column=sutun, value=ad)
        hucre.font = KALIN
        hucre.border = KENAR
        hucre.alignment = ORTALI
        hucre.fill = BASLIK_DOLGU

    satirlar = sorted(
        plan.satirlar, key=lambda s: (s.depo_kodu, s.teslimat_no, s.urun_kodu)
    )
    veri_basi = baslik_satiri + 1
    for sira, satir in enumerate(satirlar):
        y = veri_basi + sira
        degerler = [
            sira + 1,
            satir.depo_kodu,
            plan.axata_ozeti or "",
            satir.ulke_kodu or "",
            satir.siparis_no,
            satir.urun_kodu,
            satir.gosterilecek_urun_adi,
            float(satir.miktar),
            satir.bayi_adi,
            satir.sevk_adresi,
            satir.teslimat_no,
            float(satir.desi or 0),
        ]
        for sutun, deger in enumerate(degerler, start=1):
            hucre = sayfa.cell(row=y, column=sutun, value=deger)
            hucre.border = KENAR
            hucre.font = KUCUK
            hucre.alignment = SOLA if sutun in (7, 9, 10) else ORTALI

    bos_basi = veri_basi + len(satirlar)
    veri_sonu = veri_basi + max(len(satirlar), ASGARI_SATIR) - 1
    for sira, y in enumerate(range(bos_basi, veri_sonu + 1), start=len(satirlar) + 1):
        sayfa.cell(row=y, column=1, value=sira).font = KUCUK
        for sutun in range(1, SUTUN_SAYISI + 1):
            sayfa.cell(row=y, column=sutun).border = KENAR
    return veri_sonu


def _alt_blok(sayfa, plan: SevkiyatPlani, ilk_satir: int) -> int:
    y = ilk_satir + 1
    sayfa.cell(row=y, column=7, value="TOPLAM PARÇA ADEDİ :").font = KALIN
    sayfa.cell(row=y, column=8, value=float(plan.toplam_adet or 0)).font = KALIN

    y += 1
    sol = sayfa.cell(row=y, column=1, value="SEVK BİLGİLERİ")
    sol.font = KALIN
    sag = sayfa.cell(row=y, column=5, value="SİPARİŞ BİLGİLERİ")
    sag.font = KALIN
    uyari = sayfa.cell(row=y, column=7, value=HASAR_UYARISI)
    uyari.font = Font(bold=True, size=9, color="C00000")
    uyari.alignment = ORTALI
    sayfa.merge_cells(start_row=y, start_column=7, end_row=y, end_column=8)
    if plan.musteri_aciklamasi:
        # Müşteriye özel yükleme notu: hava yastığı, silika jel, paletsiz dökme ...
        not_hucresi = sayfa.cell(row=y, column=9, value=plan.musteri_aciklamasi)
        not_hucresi.font = Font(bold=True, size=9)
        not_hucresi.alignment = SOLA

    sol_alanlar = [
        ("DESİ", float(plan.toplam_desi or 0)),
        ("AĞIRLIK (KG)", float(plan.toplam_agirlik or 0)),
        ("ARAÇ TİPİ", plan.arac_tipi or ""),
        ("YÜKLEME TİPİ", plan.yukleme_tipi or ""),
        ("MAKSİMUM TONAJ", float(plan.azami_agirlik) if plan.azami_agirlik else ""),
        ("ARAÇ BİLGİLERİ", ""),
        ("ÇEKİCİNİN PLAKASI", plan.plaka or ""),
        ("DORSE/KONTEYNER NO", plan.konteyner_no or ""),
        ("MÜHÜR NO", plan.muhur_no or ""),
    ]
    sag_alanlar = [
        ("PLANLAMA TARİHİ", plan.plan_tarihi),
        ("SATIŞ DESTEK", ""),
        ("ARAÇ TEDARİK", plan.nakliyeci or ""),
        ("PLANLAMA", plan.olusturan),
        ("DEPO BİLGİLERİ", ""),
        ("TOPLAYAN", ""),
        ("SEVK KONTROL", ""),
        ("VARDİYA AMİRİ", ""),
    ]
    for sira, (etiket, deger) in enumerate(sol_alanlar):
        satir = y + 1 + sira
        sayfa.cell(row=satir, column=1, value=etiket).font = KALIN
        hucre = sayfa.cell(row=satir, column=3, value=deger)
        hucre.border = KENAR
        hucre.alignment = SOLA
        if etiket == "PLANLAMA TARİHİ":
            hucre.number_format = "DD.MM.YYYY"
    for sira, (etiket, deger) in enumerate(sag_alanlar):
        satir = y + 1 + sira
        sayfa.cell(row=satir, column=5, value=etiket).font = KALIN
        hucre = sayfa.cell(row=satir, column=6, value=deger)
        hucre.border = KENAR
        hucre.alignment = SOLA
        if etiket == "PLANLAMA TARİHİ":
            hucre.number_format = "DD.MM.YYYY"

    alt = y + 1 + max(len(sol_alanlar), len(sag_alanlar))
    if plan.marka_paylari:
        sayfa.cell(row=alt, column=5, value="FATURA YÜZDESİ").font = KALIN
        for sira, (ad, oran) in enumerate(plan.marka_paylari.items()):
            sayfa.cell(row=alt + sira, column=6, value=ad).font = KUCUK
            yuzde = sayfa.cell(row=alt + sira, column=7, value=float(oran))
            yuzde.number_format = "0%"
            yuzde.font = KALIN
        alt += len(plan.marka_paylari)
    return alt + 2


def _blok_yaz(sayfa, plan: SevkiyatPlani, ust: int) -> int:
    baslik_satiri = _ust_blok(sayfa, plan, ust) + 1
    veri_sonu = _satir_tablosu(sayfa, plan, baslik_satiri)
    return _alt_blok(sayfa, plan, veri_sonu)


def _sayfayi_hazirla(sayfa) -> None:
    sayfa.page_setup.orientation = "landscape"
    sayfa.page_setup.paperSize = 9  # A4
    sayfa.page_setup.fitToWidth = 1
    sayfa.sheet_properties.pageSetUpPr.fitToPage = True
    for kolon, genislik in KOLON_GENISLIKLERI.items():
        sayfa.column_dimensions[kolon].width = genislik


def formlari_uret(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    """Planları tek EXPORT sayfasına alt alta, her biri ayrı sayfaya basılacak şekilde yazar."""
    if not planlar:
        raise ValueError("Form üretmek için en az bir plan gerekir.")
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("EXPORT")
    _sayfayi_hazirla(sayfa)

    ust = 2
    sirali = sorted(planlar, key=lambda p: p.sefer_no)
    for sira, plan in enumerate(sirali):
        son = _blok_yaz(sayfa, plan, ust)
        ust = son + 2
        if sira < len(sirali) - 1:
            sayfa.row_breaks.append(Break(id=ust - 1))

    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef


def form_uret(plan: SevkiyatPlani, hedef: Path | None = None) -> Path:
    hedef = hedef or CIKTI_DIZIN / f"ihracat_yukleme_formu_{plan.sefer_no}.xlsx"
    return formlari_uret([plan], hedef)


def gunluk_form(planlar: list[SevkiyatPlani], hedef: Path | None = None) -> Path:
    if not planlar:
        raise ValueError("Form üretmek için en az bir plan gerekir.")
    tarih = planlar[0].plan_tarihi
    hedef = hedef or CIKTI_DIZIN / f"ihracat_yukleme_formlari_{tarih:%Y%m%d}.xlsx"
    return formlari_uret(planlar, hedef)


def plan_listesi_disa_aktar(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("İhracat Planları")
    basliklar = [
        "Sefer No", "Plan Tarihi", "Müşteri", "Ülke", "Ülke Kodu", "Araç Tipi",
        "Taşıma Modu", "Yükleme Tipi", "Durum", "Axata No", "Depo",
        "Desi", "Ağırlık (kg)", "Azami Tonaj", "Doluluk %", "Kısıtlayan",
        "Adet", "Teslimat", "Plaka", "Konteyner No", "Mühür No", "Nakliyeci",
        "Marka Payı", "Oluşturan",
    ]
    satirlar = [
        [
            plan.sefer_no,
            plan.plan_tarihi.strftime("%d.%m.%Y") if plan.plan_tarihi else "",
            plan.musteri_adi or "",
            plan.ulke or "",
            plan.ulke_kodu or "",
            plan.arac_tipi or "",
            plan.tasima_modu or "",
            plan.yukleme_tipi or "",
            plan.durum.value,
            plan.axata_ozeti or "",
            plan.depo_kodu,
            float(plan.toplam_desi or 0),
            float(plan.toplam_agirlik or 0),
            float(plan.azami_agirlik or 0),
            float(plan.doluluk_yuzdesi or 0),
            plan.kisitlayan_olcu or "",
            float(plan.toplam_adet or 0),
            plan.teslimat_sayisi,
            plan.plaka or "",
            plan.konteyner_no or "",
            plan.muhur_no or "",
            plan.nakliyeci or "",
            plan.marka_ozeti,
            plan.olusturan,
        ]
        for plan in planlar
    ]
    sayfa_yaz(sayfa, basliklar, satirlar)
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef
