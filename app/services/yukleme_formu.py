"""Yükleme formu üretimi.

Düzen, depo operasyonun kullandığı "YÜKLEME FORMLARI" (D-RİNG) sayfasının birebir
karşılığıdır: form no, sefer no, depo/AXATA kutusu, plan sevk tarihi, sipariş
satırları, toplam adet ve imza alanları. Bir çalışma kitabına birden çok plan alt
alta yazılır ve her plan ayrı sayfaya basılacak şekilde sayfa sonu konur.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from app.config import CIKTI_DIZIN
from app.models import SevkiyatPlani
from app.services.excel import sayfa_yaz, yeni_kitap

FORM_NO = "8101058099.01"
UYARI_METNI = (
    "EKSİK ÜRÜN ÇIKIŞI YAPILMASI DURUMUNDA İLGİLİ PLANLAMACIYA BİLGİ VERİLMESİ "
    "RİCA OLUNUR. HİT SİSTEMİNE YAPILAN GİRİŞ EMİRLERİNİN DÜZELTİLMESİ GEREKMEKTEDİR"
)
DEPO_SATIRLARI = ("34-DEPO", "44-DEPO", "64-D DEPO", "64-V DEPO", "74-DEPO")
"""Form üzerindeki depo/AXATA kutusunun satırları. Sıra formdakiyle aynı olmalıdır."""

BASLIKLAR = [
    "No", "İl Adi", "Sipariş No", "Belge No", "Depo ", "Ürün Kodu", "Ürün Adi",
    "Adet", "Bayii Adı", "Alıcı Firma", "Sevk Adresi", "Teslimat",
]
"""Kaynak Ring dosyasının sütun düzeni (BayiAdi / AliciFirma / SevkAdresi).

Son iki başlık önce boştu; depo formda hangi sütunun ne olduğunu göremiyordu.
"""
KOLON_GENISLIKLERI = {
    "A": 9.7, "B": 14.7, "C": 15.7, "D": 14.7, "E": 12.4, "F": 14.3,
    "G": 49.3, "H": 7.0, "I": 38.0, "J": 6.7, "K": 12.0, "L": 17.1,
}
SUTUN_SAYISI = len(BASLIKLAR)
ASGARI_SATIR = 19
"""Form her zaman en az bu kadar satır gösterir; kalanlar numaralı boş satırdır."""

UYARI_SATIR_YUKSEKLIGI = 30.0
"""Kırmızı uyarı satırının yüksekliği; metin tek satıra sığmadığı için sarılır."""

KALIN = Font(bold=True)
KUCUK = Font(size=9)
INCE = Side(style="thin")
KENAR = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)
ORTALI = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA = Alignment(horizontal="left", vertical="center", wrap_text=True)
BASLIK_DOLGU = PatternFill("solid", fgColor="D9D9D9")

KALIN_YAN = Side(style="medium")
"""Formun dış çerçevesi; iç çizgiler ince kalır."""


def kenar_ekle(hucre, **yanlar) -> None:
    """Hücrenin mevcut kenarlığını bozmadan verilen yanları değiştirir.

    Doğrudan `hucre.border = Border(...)` atamak tablo içindeki ince çizgileri
    siler; çerçeve çizerken var olanın üzerine eklemek gerekir.
    """
    mevcut = hucre.border
    hucre.border = Border(
        left=yanlar.get("left", mevcut.left),
        right=yanlar.get("right", mevcut.right),
        top=yanlar.get("top", mevcut.top),
        bottom=yanlar.get("bottom", mevcut.bottom),
    )


def cerceve_ciz(sayfa, ust: int, alt: int, sol: int = 1, sag: int = 12) -> None:
    """Verilen dikdörtgenin dışına kalın çerçeve çizer.

    Formun tamamı tek çerçeve içinde görünsün diye her plan bloğunun etrafına
    uygulanır; blok içindeki hücrelerin ince kenarlıkları korunur.
    """
    for sutun in range(sol, sag + 1):
        kenar_ekle(sayfa.cell(row=ust, column=sutun), top=KALIN_YAN)
        kenar_ekle(sayfa.cell(row=alt, column=sutun), bottom=KALIN_YAN)
    for satir in range(ust, alt + 1):
        kenar_ekle(sayfa.cell(row=satir, column=sol), left=KALIN_YAN)
        kenar_ekle(sayfa.cell(row=satir, column=sag), right=KALIN_YAN)


def gridleri_gizle(sayfa) -> None:
    """Arka plandaki hücre kılavuz çizgilerini kapatır; formun kendi çizgileri kalır."""
    sayfa.sheet_view.showGridLines = False
    sayfa.print_options.gridLines = False


def depo_etiketi(depo_kodu: str) -> str | None:
    """Plan deposunu formdaki depo kutusunun satırlarından biriyle eşler.

    Eşleşme bulunamazsa None döner; o durumda kutuya deponun kendi kodu için ek bir
    satır açılır (bkz. `_depo_satirlari`). Böylece Axata numarası hiçbir depoda
    kaybolmaz.
    """
    kod = (depo_kodu or "").strip().upper()
    for etiket in DEPO_SATIRLARI:
        if etiket.replace(" DEPO", "").replace("-DEPO", "") == kod:
            return etiket
    if kod.startswith("64"):
        return "64-V DEPO" if kod.endswith("V") else "64-D DEPO"
    for etiket in DEPO_SATIRLARI:
        if etiket.split("-")[0] == kod.split("-")[0]:
            return etiket
    return None


def _depo_satirlari(depo_kodu: str) -> tuple[list[str], str]:
    """Formun depo/AXATA kutusundaki satırlar ve planın kendi deposunun satırı."""
    etiket = depo_etiketi(depo_kodu)
    if etiket is not None:
        return list(DEPO_SATIRLARI), etiket
    ek = f"{(depo_kodu or '').strip().upper()}-DEPO"
    return [*DEPO_SATIRLARI, ek], ek


def axata_kutusu(plan, depo_satirlari: list[str], hedef_etiket: str) -> dict[str, str]:
    """Depo/AXATA kutusuna basılacak değerler: form satırı -> Axata numaraları.

    Bir planda birden çok depo olabiliyor (ör. 64 + 74) ve her depo kendi Axata iş
    emrini açıyor. Numara bir depoya bağlıysa **yalnızca o deponun satırına** yazılır;
    depo yanlış iş emriyle toplama yapmasın diye. Deposu belirtilmemiş numaralar
    (tek depolu planlar ve eski kayıtlar) planın kendi depo satırına yazılır.
    """
    kutu: dict[str, str] = {}
    baglanmamis = [a.numara for a in plan.axata_numaralari if not a.depo_kodu]
    for depo in plan.axata_depolari:
        etiket = depo_etiketi(depo)
        if etiket is None or etiket not in depo_satirlari:
            continue
        numaralar = [
            a.numara for a in plan.axata_numaralari if a.depo_kodu == depo
        ]
        if numaralar:
            kutu[etiket] = ", ".join(numaralar)
    if baglanmamis:
        mevcut = kutu.get(hedef_etiket)
        kutu[hedef_etiket] = (
            f"{mevcut}, {', '.join(baglanmamis)}" if mevcut else ", ".join(baglanmamis)
        )
    return kutu


def _blok_yaz(sayfa, plan: SevkiyatPlani, ust: int) -> int:
    """Tek bir planın formunu `ust` satırından başlayarak yazar, son satırı döner."""
    satirlar = sorted(
        plan.satirlar, key=lambda s: (s.teslimat_no, s.siparis_no, s.urun_kodu)
    )

    sayfa.cell(row=ust, column=10, value="FORM NO : ").font = KUCUK
    sayfa.cell(row=ust, column=11, value=FORM_NO).font = KUCUK

    baslik = sayfa.cell(row=ust + 1, column=2, value="YÜKLEME FORMLARI")
    baslik.font = Font(bold=True, size=12)
    baslik.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 1, start_column=2, end_row=ust + 1, end_column=4)
    sayfa.cell(row=ust + 1, column=5, value="Depo ").font = KALIN
    sayfa.cell(row=ust + 1, column=6, value="AXATA").font = KALIN

    uyari = sayfa.cell(row=ust + 1, column=7, value=UYARI_METNI)
    uyari.font = Font(bold=True, size=9, color="C00000")
    uyari.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 1, start_column=7, end_row=ust + 1, end_column=12)
    sayfa.row_dimensions[ust + 1].height = UYARI_SATIR_YUKSEKLIGI

    etiket = sayfa.cell(row=ust + 2, column=2, value="SEFER NO")
    etiket.font = KALIN
    etiket.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 2, start_column=2, end_row=ust + 3, end_column=2)
    sefer = sayfa.cell(row=ust + 2, column=3, value=plan.sefer_no)
    sefer.font = Font(bold=True, size=12)
    sefer.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 2, start_column=3, end_row=ust + 3, end_column=4)

    depo_satirlari, hedef_etiket = _depo_satirlari(plan.depo_kodu)
    kutu = axata_kutusu(plan, depo_satirlari, hedef_etiket)
    for sira, depo_adi in enumerate(depo_satirlari):
        satir = ust + 2 + sira
        hucre = sayfa.cell(row=satir, column=5, value=depo_adi)
        hucre.font = KALIN
        hucre.border = KENAR
        hucre.alignment = ORTALI
        axata = sayfa.cell(row=satir, column=6)
        axata.border = KENAR
        axata.alignment = ORTALI
        if kutu.get(depo_adi):
            axata.value = kutu[depo_adi]
            axata.font = KALIN

    # Axata numarası kutunun dışında da yazılır: kutuda satırı olmayan bir depo ya da
    # kutunun gözden kaçması hâlinde numara yine formda görünür.
    axata_etiketi = sayfa.cell(row=ust + 2, column=7, value="AXATA NO")
    axata_etiketi.font = KALIN
    axata_degeri = sayfa.cell(row=ust + 2, column=8, value=plan.axata_ozeti or "")
    axata_degeri.font = Font(bold=True, size=12)

    tarih_etiketi = sayfa.cell(row=ust + 4, column=2, value="Plan Sevk Tarihi ve Günü")
    tarih_etiketi.font = KALIN
    tarih_etiketi.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 4, start_column=2, end_row=ust + 4, end_column=4)
    tarih = sayfa.cell(row=ust + 5, column=2, value=plan.plan_tarihi)
    tarih.number_format = "DD.MM.YYYY dddd"
    tarih.alignment = ORTALI
    tarih.font = KALIN
    sayfa.merge_cells(start_row=ust + 5, start_column=2, end_row=ust + 6, end_column=4)

    # Depo kutusu 5 satırdan uzunsa (kutuda karşılığı olmayan bir depo eklendiyse)
    # tablo aşağı kayar; başlık satırının kutunun üzerine yazması engellenir.
    baslik_satiri = ust + max(7, 2 + len(depo_satirlari))
    for sutun, ad in enumerate(BASLIKLAR, start=1):
        hucre = sayfa.cell(row=baslik_satiri, column=sutun, value=ad)
        hucre.font = KALIN
        hucre.border = KENAR
        hucre.alignment = ORTALI
        hucre.fill = BASLIK_DOLGU

    veri_basi = baslik_satiri + 1
    for sira, satir in enumerate(satirlar):
        y = veri_basi + sira
        degerler = [
            None,
            satir.sehir,
            satir.siparis_no,
            plan.sefer_no,
            satir.depo_kodu,
            satir.urun_kodu,
            satir.gosterilecek_urun_adi,
            float(satir.miktar),
            satir.bayi_gosterimi,
            satir.alici_gosterimi,
            satir.adres_metni,
            satir.teslimat_no,
        ]
        for sutun, deger in enumerate(degerler, start=1):
            hucre = sayfa.cell(row=y, column=sutun, value=deger)
            hucre.border = KENAR
            hucre.font = KUCUK
            hucre.alignment = SOLA if sutun in (7, 9, 10, 11) else ORTALI

    # Form sabit yükseklikte basılır: kalan satırlar numaralanmış boş satır olarak durur.
    bos_basi = veri_basi + len(satirlar)
    veri_sonu = veri_basi + max(len(satirlar), ASGARI_SATIR) - 1
    for sira, y in enumerate(range(bos_basi, veri_sonu + 1), start=len(satirlar) + 1):
        sayfa.cell(row=y, column=1, value=sira).font = KUCUK
        for sutun in range(1, 13):
            sayfa.cell(row=y, column=sutun).border = KENAR

    arac = sayfa.cell(row=veri_basi, column=1, value="RİNG")
    arac.font = KALIN
    arac.alignment = ORTALI
    if len(satirlar) > 1:
        sayfa.merge_cells(
            start_row=veri_basi, start_column=1,
            end_row=veri_basi + len(satirlar) - 1, end_column=1,
        )

    toplam_satiri = veri_sonu + 1
    if plan.yukleme_notu:
        # Planlamacının depoya yazdığı serbest not; tablonun hemen altında durur.
        notu = sayfa.cell(row=toplam_satiri, column=1, value=f"NOT: {plan.yukleme_notu}")
        notu.font = Font(bold=True, size=10, color="2A507C")
        notu.alignment = SOLA
        sayfa.merge_cells(
            start_row=toplam_satiri, start_column=1,
            end_row=toplam_satiri, end_column=3,
        )
    sayfa.cell(row=toplam_satiri, column=4, value="PLANLAYAN").font = KALIN
    sayfa.cell(row=toplam_satiri, column=5, value=plan.olusturan)
    sayfa.cell(row=toplam_satiri, column=7, value="TOPLAM ADET").font = KALIN
    toplam = sayfa.cell(row=toplam_satiri, column=8, value=float(plan.toplam_adet or 0))
    toplam.font = KALIN

    if plan.istisna_asim:
        uyari = sayfa.cell(
            row=toplam_satiri, column=10,
            value="DİKKAT: Tek teslimat üst limiti aşıyor (istisna planı)",
        )
        uyari.font = Font(bold=True, size=9, color="C00000")

    # Navlun faturasının markalar arasında dağıtımı.
    marka_satiri = toplam_satiri + 1
    if plan.marka_paylari:
        etiket = sayfa.cell(row=marka_satiri, column=4, value="FATURA YÜZDESİ")
        etiket.font = KALIN
        for sira, (ad, oran) in enumerate(plan.marka_paylari.items()):
            sayfa.cell(row=marka_satiri + sira, column=5, value=ad).font = KUCUK
            yuzde = sayfa.cell(row=marka_satiri + sira, column=6, value=float(oran))
            yuzde.number_format = "0%"
            yuzde.font = KALIN
        marka_satiri += len(plan.marka_paylari) - 1

    # Depo operasyonu birden fazla Axata numarası verebiliyor; hepsi listelenir.
    if len(plan.axata_numaralari) > 1:
        etiket = sayfa.cell(row=marka_satiri + 1, column=4, value="AXATA NUMARALARI")
        etiket.font = KALIN
        for sira, axata in enumerate(plan.axata_numaralari):
            sayfa.cell(
                row=marka_satiri + 1 + sira,
                column=5,
                value=f"{axata.numara}" + (f" — {axata.aciklama}" if axata.aciklama else ""),
            ).font = KUCUK
        marka_satiri += len(plan.axata_numaralari)

    imza_basi = marka_satiri + 3
    for sira, etiket in enumerate(("Sevk Kontrol", "Adı Soyadı:", "İmzası:")):
        sayfa.cell(row=imza_basi + sira, column=2, value=etiket).font = KALIN
    alt = imza_basi + 2
    # Formun tamamı tek çerçeve içinde görünür.
    cerceve_ciz(sayfa, ust, alt, 1, SUTUN_SAYISI)
    return alt


def _sayfayi_hazirla(sayfa) -> None:
    gridleri_gizle(sayfa)
    sayfa.page_setup.orientation = "landscape"
    sayfa.page_setup.paperSize = 9  # A4
    sayfa.page_setup.fitToWidth = 1
    sayfa.sheet_properties.pageSetUpPr.fitToPage = True
    for kolon, genislik in KOLON_GENISLIKLERI.items():
        sayfa.column_dimensions[kolon].width = genislik


def formlari_uret(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    """Birden çok planı tek çalışma kitabına, her biri ayrı sayfaya basılacak şekilde yazar."""
    if not planlar:
        raise ValueError("Form üretmek için en az bir plan gerekir.")
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("D-RİNG")
    _sayfayi_hazirla(sayfa)

    ust = 1
    for sira, plan in enumerate(planlar):
        son = _blok_yaz(sayfa, plan, ust)
        ust = son + 2
        if sira < len(planlar) - 1:
            sayfa.row_breaks.append(Break(id=ust - 1))

    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef


def form_uret(plan: SevkiyatPlani, hedef: Path | None = None) -> Path:
    hedef = hedef or CIKTI_DIZIN / f"yukleme_formu_{plan.sefer_no}.xlsx"
    return formlari_uret([plan], hedef)


def plan_listesi_disa_aktar(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    """Rapor ekranlarındaki plan listesini Excel'e aktarır."""
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("Planlar")
    basliklar = [
        "Sefer No", "Plan Tarihi", "Durum", "Axata No", "Marka Payı", "Depo",
        "Ürün Grubu / Anahtar",
        "Ürünler", "Teslimat Sayısı", "Ölçü", "Toplam Palet", "Toplam Anahtar",
        "Doluluk %", "Toplam Adet", "Toplam Ağırlık", "Mix", "İstisna",
        "Mail Tarihi", "Oluşturan",
    ]
    satirlar = [
        [
            plan.sefer_no,
            plan.plan_tarihi.strftime("%d.%m.%Y") if plan.plan_tarihi else "",
            plan.durum.value,
            plan.axata_ozeti or "",
            plan.marka_ozeti,
            plan.depo_kodu,
            plan.planlama_anahtari,
            plan.urun_kodlari,
            plan.teslimat_sayisi,
            plan.olcu,
            float(plan.toplam_palet or Decimal(0)),
            round(float(plan.toplam_anahtar or Decimal(0)), 4),
            float(plan.doluluk_yuzdesi or Decimal(0)),
            float(plan.toplam_adet or Decimal(0)),
            float(plan.toplam_agirlik or Decimal(0)),
            "E" if plan.mix_mi else "H",
            "E" if plan.istisna_asim else "H",
            plan.mail_gonderim_tarihi.strftime("%d.%m.%Y %H:%M") if plan.mail_gonderim_tarihi else "",
            plan.olusturan,
        ]
        for plan in planlar
    ]
    sayfa_yaz(sayfa, basliklar, satirlar)
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef
