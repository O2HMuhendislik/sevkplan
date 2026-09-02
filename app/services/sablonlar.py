"""Kullanıcıya verilecek Excel şablonlarını üretir.

Her şablonun ikinci sayfası "Açıklama" olup kolonun ne anlama geldiğini, zorunlu olup
olmadığını ve örnek değeri gösterir.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Alignment, Font

from app.services.excel import sayfa_yaz, yeni_kitap
from app.services.veri_formatlari import (
    Alan,
    MUSTERI_ALANLARI,
    SIPARIS_ALANLARI,
    URUN_ALANLARI,
)


def _sablon_uret(alanlar: tuple[Alan, ...], sayfa_adi: str, baslik: str, hedef: Path) -> Path:
    kitap = yeni_kitap()
    veri = kitap.create_sheet(sayfa_adi)
    sayfa_yaz(
        veri,
        [alan.baslik for alan in alanlar],
        [[alan.ornek for alan in alanlar]],
        [max(16, len(alan.baslik) + 6) for alan in alanlar],
    )

    aciklama = kitap.create_sheet("Açıklama")
    aciklama["A1"] = baslik
    aciklama["A1"].font = Font(bold=True, size=13)
    aciklama.append([])
    sayfa_yaz_satir = [
        [alan.baslik, "Zorunlu" if alan.zorunlu else "Opsiyonel", alan.aciklama,
         ", ".join(alan.aliaslar) or "-"]
        for alan in alanlar
    ]
    aciklama.append(["Kolon", "Durum", "Açıklama", "Kabul edilen diğer başlıklar"])
    for hucre in aciklama[3]:
        hucre.font = Font(bold=True)
    for satir in sayfa_yaz_satir:
        aciklama.append(satir)
    for kolon, genislik in zip("ABCD", (26, 12, 78, 46)):
        aciklama.column_dimensions[kolon].width = genislik
    for satir in aciklama.iter_rows(min_row=3):
        for hucre in satir:
            hucre.alignment = Alignment(vertical="top", wrap_text=True)

    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef


def urun_sablonu(hedef: Path) -> Path:
    return _sablon_uret(
        URUN_ALANLARI,
        "Ürünler",
        "Ürün Master Data Şablonu — kolon başlıklarını değiştirmeyin, satırları doldurun.",
        hedef,
    )


def siparis_sablonu(hedef: Path) -> Path:
    return _sablon_uret(
        SIPARIS_ALANLARI,
        "Siparişler",
        "Sipariş Aktarım Şablonu — her satır bir sipariş kalemidir.",
        hedef,
    )


def musteri_sablonu(hedef: Path) -> Path:
    return _sablon_uret(
        MUSTERI_ALANLARI,
        "Müşteriler",
        "İç Piyasa Müşteri Master Data Şablonu — her satır bir bayi/müşteridir.",
        hedef,
    )
