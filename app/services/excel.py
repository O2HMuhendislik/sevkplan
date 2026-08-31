"""Excel okuma/yazma yardımcıları.

Kolon başlıkları normalize edilerek eşleştirilir: büyük/küçük harf, Türkçe karakter,
boşluk ve alt çizgi farkları tolere edilir. Böylece kaynak sistemden gelen dosyanın
başlıkları birebir aynı yazılmak zorunda kalmaz.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_TR_HARFLER = str.maketrans("ıİşŞğĞüÜöÖçÇ", "iIsSgGuUoOcC")

BASLIK_DOLGU = PatternFill("solid", fgColor="1F4E79")
BASLIK_YAZI = Font(color="FFFFFF", bold=True)


def normalize(metin: Any) -> str:
    if metin is None:
        return ""
    return (
        str(metin).translate(_TR_HARFLER).strip().lower().replace(" ", "_").replace("-", "_")
    )


class ExcelHatasi(Exception):
    pass


def _kolonlari_esle(
    basliklar: list[str], alias_haritasi: dict[str, tuple[str, ...]]
) -> dict[str, int]:
    """Başlık satırını alan adlarına eşler.

    Aynı başlığın iki kez geçtiği dosyalar var (kaynak sistemde iki ayrı 'Not'
    sütunu gibi); ilk eşleşen sütun kullanılır, sonrakiler serbest kalır.
    """
    kolon_indeksi: dict[str, int] = {}
    kullanilan: set[int] = set()
    for alan, aliaslar in alias_haritasi.items():
        adaylar = {normalize(a) for a in aliaslar} | {alan}
        for idx, baslik in enumerate(basliklar):
            if idx not in kullanilan and baslik in adaylar:
                kolon_indeksi[alan] = idx
                kullanilan.add(idx)
                break
    return kolon_indeksi


def sayfa_bul(
    calisma_kitabi,
    alias_haritasi: dict[str, tuple[str, ...]],
    zorunlu: tuple[str, ...],
    sayfa_adi: str | None = None,
    taranacak_satir: int = 10,
):
    """Veri sayfasını ve başlık satırını otomatik bulur.

    Kaynak dosyalarda aranan tablo çoğu zaman ilk sayfada ve ilk satırda değildir.
    Bütün sayfaların ilk birkaç satırı taranır, zorunlu alanların en çoğunu
    karşılayan satır başlık kabul edilir.
    """
    en_iyi = None
    for sayfa in calisma_kitabi.worksheets:
        if sayfa_adi and normalize(sayfa.title) != normalize(sayfa_adi):
            continue
        for satir_no, satir in enumerate(sayfa.iter_rows(values_only=True), start=1):
            if satir_no > taranacak_satir:
                break
            basliklar = [normalize(h) for h in satir]
            kolonlar = _kolonlari_esle(basliklar, alias_haritasi)
            puan = sum(1 for alan in zorunlu if alan in kolonlar)
            if en_iyi is None or puan > en_iyi[0]:
                en_iyi = (puan, sayfa.title, satir_no, kolonlar)
            if puan == len(zorunlu):
                return sayfa.title, satir_no, kolonlar
    if en_iyi is None:
        raise ExcelHatasi("Dosyada okunabilir sayfa bulunamadı.")
    return en_iyi[1], en_iyi[2], en_iyi[3]


def satirlari_oku(
    dosya: Path | Any,
    alias_haritasi: dict[str, tuple[str, ...]],
    zorunlu: tuple[str, ...] = (),
    sayfa_adi: str | None = None,
) -> list[dict[str, Any]]:
    """Veri sayfasını bulup her satırı normalize alan adlarıyla sözlüğe çevirir."""
    calisma_kitabi = load_workbook(dosya, data_only=True, read_only=True)
    try:
        bulunan_sayfa, baslik_satiri, kolon_indeksi = sayfa_bul(
            calisma_kitabi, alias_haritasi, zorunlu, sayfa_adi
        )
        sayfa = calisma_kitabi[bulunan_sayfa]
        kayitlar: list[dict[str, Any]] = []
        for satir_no, satir in enumerate(sayfa.iter_rows(values_only=True), start=1):
            if satir_no <= baslik_satiri:
                continue
            if satir is None or all(h is None or str(h).strip() == "" for h in satir):
                continue
            kayit: dict[str, Any] = {"_satir_no": satir_no, "_sayfa": bulunan_sayfa}
            for alan, idx in kolon_indeksi.items():
                kayit[alan] = satir[idx] if idx < len(satir) else None
            kayitlar.append(kayit)
        return kayitlar
    finally:
        calisma_kitabi.close()


def eksik_kolonlar(
    dosya: Path | Any,
    alias_haritasi: dict[str, tuple[str, ...]],
    zorunlu: tuple[str, ...],
    sayfa_adi: str | None = None,
) -> list[str]:
    calisma_kitabi = load_workbook(dosya, data_only=True, read_only=True)
    try:
        _, _, kolonlar = sayfa_bul(calisma_kitabi, alias_haritasi, zorunlu, sayfa_adi)
        return [alan for alan in zorunlu if alan not in kolonlar]
    finally:
        calisma_kitabi.close()


GECERSIZ_DEGERLER = {"#n/a", "#yok", "#value!", "#ref!", "na", "n/a", "-", "header"}


def bos_mu(deger: Any) -> bool:
    """Kaynak dosyalarda formül hatası olarak gelen değerleri boş sayar."""
    return deger is None or str(deger).strip().lower() in GECERSIZ_DEGERLER | {""}


def metin(deger: Any) -> str | None:
    if bos_mu(deger):
        return None
    sonuc = str(deger).strip()
    if sonuc.endswith(".0") and sonuc[:-2].isdigit():
        sonuc = sonuc[:-2]  # Excel sayısal hücreleri metne çevirirken oluşan .0 kuyruğu
    return sonuc or None


def sayi_ya_da(deger: Any) -> Decimal | None:
    """Sayıya çevrilebiliyorsa Decimal, çevrilemiyorsa None döner (hata atmaz)."""
    if bos_mu(deger):
        return None
    try:
        return Decimal(str(deger).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def tam_sayi_ya_da(deger: Any) -> int | None:
    sonuc = sayi_ya_da(deger)
    return int(sonuc) if sonuc is not None else None


def sayi(deger: Any, alan: str) -> Decimal:
    if bos_mu(deger):
        raise ExcelHatasi(f"{alan} boş olamaz")
    try:
        return Decimal(str(deger).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        raise ExcelHatasi(f"{alan} sayı olmalı: {deger!r}") from None


def tam_sayi(deger: Any, alan: str) -> int:
    return int(sayi(deger, alan))


def tarih(deger: Any) -> date | None:
    if bos_mu(deger):
        return None
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    ham = str(deger).strip()
    for desen in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y", "%Y%m%d"):
        try:
            return datetime.strptime(ham, desen).date()
        except ValueError:
            continue
    raise ExcelHatasi(f"Tarih çözümlenemedi: {deger!r} (beklenen: GG.AA.YYYY)")


def evet_hayir(deger: Any, varsayilan: bool = False) -> bool:
    if bos_mu(deger):
        return varsayilan
    return normalize(deger) in {"e", "evet", "h_evet", "1", "true", "x", "var", "yes"}


def sayfa_yaz(
    sayfa, basliklar: list[str], satirlar: list[list[Any]], genislikler: list[int] | None = None
) -> None:
    sayfa.append(basliklar)
    for hucre in sayfa[1]:
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(horizontal="center", vertical="center")
    for satir in satirlar:
        sayfa.append(satir)
    for idx, baslik in enumerate(basliklar, start=1):
        genislik = genislikler[idx - 1] if genislikler else max(14, len(str(baslik)) + 4)
        sayfa.column_dimensions[get_column_letter(idx)].width = genislik
    sayfa.freeze_panes = "A2"


def yeni_kitap() -> Workbook:
    kitap = Workbook()
    kitap.remove(kitap.active)
    return kitap
