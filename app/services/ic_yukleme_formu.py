"""İç piyasa yükleme formu.

Düzen, sahadaki "… Tarihli N. Bölge Yükleme Formları" çalışma kitabının karşılığıdır.
Ring formundan iki farkı var:

* Sevkiyat tipleri **ayrı sayfalara** yazılır (`S-FTL Sevk`, `R-Rutin`, `K-KARGO`).
* Satır tablosunun altında araç, rota ve fatura bilgilerini taşıyan bir **alt blok**
  bulunur: toplam parça, araç tipi, il + durak sayısı (`İZMİR2YER`), '+' ile birleşik
  ilçeler, yer miktarı, plaka/şoför, nakliyeci, marka bazlı fatura yüzdesi ve yükleme
  yapacak depoların kalem/adet dökümü.

Ortak yüklemede (64 + 74 + -1) malı başka depoda olan satırın karşısına
"… depoya gönderilmelidir" notu yazılır; bu aktarma için ayrı plan üretilmez.
"""
from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl.styles import Font
from openpyxl.worksheet.pagebreak import Break

from app.config import CIKTI_DIZIN
from app.domain.ic_piyasa import SevkiyatTipi
from app.models import SevkiyatPlani
from app.services.excel import sayfa_yaz, yeni_kitap
from app.services.yukleme_formu import (
    ASGARI_SATIR,
    BASLIK_DOLGU,
    FORM_NO,
    KALIN,
    KENAR,
    KUCUK,
    ORTALI,
    SOLA,
    UYARI_METNI,
    _depo_satirlari,
)

SAYFA_ADLARI = {
    SevkiyatTipi.FTL: "S-FTL Sevk",
    SevkiyatTipi.RUTIN: "R-Rutin",
    SevkiyatTipi.KARGO: "K-KARGO",
}

ARAC_ETIKETLERI = {
    SevkiyatTipi.FTL: "TIR",
    SevkiyatTipi.RUTIN: "FTL RUTİN",
    SevkiyatTipi.KARGO: "KARGO",
}

BASLIKLAR = [
    "No", "İl Adi", "Sipariş No", "Belge No", "Depo ", "Ürün Kodu", "Ürün Adi",
    "Adet", "Bayii Adı", "Alıcı Firma", "Sevk Adresi", "İlçe", "Teslimat",
    "Axata", "Not",
]
KOLON_GENISLIKLERI = {
    "A": 9.7, "B": 14.7, "C": 15.7, "D": 14.7, "E": 9.0, "F": 14.3,
    "G": 40.0, "H": 7.0, "I": 32.0, "J": 30.0, "K": 34.0, "L": 14.0,
    "M": 15.0, "N": 12.0, "O": 28.0,
}
SUTUN_SAYISI = len(BASLIKLAR)


def _plan_tipi(plan: SevkiyatPlani) -> SevkiyatTipi:
    try:
        return SevkiyatTipi(plan.sevkiyat_tipi or "")
    except ValueError:
        return SevkiyatTipi.FTL


def _depo_dokumu(plan: SevkiyatPlani) -> list[tuple[str, int, Decimal]]:
    """Formun "Yükleme yapacak depolar" bloğu: depo + marka, kalem sayısı, adet.

    Kalem = o depodan çıkacak ürün kodu sayısı; depo operasyonu toplama listesini
    buna göre böler.
    """
    from app.domain.marka import marka

    kalemler: dict[str, set[str]] = defaultdict(set)
    adetler: dict[str, Decimal] = defaultdict(Decimal)
    for satir in plan.satirlar:
        etiket = f"{satir.depo_kodu} {marka(satir.depo_kodu)}"
        kalemler[etiket].add(satir.urun_kodu)
        adetler[etiket] += Decimal(satir.miktar)
    return [
        (etiket, len(kalemler[etiket]), adetler[etiket])
        for etiket in sorted(adetler, key=lambda e: -adetler[e])
    ]


def _ust_blok(sayfa, plan: SevkiyatPlani, ust: int) -> int:
    """Form başlığı, sefer no, depo/AXATA kutusu ve sevk tarihi. Tablo başlığını döner."""
    sayfa.cell(row=ust, column=12, value="FORM NO : ").font = KUCUK
    sayfa.cell(row=ust, column=13, value=FORM_NO).font = KUCUK

    baslik = sayfa.cell(row=ust + 1, column=2, value="YÜKLEME FORMLARI")
    baslik.font = Font(bold=True, size=12)
    baslik.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 1, start_column=2, end_row=ust + 1, end_column=4)
    sayfa.cell(row=ust + 1, column=5, value="Depo ").font = KALIN
    sayfa.cell(row=ust + 1, column=6, value="AXATA").font = KALIN

    uyari = sayfa.cell(row=ust + 1, column=7, value=UYARI_METNI)
    uyari.font = Font(bold=True, size=9, color="C00000")
    uyari.alignment = ORTALI
    sayfa.merge_cells(
        start_row=ust + 1, start_column=7, end_row=ust + 1, end_column=SUTUN_SAYISI
    )

    etiket = sayfa.cell(row=ust + 2, column=2, value="SEFER NO")
    etiket.font = KALIN
    etiket.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 2, start_column=2, end_row=ust + 3, end_column=2)
    sefer = sayfa.cell(row=ust + 2, column=3, value=plan.sefer_no)
    sefer.font = Font(bold=True, size=12)
    sefer.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 2, start_column=3, end_row=ust + 3, end_column=4)

    depo_satirlari, hedef_etiket = _depo_satirlari(plan.yukleme_deposu or plan.depo_kodu)
    for sira, depo_adi in enumerate(depo_satirlari):
        satir = ust + 2 + sira
        hucre = sayfa.cell(row=satir, column=5, value=depo_adi)
        hucre.font = KALIN
        hucre.border = KENAR
        hucre.alignment = ORTALI
        axata = sayfa.cell(row=satir, column=6)
        axata.border = KENAR
        axata.alignment = ORTALI
        if depo_adi == hedef_etiket:
            axata.value = plan.axata_ozeti or ""
            axata.font = KALIN

    sayfa.cell(row=ust + 2, column=7, value="AXATA NO").font = KALIN
    sayfa.cell(row=ust + 2, column=8, value=plan.axata_ozeti or "").font = Font(
        bold=True, size=12
    )
    sayfa.cell(row=ust + 3, column=7, value="SEVKİYAT TİPİ").font = KALIN
    sayfa.cell(row=ust + 3, column=8, value=plan.sevkiyat_tipi_adi).font = KALIN
    sayfa.cell(row=ust + 4, column=7, value="BÖLGE").font = KALIN
    sayfa.cell(row=ust + 4, column=8, value=plan.bolge_adi).font = KALIN

    tarih_etiketi = sayfa.cell(row=ust + 4, column=2, value="Plan Sevk Tarihi ve Günü")
    tarih_etiketi.font = KALIN
    tarih_etiketi.alignment = ORTALI
    sayfa.merge_cells(start_row=ust + 4, start_column=2, end_row=ust + 4, end_column=4)
    tarih = sayfa.cell(row=ust + 5, column=2, value=plan.plan_tarihi)
    tarih.number_format = "DD.MM.YYYY dddd"
    tarih.alignment = ORTALI
    tarih.font = KALIN
    sayfa.merge_cells(start_row=ust + 5, start_column=2, end_row=ust + 6, end_column=4)

    return ust + max(7, 2 + len(depo_satirlari))


def _satir_tablosu(sayfa, plan: SevkiyatPlani, baslik_satiri: int) -> int:
    """Sipariş satırlarını yazar; tablonun son satırını döner."""
    from app.domain.iller import mesafe

    for sutun, ad in enumerate(BASLIKLAR, start=1):
        hucre = sayfa.cell(row=baslik_satiri, column=sutun, value=ad)
        hucre.font = KALIN
        hucre.border = KENAR
        hucre.alignment = ORTALI
        hucre.fill = BASLIK_DOLGU

    # Satırlar rota sırasına göre dizilir: yakın duraktan uzağa, en uzak il en altta.
    def sira_anahtari(satir):
        uzaklik = mesafe(satir.sehir)
        return (
            uzaklik if uzaklik is not None else 9999,
            satir.sehir or "",
            satir.bayi_adi or "",
            satir.teslimat_no,
            satir.urun_kodu,
        )

    satirlar = sorted(plan.satirlar, key=sira_anahtari)
    veri_basi = baslik_satiri + 1
    for sira, satir in enumerate(satirlar):
        y = veri_basi + sira
        degerler = [
            None,
            satir.sehir,
            satir.siparis_no,
            plan.sefer_no,
            satir.depo_kodu,
            satir.urun_kodu,
            satir.gosterilecek_urun_adi,
            float(satir.miktar),
            satir.bayi_adi,
            satir.alici_firma,
            satir.sevk_adresi,
            satir.ilce,
            satir.teslimat_no,
            plan.axata_ozeti or "",
            plan.aktarma_notu(satir),
        ]
        for sutun, deger in enumerate(degerler, start=1):
            hucre = sayfa.cell(row=y, column=sutun, value=deger)
            hucre.border = KENAR
            hucre.font = KUCUK
            hucre.alignment = SOLA if sutun in (7, 9, 10, 11, 15) else ORTALI
        if degerler[-1]:
            sayfa.cell(row=y, column=SUTUN_SAYISI).font = Font(
                bold=True, size=9, color="C00000"
            )

    bos_basi = veri_basi + len(satirlar)
    veri_sonu = veri_basi + max(len(satirlar), ASGARI_SATIR) - 1
    for sira, y in enumerate(range(bos_basi, veri_sonu + 1), start=len(satirlar) + 1):
        sayfa.cell(row=y, column=1, value=sira).font = KUCUK
        for sutun in range(1, SUTUN_SAYISI + 1):
            sayfa.cell(row=y, column=sutun).border = KENAR

    arac = sayfa.cell(row=veri_basi, column=1, value=ARAC_ETIKETLERI[_plan_tipi(plan)])
    arac.font = KALIN
    arac.alignment = ORTALI
    if len(satirlar) > 1:
        sayfa.merge_cells(
            start_row=veri_basi, start_column=1,
            end_row=veri_basi + len(satirlar) - 1, end_column=1,
        )
    return veri_sonu


def _alt_blok(sayfa, plan: SevkiyatPlani, ilk_satir: int) -> int:
    """Araç, rota ve fatura bilgileri. Ring formunda olmayan kısım."""
    y = ilk_satir
    sayfa.cell(row=y, column=6, value="Toplam parça sayısı").font = KALIN
    sayfa.cell(row=y, column=8, value=float(plan.toplam_adet or 0)).font = KALIN

    y += 1
    sol = sayfa.cell(
        row=y, column=2, value="Araç, Yer, Sürücü ve Yükleme Yapanın Bilgileri"
    )
    sol.font = KALIN
    sol.alignment = ORTALI
    sayfa.merge_cells(start_row=y, start_column=2, end_row=y, end_column=5)
    sag = sayfa.cell(row=y, column=6, value="Sevkiyat Palet Detayı ve Ürün Bilgileri")
    sag.font = KALIN
    sag.alignment = ORTALI
    sayfa.merge_cells(start_row=y, start_column=6, end_row=y, end_column=9)

    # Sol sütun: araç ve rota bilgileri.
    alanlar = [
        ("Araç Tipi", ARAC_ETIKETLERI[_plan_tipi(plan)]),
        ("İl", plan.il_yeri_metni),
        ("İlçe", plan.ilce_metni),
        ("Yer Miktarı", plan.durak_sayisi),
        ("Plaka", plan.plaka or ""),
        ("Şöför Adı", plan.surucu or ""),
        ("Telefon", plan.surucu_telefon or ""),
        ("Nak.Firma", plan.nakliyeci or ""),
    ]
    for sira, (etiket, deger) in enumerate(alanlar):
        satir = y + 1 + sira
        hucre = sayfa.cell(row=satir, column=2, value=etiket)
        hucre.font = KALIN
        hucre.border = KENAR
        deger_hucresi = sayfa.cell(row=satir, column=3, value=deger)
        deger_hucresi.border = KENAR
        deger_hucresi.alignment = SOLA
        sayfa.merge_cells(
            start_row=satir, start_column=3, end_row=satir, end_column=5
        )

    # Sağ sütun: fatura yüzdesi ve yükleme yapacak depolar.
    sag_satiri = y + 1
    if plan.marka_paylari:
        sayfa.cell(row=sag_satiri, column=6, value="Faturalama").font = KALIN
        for ad, oran in plan.marka_paylari.items():
            sayfa.cell(row=sag_satiri, column=7, value=ad).font = KUCUK
            yuzde = sayfa.cell(row=sag_satiri, column=8, value=float(oran))
            yuzde.number_format = "0%"
            yuzde.font = KALIN
            sag_satiri += 1
        sag_satiri += 1

    sayfa.cell(row=sag_satiri, column=6, value="Yükleme yapacak depolar").font = KALIN
    sayfa.cell(row=sag_satiri, column=8, value="Kalem").font = KALIN
    sayfa.cell(row=sag_satiri, column=9, value="Adet").font = KALIN
    for etiket, kalem, adet in _depo_dokumu(plan):
        sag_satiri += 1
        sayfa.cell(row=sag_satiri, column=6, value=etiket).font = KUCUK
        sayfa.cell(row=sag_satiri, column=8, value=kalem).font = KUCUK
        sayfa.cell(row=sag_satiri, column=9, value=float(adet)).font = KUCUK

    if len(plan.axata_numaralari) > 1:
        sag_satiri += 2
        sayfa.cell(row=sag_satiri, column=6, value="AXATA NUMARALARI").font = KALIN
        for axata in plan.axata_numaralari:
            sag_satiri += 1
            sayfa.cell(
                row=sag_satiri,
                column=7,
                value=axata.numara
                + (f" — {axata.aciklama}" if axata.aciklama else ""),
            ).font = KUCUK

    alt = max(y + len(alanlar), sag_satiri) + 2
    sayfa.cell(row=alt, column=2, value="Planlayan").font = KALIN
    sayfa.cell(row=alt, column=3, value=plan.olusturan)
    for sira, etiket in enumerate(("Sevk Kontrol", "Adı Soyadı:", "İmzası:")):
        sayfa.cell(row=alt + 1 + sira, column=2, value=etiket).font = KALIN
    return alt + 3


def _blok_yaz(sayfa, plan: SevkiyatPlani, ust: int) -> int:
    baslik_satiri = _ust_blok(sayfa, plan, ust)
    veri_sonu = _satir_tablosu(sayfa, plan, baslik_satiri)
    return _alt_blok(sayfa, plan, veri_sonu + 1)


def _sayfayi_hazirla(sayfa) -> None:
    sayfa.page_setup.orientation = "landscape"
    sayfa.page_setup.paperSize = 9  # A4
    sayfa.page_setup.fitToWidth = 1
    sayfa.sheet_properties.pageSetUpPr.fitToPage = True
    for kolon, genislik in KOLON_GENISLIKLERI.items():
        sayfa.column_dimensions[kolon].width = genislik


def formlari_uret(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    """Planları sevkiyat tipine göre ayrı sayfalara yazar.

    Aynı sayfadaki planlar alt alta gelir ve her biri ayrı sayfaya basılacak şekilde
    sayfa sonu konur — sahadaki günlük bölge formatı budur.
    """
    if not planlar:
        raise ValueError("Form üretmek için en az bir plan gerekir.")
    kitap = yeni_kitap()

    gruplar: dict[SevkiyatTipi, list[SevkiyatPlani]] = defaultdict(list)
    for plan in planlar:
        gruplar[_plan_tipi(plan)].append(plan)

    for tip in SevkiyatTipi:
        grup = gruplar.get(tip)
        if not grup:
            continue
        sayfa = kitap.create_sheet(SAYFA_ADLARI[tip])
        _sayfayi_hazirla(sayfa)
        ust = 1
        for sira, plan in enumerate(sorted(grup, key=lambda p: p.sefer_no)):
            son = _blok_yaz(sayfa, plan, ust)
            ust = son + 2
            if sira < len(grup) - 1:
                sayfa.row_breaks.append(Break(id=ust - 1))

    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef


def form_uret(plan: SevkiyatPlani, hedef: Path | None = None) -> Path:
    hedef = hedef or CIKTI_DIZIN / f"ic_yukleme_formu_{plan.sefer_no}.xlsx"
    return formlari_uret([plan], hedef)


def gunluk_form(planlar: list[SevkiyatPlani], hedef: Path | None = None) -> Path:
    """Bir günün bütün iç piyasa planlarını tek kitapta toplar."""
    if not planlar:
        raise ValueError("Form üretmek için en az bir plan gerekir.")
    tarih = planlar[0].plan_tarihi
    hedef = hedef or CIKTI_DIZIN / f"ic_yukleme_formlari_{tarih:%Y%m%d}.xlsx"
    return formlari_uret(planlar, hedef)


def plan_listesi_disa_aktar(planlar: list[SevkiyatPlani], hedef: Path) -> Path:
    """İç piyasa plan listesini Excel'e aktarır."""
    kitap = yeni_kitap()
    sayfa = kitap.create_sheet("İç Piyasa Planları")
    basliklar = [
        "Sefer No", "Plan Tarihi", "Tip", "Durum", "Axata No", "Bölge",
        "İller", "İlçeler", "Son Uğrak", "Son Uğrak %", "Durak", "Müşteri",
        "Yükleme Deposu", "Depolar", "Doluluk %", "Anahtar", "Palet", "Adet",
        "Desi", "Ağırlık", "Marka Payı", "Nakliyeci", "Plaka", "Oluşturan",
    ]
    satirlar = [
        [
            plan.sefer_no,
            plan.plan_tarihi.strftime("%d.%m.%Y") if plan.plan_tarihi else "",
            plan.sevkiyat_tipi_adi,
            plan.durum.value,
            plan.axata_ozeti or "",
            plan.bolge_adi,
            plan.iller_metni or "",
            plan.ilce_metni,
            plan.son_ugrak or "",
            round(float(plan.son_ugrak_orani or 0) * 100, 1),
            plan.durak_sayisi,
            plan.musteri_sayisi,
            plan.yukleme_deposu or "",
            ", ".join(sorted({s.depo_kodu for s in plan.satirlar})),
            float(plan.doluluk_yuzdesi or 0),
            round(float(plan.toplam_anahtar or 0), 4),
            float(plan.toplam_palet or 0),
            float(plan.toplam_adet or 0),
            float(plan.toplam_desi or 0),
            float(plan.toplam_agirlik or 0),
            plan.marka_ozeti,
            plan.nakliyeci or "",
            plan.plaka or "",
            plan.olusturan,
        ]
        for plan in planlar
    ]
    sayfa_yaz(sayfa, basliklar, satirlar)
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef
