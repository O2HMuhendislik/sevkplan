"""Sahanın cari kod ve teslimat tipi listelerini müşteri master datasına işler.

Kaynak, iki sayfalı bir çalışma kitabıdır:

* **CARİ KODLAR** — "Teslim Yapılacak Müşteri Adı" + "Müşteri Kodu". Bayi kodu
  bugüne kadar sistemde yoktu; eşleştirme hep adla yapılıyordu.
* **TESLİMAT TİPİ** — "BAYİ ADI" + "SEVK TİPİ" + "ÖZEL DURUM". Sevk tipi metni araç
  tipini, cumartesi mal kabulünü ve e-irsaliye gereğini birlikte taşır.

Sayfa adları esnektir; başlıklardan tanınır. Eşleşmeyen adlar sessizce atılmaz,
gerekçesiyle raporlanır (bkz. musteri_eslestirme).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Musteri
from app.services import excel
from app.services.musteri_eslestirme import MusteriEslestirici

TIR_GIREN = {"TIR"}
TIR_GIREMEZ = {"KAMYON", "KAMYONET", "RUTIN", "SADECE KAMYON"}
"""Sevk tipinin araç kısmı. Kamyon/kamyonet/rutin yazan adrese tır giremiyor."""

CUMARTESI_YOK = ("C.TESI YOK", "CTESI YOK", "CUMARTESI YOK")
E_IRSALIYE = ("EIRSALIYE", "E IRSALIYE", "E-IRSALIYE")


@dataclass
class EkBilgiSonucu:
    cari_okunan: int = 0
    cari_yazilan: int = 0
    tip_okunan: int = 0
    tip_yazilan: int = 0
    tir_degisen: int = 0
    """Tır girişi bilgisi değişen müşteri sayısı."""
    eslesmeyenler: list[dict] = field(default_factory=list)
    """Eşleşmeyen ve belirsiz kalan adlar; inceleme dosyasına yazılır."""

    def ozet(self) -> str:
        parcalar = []
        if self.cari_okunan:
            parcalar.append(
                f"cari kod: {self.cari_okunan} okundu, {self.cari_yazilan} işlendi"
            )
        if self.tip_okunan:
            parcalar.append(
                f"sevk tipi: {self.tip_okunan} okundu, {self.tip_yazilan} işlendi "
                f"({self.tir_degisen} müşteride tır girişi değişti)"
            )
        if self.eslesmeyenler:
            parcalar.append(f"{len(self.eslesmeyenler)} ad eşleşmedi")
        return " · ".join(parcalar) or "İşlenecek kayıt bulunamadı."


def _normalize(deger: Any) -> str:
    return (str(deger) if deger is not None else "").strip()


def sevk_tipini_coz(metin: str) -> dict:
    """Sevk tipi metnini alanlara ayırır.

    "TIR-C.TESİ YOK-EİRSALİYE" -> tır girer, cumartesi yok, e-irsaliye var.
    Araç kısmı ilk parçadır; tanınmayan değerde tır girişi '?' kalır — sahanın
    yazdığı metin yine saklanır ve ekranda görünür.
    """
    ham = _normalize(metin)
    buyuk = excel.normalize(ham).upper().replace("_", " ")
    arac = buyuk.split("-")[0].strip()
    # "TIRE-..." gibi yazım hataları: baştaki TIR yakalanır.
    if arac.startswith("TIR"):
        tir = "E"
    elif arac in TIR_GIREMEZ or arac.startswith("KAMYON") or arac.startswith("RUTIN"):
        tir = "H"
    elif "TIR" in buyuk and "ZORDA" in buyuk:
        # "ZORDA KALIRSAN TIR": tır girebiliyor ama tercih edilmiyor.
        tir = "E"
    else:
        tir = "?"
    return {
        "sevk_tipi": ham,
        "tir_girisi": tir,
        "cumartesi_teslimat": not any(k in buyuk for k in CUMARTESI_YOK),
        "e_irsaliye": any(k in buyuk for k in E_IRSALIYE),
    }


def _sayfa_bul(kitap, *anahtarlar: str):
    """Başlık satırında verilen kelimeleri taşıyan sayfayı döner."""
    for sayfa in kitap.worksheets:
        for satir in sayfa.iter_rows(min_row=1, max_row=3, values_only=True):
            basliklar = " ".join(excel.normalize(h) for h in satir if h)
            if all(a in basliklar for a in anahtarlar):
                return sayfa
    return None


def _satirlar(sayfa) -> list[tuple]:
    """Başlık satırını atlayıp dolu satırları döner."""
    kayitlar = []
    baslik_gecildi = False
    for satir in sayfa.iter_rows(values_only=True):
        if not baslik_gecildi:
            baslik_gecildi = True
            continue
        if satir and any(h is not None and str(h).strip() for h in satir):
            kayitlar.append(satir)
    return kayitlar


GECERSIZ_KOD = {"-", "--", "YOK", "0"}
"""Cari kod sütununda kod yerine konan yer tutucular."""


def ek_bilgileri_aktar(db: Session, dosya: Path | Any) -> EkBilgiSonucu:
    """Cari kod ve sevk tipi sayfalarını müşteri kayıtlarına işler.

    İki geçişlidir. Önce bütün satırlar eşleştirilir, sonra **aynı müşteriye farklı
    değer yazmak isteyen satırlar ayıklanır**. Dosyada aynı bayi birden çok satırda
    farklı yazılabiliyor ("AS MÜHENDİSLİK" bir satırda KAMYON, diğerinde TIR) ve
    ikisi de aynı kayda düşüyor; son satırın kazanması sessizce yanlış araç kararı
    üretirdi. Çakışanlar yazılmaz, gerekçesiyle raporlanır.
    """
    kitap = load_workbook(dosya, data_only=True)
    try:
        sonuc = EkBilgiSonucu()
        eslestirici = MusteriEslestirici(db)

        cari = _sayfa_bul(kitap, "musteri", "kod")
        if cari is not None:
            istekler = []
            for satir in _satirlar(cari):
                ad = _normalize(satir[0])
                kod = _normalize(satir[1] if len(satir) > 1 else "")
                if not ad or not kod or kod.upper() in GECERSIZ_KOD:
                    continue
                sonuc.cari_okunan += 1
                istekler.append((ad, kod, {"bayi_kodu": kod}))
            sonuc.cari_yazilan = _uygula(db, eslestirici, istekler, sonuc, "cari kod")

        tip = _sayfa_bul(kitap, "sevk", "tip")
        if tip is not None:
            istekler = []
            for satir in _satirlar(tip):
                ad = _normalize(satir[0])
                sevk = _normalize(satir[1] if len(satir) > 1 else "")
                ozel = _normalize(satir[2] if len(satir) > 2 else "")
                if not ad or not sevk:
                    continue
                sonuc.tip_okunan += 1
                alanlar = sevk_tipini_coz(sevk)
                if ozel:
                    alanlar["ozel_durum"] = ozel
                istekler.append((ad, sevk, alanlar))
            sonuc.tip_yazilan = _uygula(db, eslestirici, istekler, sonuc, "sevk tipi")
        db.flush()
        return sonuc
    finally:
        kitap.close()


def _uygula(
    db: Session,
    eslestirici: MusteriEslestirici,
    istekler: list[tuple[str, str, dict]],
    sonuc: EkBilgiSonucu,
    kaynak: str,
) -> int:
    """Eşleşen ve çakışmayan istekleri yazar; yazılan müşteri sayısını döner."""
    hedefler: dict[int, list[tuple[str, str, dict]]] = {}
    kayitlar: dict[int, Musteri] = {}
    for ad, deger, alanlar in istekler:
        musteri, adaylar = eslestirici.esle(ad)
        if musteri is None:
            sonuc.eslesmeyenler.append(
                {
                    "kaynak": kaynak,
                    "ad": ad,
                    "deger": deger,
                    "sebep": "birden fazla aday" if adaylar else "eşleşme yok",
                    "adaylar": ", ".join(a.bayi_adi for a in adaylar[:4]) or "",
                }
            )
            continue
        kayitlar[musteri.id] = musteri
        hedefler.setdefault(musteri.id, []).append((ad, deger, alanlar))

    yazilan = 0
    for musteri_id, grup in hedefler.items():
        musteri = kayitlar[musteri_id]
        if len({deger for _, deger, _ in grup}) > 1:
            # Aynı kayda farklı değer isteyen satırlar: hangisi doğru belli değil.
            for ad, deger, _ in grup:
                sonuc.eslesmeyenler.append(
                    {
                        "kaynak": kaynak,
                        "ad": ad,
                        "deger": deger,
                        "sebep": "çakışma — aynı müşteriye farklı değer",
                        "adaylar": musteri.bayi_adi,
                    }
                )
            continue
        _, _, alanlar = grup[0]
        if "tir_girisi" in alanlar and musteri.tir_girisi != alanlar["tir_girisi"]:
            sonuc.tir_degisen += 1
        for alan, deger in alanlar.items():
            setattr(musteri, alan, deger)
        yazilan += 1
    return yazilan


def eslesmeyen_raporu(sonuc: EkBilgiSonucu, hedef: Path) -> Path:
    """Eşleşmeyen adları inceleme dosyasına yazar.

    Kullanıcı doğru bayiyi bulup Müşteriler ekranından işleyebilsin diye adayları
    da yazar; tahmine dayalı otomatik eşleştirme yapılmadığı için bu liste işin
    kalan kısmıdır.
    """
    kitap = excel.yeni_kitap()
    sayfa = kitap.create_sheet("Eşleşmeyenler")
    excel.sayfa_yaz(
        sayfa,
        ["Kaynak", "Dosyadaki ad", "Değer", "Sebep", "Sistemdeki aday kayıtlar"],
        [
            [k["kaynak"], k["ad"], k["deger"], k["sebep"], k["adaylar"]]
            for k in sonuc.eslesmeyenler
        ],
        genislikler=[14, 46, 32, 20, 60],
    )
    hedef.parent.mkdir(parents=True, exist_ok=True)
    kitap.save(hedef)
    return hedef
